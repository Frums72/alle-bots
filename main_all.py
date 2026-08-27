# v60 – Rückumstellung von API-Football auf livescore-api.com
# ============================================================
# Änderungen v60:
#   • Komplette Rückumstellung von API-Football (v3.football.api-sports.io) auf
#     livescore-api.com – gleiches internes Datenformat wie vorher, damit die
#     komplette Bot-Logik (Ecken/Torwart/Druck/Comeback/etc.) unverändert bleibt.
#     Nur die ca. 15 Funktionen, die direkt mit der API sprechen, wurden umgebaut.
#   • Auth läuft jetzt über 2 Railway-Variablen: LIVESCORE_API_KEY + LIVESCORE_API_SECRET
#     (statt vorher 1 Variable API_FOOTBALL_KEY mit Header-Auth)
#   • QUOTEN DEAKTIVIERT: livescore-api.com liefert nur eine einfache 1/X/2-Siegquote,
#     keine Über/Unter-Linien und keine Ecken-Märkte. Alle Quoten-Funktionen geben
#     jetzt bewusst "keine Quote" zurück (siehe Kommentar dort), Value-Bot und
#     Arbitrage-Bot sind deshalb aus der Bot-Liste rausgenommen (pausiert, nicht
#     gelöscht) – sie könnten ohne echte Quoten sowieso nie mehr etwas finden.
#   • Prematch-Fixtures (fixtures/list.json) kommen bei livescore-api.com in UTC und
#     max. 30 pro Seite – Zeiten werden jetzt wieder auf deutsche Zeit umgerechnet
#     (+2h, wie der Rest vom Code das auch macht) und über mehrere Seiten geblättert.
#
# ============================================================
# v59 – Reaktions-Auswertung, Duplikat-Fix, Tagesbericht-Fix
# ============================================================
# Änderungen v59 (auf Wunsch, drei Fixes):
#   1) NEU: Discord-Bot postet nach der Auswertung eine ✅/❌ Reaktion direkt
#      auf die ursprüngliche Signal-Nachricht (Ecken/Torwart/Druck/Comeback/
#      Torflut/HZ1-Tore/VZ-Tore). Dafür wird ein echter Discord-Bot-Token
#      benötigt (DISCORD_BOT_TOKEN) – Webhooks allein können keine Reaktionen
#      setzen. send_discord_embed() gibt bei wait_for_message=True jetzt die
#      message_id/channel_id zurück, die im Signal-Tracker mitgespeichert wird.
#   2) FIX Duplikate: notified_sets_laden() hat die gespeicherten "bereits
#      gepostet"-Listen bisher verworfen, wenn sie älter als 30 Minuten waren
#      (auch am selben Tag). Nach jedem Neustart (Redeploy/Crash) hat der Bot
#      dadurch alle bereits geposteten Spiele vergessen und teils erneut
#      getippt -> exakt das "gleiches Spiel, gleicher Tipp"-Problem. Der
#      30-Minuten-Frische-Check ist raus, es zählt nur noch der Tages-Check.
#   3) FIX Tagesbericht: bot_auswertung_und_berichte nutzte eine lokale
#      Variable statt des persistierten globalen Markers -> nach einem
#      Neustart rund um 02:00 Uhr fiel der Bericht für den ganzen Tag aus.
#      Jetzt Nachhol-Logik (now.hour >= TAGESBERICHT_UHRZEIT) + globaler
#      Marker. Zusätzlich postet send_tagesbericht() jetzt auch als Discord-
#      Embed in den Bilanz-Channel (Bankroll, Streak, Bot-Rangliste,
#      Virtuelle Konten) – vorher ging der Bericht nur an Telegram.
#
# ============================================================
# Änderungen v58 (API-Football Migration):
#   • Komplette Umstellung von livescore-api.com auf API-Football (v3.football.api-sports.io)
#   • Auth jetzt per Header (x-apisports-key), nur EIN Key nötig, kein Secret
#   • Alle Datenabruf-Funktionen (Live-Spiele, Statistiken, Events, H2H, Standings,
#     Lineups, Fixtures) intern auf die gleiche Datenstruktur gemappt wie vorher,
#     damit die Bot-Logik selbst unverändert bleibt
#   • "Dangerous Attacks" und "Free Kicks" komplett entfernt (bei API-Football
#     nicht/nur unzuverlässig verfügbar) – betroffene Zusatzfilter in
#     Ecken-/Druck-/Comeback-/xG-/Anomalie-Bot laufen jetzt ohne diese Werte
#
# Änderungen aus v57 (unverändert übernommen):
#   • FIX: Druck-Bot / Comeback-Bot / CornerRush-Bot sperrten Spiele SOFORT beim
#     ersten Sehen (vor jeder Bedingungsprüfung) -> Spiele wurden nie erneut
#     geprüft, wenn die Bedingung beim ersten Durchlauf noch nicht erfüllt war.
#     Jetzt: notified_X.add() erst unmittelbar VOR dem tatsächlichen Versand.
#   • NEU: Hartes Tageslimit für die livescore-api.com Anfragen (75.000/Tag,
#     Hard-Stop bei 73.000 als Sicherheitspuffer). rate_limit_check() blockiert
#     weitere Calls ab dem Hard-Stop, bis der Tag (de_now()) wechselt.
#   • CornerRush-Bot hat jetzt zusätzlich signal_spam_check() wie die anderen
#     Live-Bots.
#
# Änderungen aus v55/v56 (unverändert übernommen):
#   • Bots entfernt: Ecken-Über, Karten, Rotkarte, CS2 (später reaktivierbar)
#   • Ecken-Unter: Erweiterte Live-Validierung (Schüsse, DA-Ratio, Ballbesitz, Tordiff, FK)
#   • Auswertung: 90s Retry-Intervall, max. 20 Versuche, Soft-Verify, Live-Liste-Check
#   • Discord: Alle Webhooks als env vars (kein hardcoded Fallback = kein Kanal-Chaos)
#   • notified_*: 6 fehlende Sets persistiert (corner_rush/anomalie/early_goal/hz2/rk_ecken/value)
#   • Telegram: Auth-Check – nur erlaubte Chat-IDs dürfen Befehle senden
#   • Memory: cleanup_memory() täglich 03:00 Uhr (bereits_getippt + aktive_tipps)
#   • Health: /health auf Port 8080 konsolidiert (kein separater Port 8081 mehr)
#   • Startup: API-Key Validierung mit klarer Fehlermeldung
#   • Spam: signal_spam_check() in Ecken/Druck/Comeback aktiv
#   • EV-Score: Wird jetzt in Signal-Nachrichten angezeigt
#   • TELEGRAM_CHAT_PREMATCH: als env var
import os, requests, re, time, threading
from datetime import datetime, timezone, timedelta

# ============================================================
#  BUILD-VERSION – bei jedem Deploy prüfbar im Railway-Log (ganz oben) und in der
#  Telegram-Startmeldung. So lässt sich immer zweifelsfrei sehen, welcher Code-Stand
#  tatsächlich läuft, statt zu raten ob ein Deploy wirklich durchgekommen ist.
# ============================================================
BOT_VERSION      = "v60.0"
BOT_VERSION_INFO = "Rueckumstellung von API-Football auf livescore-api.com; Quoten (Ueber/Unter, Ecken-Markt) deaktiviert, Value-/Arbitrage-Bot pausiert"

# ============================================================
#  KONFIGURATION
# ============================================================
API_KEY            = os.environ.get("LIVESCORE_API_KEY","")     # v60: livescore-api.com Key (Railway-Variable setzen!)
API_SECRET         = os.environ.get("LIVESCORE_API_SECRET","")  # v60: livescore-api.com Secret (Railway-Variable setzen!)
TELEGRAM_BOT_TOKEN = os.environ.get("TELEGRAM_TOKEN",  "8706066107:AAHHVuC-M_gz-sr5sbm-7zkb7QGygfoRHoM")
TELEGRAM_CHAT_ID   = os.environ.get("TELEGRAM_CHAT_ID","@BettingLabPubDE")
VIP_TELEGRAM_HANDLE = os.environ.get("VIP_TELEGRAM_HANDLE", "BettingxLabs")  # v60.2 NEU: euer @Handle für die VIP-Werbung (ohne @)
VIP_WERBUNG_UHRZEIT = 22  # v60.2 NEU: Uhrzeit (deutsche Zeit), zu der die tägliche VIP-Werbung raus geht

# ── Discord Webhooks ─────────────────────────────────────────
# Alle Webhooks MÜSSEN als Railway-Umgebungsvariablen gesetzt sein.
# Kein hardcoded Fallback – verhindert dass alle Signale in denselben Kanal laufen.
# v59.10 SICHERHEITSFIX: Alle Fallback-URLs entfernt. Diese Webhook-URLs standen vorher als
# Klartext-Fallback im Code – wurde die Datei jemals öffentlich sichtbar (z.B. GitHub), waren
# sie damit für jeden nutzbar (eine Webhook-URL wirkt wie ein Passwort: wer sie kennt, kann ohne
# jede Anmeldung Nachrichten in den Kanal posten). Genau das war vermutlich die Ursache für die
# Spam-Nachrichten in euren Discord-Kanälen. Jetzt MÜSSEN alle Webhook-URLs als Railway-
# Umgebungsvariablen gesetzt sein – ohne gesetzte Variable bleibt der jeweilige Kanal stumm
# (send_discord_embed() prüft das ab und überspringt den Versand, siehe unten).
DISCORD_WEBHOOK_ECKEN    = os.environ.get("DISCORD_WEBHOOK_ECKEN",    "")
DISCORD_WEBHOOK_KARTEN   = os.environ.get("DISCORD_WEBHOOK_KARTEN",   "")
DISCORD_WEBHOOK_TORWART  = os.environ.get("DISCORD_WEBHOOK_TORWART",  "")
DISCORD_WEBHOOK_BILANZ   = os.environ.get("DISCORD_WEBHOOK_BILANZ",   "")
DISCORD_WEBHOOK_DRUCK    = os.environ.get("DISCORD_WEBHOOK_DRUCK",    "")
DISCORD_WEBHOOK_COMEBACK = os.environ.get("DISCORD_WEBHOOK_COMEBACK", "")
DISCORD_WEBHOOK_TORFLUT  = os.environ.get("DISCORD_WEBHOOK_TORFLUT",  "")
DISCORD_WEBHOOK_ROTKARTE = os.environ.get("DISCORD_WEBHOOK_ROTKARTE", "")
DISCORD_WEBHOOK_HZ1TORE  = os.environ.get("DISCORD_WEBHOOK_HZ1TORE",  "")
DISCORD_WEBHOOK_VZTORE   = os.environ.get("DISCORD_WEBHOOK_VZTORE",   "")
DISCORD_WEBHOOK_TORE     = os.environ.get("DISCORD_WEBHOOK_TORE",     "")
DISCORD_WEBHOOK_VALUE    = os.environ.get("DISCORD_WEBHOOK_VALUE",    "")
DISCORD_WEBHOOK_CS2      = os.environ.get("DISCORD_WEBHOOK_CS2",      "")
DISCORD_WEBHOOK_TAGESTIPP = os.environ.get("DISCORD_WEBHOOK_TAGESTIPP", "")  # v59.29 NEU: eigener Kanal für "Tipp des Tages" (vorher fälschlich im Auswertungs-/Bilanz-Kanal)

# ── v59: Discord-Bot-Token für Reaktionen (✅/❌) auf ausgewertete Signale ──
# Webhooks können KEINE Reaktionen setzen – dafür braucht es einen echten Bot
# mit "Add Reactions" + "View Channel"-Rechten in den jeweiligen Kanälen.
DISCORD_BOT_TOKEN = os.environ.get("DISCORD_BOT_TOKEN", "")

def discord_add_reaction(channel_id: str, message_id: str, emoji: str):
    """v59: Setzt eine Reaktion (URL-encodetes Emoji, z.B. %E2%9C%85 = ✅) auf eine Nachricht."""
    if not DISCORD_BOT_TOKEN:
        print(f"  [Discord Reaction] Übersprungen – DISCORD_BOT_TOKEN nicht gesetzt")
        return
    if not channel_id or not message_id:
        print(f"  [Discord Reaction] Übersprungen – keine message_id/channel_id gespeichert (Signal von vor dem Update?)")
        return
    try:
        url = f"https://discord.com/api/v10/channels/{channel_id}/messages/{message_id}/reactions/{emoji}/@me"
        resp = requests.put(url, headers={"Authorization": f"Bot {DISCORD_BOT_TOKEN}"}, timeout=10)
        if resp.status_code in (204, 200):
            print(f"  [Discord Reaction] ✅ Reaktion {emoji} gesetzt (message_id={message_id})")
        else:
            print(f"  [Discord Reaction Fehler] {resp.status_code}: {resp.text}")
    except Exception as e:
        print(f"  [Discord Reaction Fehler] {e}")

ODDS_API_KEY       = os.environ.get("ODDS_API_KEY",   "")
PANDASCORE_API_KEY = os.environ.get("PANDASCORE_KEY", "")
FOOTBALLDATA_KEY   = os.environ.get("FOOTBALLDATA_KEY","")
ANTHROPIC_API_KEY  = os.environ.get("ANTHROPIC_KEY",  "")
CLAUDE_MAX_PRO_TAG = 3

TELEGRAM_FILTER_DATEI = "telegram_filter.json"
_telegram_deaktiviert = set()

TELEGRAM_BOT_NAMEN = {
    "ecken":       "📐 Ecken-Unter",
    "torwart":     "🧤 Torwart-Bot",
    "druck":       "🔥 Druck-Bot",
    "comeback":    "🔄 Comeback-Bot",
    "torflut":     "🌊 Torflut-Bot",
    "value":       "💎 Value-Bot",
    "arbitrage":   "💰 Arbitrage-Bot",
    "xg":          "📊 xG-Bot",
    "earlygoal":   "⚡ EarlyGoal-Bot",
    "hz2tore":     "🥅 HZ2-Tore-Bot",
    "cornerrush":  "📐 CornerRush-Bot",
    "sharp":       "💼 Sharp-Money-Bot",
    "anomalie":    "🚨 Anomalie-Bot",
    "prematch":    "📅 PreMatch-Bot",
    "tagesoverview":"🌅 Morgen-Übersicht",
}

def telegram_filter_laden():
    import json, os as _os
    global _telegram_deaktiviert
    if not _os.path.exists(TELEGRAM_FILTER_DATEI):
        return
    try:
        with open(TELEGRAM_FILTER_DATEI) as f:
            _telegram_deaktiviert = set(json.load(f))
        if _telegram_deaktiviert:
            print(f"  [TG-Filter] Deaktiviert: {', '.join(_telegram_deaktiviert)}")
    except Exception as e:
        print(f"  [TG-Filter] Ladefehler: {e}")

def telegram_filter_speichern():
    import json
    try:
        with open(TELEGRAM_FILTER_DATEI, "w") as f:
            json.dump(list(_telegram_deaktiviert), f)
    except Exception as e:
        print(f"  [TG-Filter] Speicherfehler: {e}")

def dynamische_filter_laden():
    import json, os as _os
    global DYNAMISCHE_FILTER
    pfad = "dynamische_filter.json"
    if not _os.path.exists(pfad):
        return
    try:
        with open(pfad, "r") as f:
            data = json.load(f)
        for typ, werte in data.items():
            if typ in DYNAMISCHE_FILTER:
                DYNAMISCHE_FILTER[typ].update(werte)
        print(f"  [DynFilter] Geladen: {list(data.keys())}")
    except Exception as e:
        print(f"  [DynFilter] Ladefehler: {e}")

def dynamische_filter_speichern():
    import json
    try:
        with open("dynamische_filter.json", "w") as f:
            json.dump(DYNAMISCHE_FILTER, f, indent=2)
    except Exception as e:
        print(f"  [DynFilter] Speicherfehler: {e}")

def telegram_signal_erlaubt(bot_key: str) -> bool:
    return bot_key not in _telegram_deaktiviert

_claude_calls_heute = 0
_claude_calls_datum = ""

TOP_LIGEN_CLAUDE = {
    "Premier League","Bundesliga","La Liga","Serie A","Ligue 1",
    "Champions League","Europa League","Eredivisie","Primeira Liga",
    "World Cup","European Championship","DFB-Pokal","FA Cup",
    "Copa del Rey","Coppa Italia",
}

def claude_budget_verfuegbar(liga: str = "") -> bool:
    global _claude_calls_heute, _claude_calls_datum
    heute = de_now().strftime("%Y-%m-%d")
    if _claude_calls_datum != heute:
        _claude_calls_heute = 0
        _claude_calls_datum = heute
    if _claude_calls_heute >= CLAUDE_MAX_PRO_TAG:
        print(f"  [Claude] Tages-Limit erreicht – Fallback")
        return False
    if liga and liga not in TOP_LIGEN_CLAUDE:
        return False
    return True

def claude_budget_erhoehen():
    global _claude_calls_heute
    _claude_calls_heute += 1

EINSATZ = 10.0
PREMATCH_UHRZEITEN   = [10, 16, 20]
PREMATCH_MAX_TIPPS   = 3
PREMATCH_LIGEN       = {
    "premier league","bundesliga","la liga","serie a","ligue 1",
    "champions league","europa league","eredivisie","primeira liga",
    "uefa champions league","uefa europa league","uefa conference league",
    "primera division","süper lig","scottish premiership",
}
# v58.4: Liga-Name allein reicht nicht (viele Länder haben eine eigene "Premier
# League", z.B. Kasachstan) - deshalb zusätzlich das Land prüfen. Europäische
# Wettbewerbe (Champions/Europa/Conference League) sind länderübergreifend.
TOP_LIGEN_LAENDER = {
    "premier league":       "england",
    "bundesliga":           "germany",
    "la liga":               "spain",
    "primera division":      "spain",
    "serie a":               "italy",
    "ligue 1":               "france",
    "eredivisie":            "netherlands",
    "primeira liga":         "portugal",
    "süper lig":             "turkey",
    "scottish premiership":  "scotland",
}
EUROPA_WETTBEWERBE = {
    "champions league","europa league","uefa champions league",
    "uefa europa league","uefa conference league",
}

def ist_top_liga(liga_name: str, country_name: str = "") -> bool:
    liga_lower    = (liga_name or "").lower()
    country_lower = (country_name or "").lower()
    for teil in EUROPA_WETTBEWERBE:
        if teil in liga_lower:
            return True
    for teil,land in TOP_LIGEN_LAENDER.items():
        if teil in liga_lower and land in country_lower:
            return True
    return False

TELEGRAM_CHAT_PREMATCH = os.environ.get("TELEGRAM_CHAT_PREMATCH", "-1001510152037")  # Besser: env var setzen

MAX_CORNERS          = 5
MIN_KARTEN           = 2
KARTEN_BIS_MINUTE    = 40
MIN_SHOTS_ON_TARGET  = 3
FUSSBALL_INTERVAL    = 2
TAGESBERICHT_UHRZEIT = 2   # 02:00 Uhr – alle Spiele sind dann sicher beendet
MIN_QUOTE            = 1.3
KELLY_FRACTION       = 0.25
KELLY_MAX_EINSATZ    = 50.0
KELLY_MIN_EINSATZ    = 2.0
WETTER_WIND_GRENZE   = 35
WETTER_REGEN_GRENZE  = 2
LIGA_MIN_TIPPS        = 10
LIGA_MIN_TREFFERQUOTE = 0.40
ECKEN_HISTORY_SPIELE = 5
ECKEN_TOLERANZ       = 1.5
H2H_MIN_SPIELE       = 3
H2H_MAX_SPIELE       = 10
H2H_SIGNAL_BIS_MIN   = 15
HZ1_UEBER_GRENZE     = 1.2
HZ1_UNTER_GRENZE     = 0.7
VZ_UEBER_GRENZE      = 2.7
VZ_UNTER_GRENZE      = 1.8
MIN_BOOKMAKER_ANZAHL = 0
CLV_WARNSCHWELLE     = 0.05
MIN_DRUCK_ECKEN      = 6
DRUCK_RATIO          = 2.5
COMEBACK_AB_MINUTE   = 30
SPAETESTE_SIGNAL_MINUTE = 80  # v59.13 FIX: Live-Signale mit Zeitkomponente (nächstes Tor/Ecke)
                              # werden ab dieser Minute nicht mehr gepostet – zu wenig Restzeit,
                              # damit der Tipp noch realistisch eintreten kann.
TORFLUT_MIN_TORE     = 3
BANKROLL             = 100.0
BANKROLL_DATEI       = "bankroll.json"
MULTI_SIGNAL_BONUS   = 2
API_MONITOR_DATEI    = "api_monitor.json"
BEOBACHTETE_DATEI    = "beobachtete_spiele.json"
BOT_PAUSIERT         = False
GITHUB_TOKEN          = os.environ.get("GITHUB_TOKEN", "")
GITHUB_REPO           = os.environ.get("GITHUB_REPO",  "Frums72/alle-bots")
GITHUB_BACKUP_UHRZEIT = 2
MAX_BEOBACHTUNG_STUNDEN = 3
SIGNAL_TRACKER_DATEI = "signal_tracker.json"
VALUE_BET_MIN_QUOTE  = 1.6
VALUE_BET_MIN_VALUE  = 0.25  # v59.17: von 15% auf 25% erhöht (Qualität statt Quantität)
MIN_KONFIDENZ_VERSAND = 6    # v59.17: Live-Signale unter dieser Konfidenz werden nicht mehr gesendet

# ── v57: Hartes Tageslimit für livescore-api.com Anfragen ───────────
# v60.1: Von 75.000 auf 45.000 runtergesetzt (kleinerer API-Plan). Der Hard-Stop
# unten ist eine HARTE Grenze, die NIE überschritten wird, egal was passiert.
API_DAILY_LIMIT     = 45000
API_DAILY_HARD_STOP = 43000   # Sicherheitspuffer – ab hier werden Calls blockiert
_api_hard_stop_alarm_gesendet = False

# v60.1 NEU: Nacht-Drosselung. Ziel: das Kontingent soll sich auf den Tag
# (Morgens bis Abends) konzentrieren, nicht nachts verbraucht werden, wo eh
# weniger relevante Spiele laufen. Es gibt KEIN festes Nacht-Kontingent –
# stattdessen wird nachts hochgerechnet, ob beim aktuellen Verbrauchstempo das
# Tageslimit bis Mitternacht gerissen würde. Nur DANN wird pausiert; läuft der
# Tag ruhig, darf der Bot auch nachts ganz normal weiterlaufen.
NACHT_ENDE_STUNDE   = 7     # ab dieser Uhrzeit (deutsche Zeit) gilt es nicht mehr als "Nacht"
NACHT_DROSSEL_MARGE = 0.90  # Drosselung greift, wenn Hochrechnung >90% vom Tageslimit ergibt
_nacht_drosselung_alarm_gesendet = False

class ApiTagesLimitErreicht(Exception):
    """Wird geworfen wenn das tägliche livescore-api.com Kontingent erschöpft ist
    ODER wenn die Nacht-Drosselung greift (siehe NACHT_ENDE_STUNDE oben)."""
    pass

KOMBI_SIGNAL_TYPEN = {
    frozenset(["hz1tore","torflut"]): "Torreiches Spiel",
    frozenset(["hz1tore","vztore"]):  "Tore-Dominanz",
    frozenset(["druck","comeback"]):  "Spannendes Duell",
    frozenset(["torwart","druck"]):   "Druck + Chancen",
}
kombi_gesendet = set()

DYNAMISCHE_FILTER = {
    "comeback": {"COMEBACK_AB_MINUTE": COMEBACK_AB_MINUTE},
    "druck":    {"DRUCK_RATIO":        DRUCK_RATIO},
    "torwart":  {"MIN_SHOTS_ON_TARGET":MIN_SHOTS_ON_TARGET},
    "karten":   {"KARTEN_BIS_MINUTE":  KARTEN_BIS_MINUTE},
}

# ============================================================
#  v60: LIVESCORE-API.COM – Basis-URL, Auth, Transform
# ============================================================
# v60 ROLLBACK auf livescore-api.com (vorher API-Football, davor auch schon
# livescore-api.com – siehe v57/v58-Kommentare oben). Auth läuft bei
# livescore-api.com NICHT über einen Header, sondern über 2 Felder als
# normale URL-Parameter: key + secret. Deshalb hängt api_get_with_retry()
# unten die beiden Werte automatisch an JEDEN Call an.
LS_BASE = "https://livescore-api.com/api-client"

KARTEN_TYPEN   = {"Yellow Card","Red Card","Yellow Red Card"}
ROTKARTE_TYPEN = {"Red Card","Yellow Red Card"}

# livescore-api.com nutzt eigene Status-Texte, die zum Glück fast 1:1 zu
# unseren internen Status-Werten passen (die intern genutzten Werte wurden
# ursprünglich genau an diese API angelehnt). "ADDED TIME" wird wie "IN PLAY"
# behandelt, damit der Rest vom Code (der nur IN PLAY/HALF TIME BREAK/FT/NS
# kennt) unverändert funktioniert.
_LS_STATUS_MAP = {
    "NOT STARTED":"NS","IN PLAY":"IN PLAY","ADDED TIME":"IN PLAY",
    "HALF TIME BREAK":"HALF TIME BREAK","FINISHED":"FT",
    "INSUFFICIENT DATA":"IN PLAY","POSTPONED":"PST","CANCELLED":"CANC",
    "ABANDONED":"ABD","AWARDED":"AWD","WALKOVER":"WO",
}
_LS_ZEIT_MAP = {"FT":"FT","AET":"AET","AP":"AP","HT":"HT"}

_af_home_team_cache = {}

def aktuelle_saison() -> int:
    now = de_now()
    return now.year if now.month >= 7 else now.year-1

def _ls_status(m: dict) -> str:
    return _LS_STATUS_MAP.get((m.get("status") or "").upper(), m.get("status",""))

def _af_fixture_zu_intern(m: dict) -> dict:
    """Wandelt ein Match-Objekt aus der livescore-api.com Antwort (matches/live.json,
    matches/history.json, fixtures/list.json, teams/head2head.json, ...) in unser
    internes Format um. Die livescore-api.com Antwort hat praktisch schon genau diese
    Struktur (home/away/competition/country/scores), deshalb ist das hier größtenteils
    ein reines Durchreichen mit sauberen Strings/Defaults."""
    home = m.get("home") or {}
    away = m.get("away") or {}
    comp = m.get("competition") or {}
    country = m.get("country") or {}
    scores  = m.get("scores") or {}

    match_id = str(m.get("id",""))
    home_id  = str(home.get("id",""))
    if match_id and home_id:
        _af_home_team_cache[match_id] = home_id

    zeit_roh = m.get("time","")
    zeit = _LS_ZEIT_MAP.get(zeit_roh, zeit_roh)  # z.B. "43" bleibt "43", "FT"/"HT" bleiben

    return {
        "id": match_id,
        "status": _ls_status(m),
        "time": zeit,
        "home": {"id":home_id,"name":home.get("name","?")},
        "away": {"id":str(away.get("id","")),"name":away.get("name","?")},
        "competition": {"id":str(comp.get("id","")),"name":comp.get("name","?")},
        "country": {"name":(country.get("name") if country else "International")},
        "scores": {
            "score":    scores.get("score","0 - 0") or "0 - 0",
            "ht_score": scores.get("ht_score","") or "",
        },
    }

def _af_fixture_zu_prematch(fx: dict) -> dict:
    """Wandelt einen Eintrag aus fixtures/list.json (livescore-api.com) in unser
    internes Prematch-Format um."""
    home = fx.get("home") or {}
    away = fx.get("away") or {}
    comp = fx.get("competition") or {}
    country = fx.get("country") or {}
    # v60: livescore-api.com liefert Uhrzeiten in UTC (kein timezone-Parameter wie
    # vorher bei API-Football). +2h passt zur restlichen Codebasis, die für
    # "deutsche Zeit" ebenfalls fest +2h auf UTC rechnet (siehe de_now() oben).
    zeit = fx.get("time","?") or "?"
    try:
        h,mi = zeit.split(":")[:2]
        stunde = (int(h)+2) % 24
        zeit = f"{stunde:02d}:{mi}"
    except Exception:
        zeit = zeit[:5] if len(zeit) >= 5 else zeit

    home_id = str(home.get("id",""))
    fixture_id = str(fx.get("id",""))
    if fixture_id and home_id:
        _af_home_team_cache[fixture_id] = home_id
    return {
        "id": fixture_id,
        "status": "NS",  # fixtures/list.json liefert nur zukünftige, noch nicht gestartete Spiele
        "home_name": home.get("name","?"),
        "away_name": away.get("name","?"),
        "home": {"id":home_id,"name":home.get("name","?")},
        "away": {"id":str(away.get("id","")),"name":away.get("name","?")},
        "competition": {"id":str(comp.get("id","")),"name":comp.get("name","?")},
        "country": {"name":(country.get("name") if country else "International")},
        "time": zeit,
    }

_cache_matches    = []
_cache_timestamp  = 0
_cache_lock       = threading.Lock()
CACHE_TTL         = 20

_stats_cache       = {}
_events_cache      = {}
_stats_cache_lock  = threading.Lock()
_events_cache_lock = threading.Lock()
STATS_CACHE_TTL    = 60
EVENTS_CACHE_TTL   = 45
_statistik_lock    = threading.Lock()

notified_ecken      = set()
notified_ecken_over = set()
notified_karten     = set()
notified_torwart    = set()
notified_druck      = set()
notified_comeback   = set()
notified_torflut    = set()
notified_rotkarte   = set()
notified_hz1tore    = set()
notified_vztore     = set()
notified_xg         = set()
notified_cs2        = set()
notified_sharp      = set()
beobachtete_spiele  = {}
auswertung_done     = set()
NOTIFIED_DATEI      = "notified_sets.json"

bereits_getippt  = {}
fehler_zaehler   = {}
FEHLER_ALERT_AB  = 3
_api_calls_log   = []
_api_calls_lock  = threading.Lock()
MAX_API_PER_MIN  = 400  # v59.27 FIX: von 50 auf 400 erhöht. Euer API-Football-Abo ist der
                         # "Ultra"-Plan (75.000 Requests/Tag) – der erlaubt laut API-Football
                         # offiziell 450 Calls/Minute, nicht 50. Das selbst gesetzte Limit war
                         # damit fast 10x enger als nötig und hat unnötig zu ständigem
                         # "Rate-Limit"-Warten geführt, wodurch auch die Auswertungs-Calls des
                         # Nachschau-Bots im Stau steckenblieben. 400 statt 450 als Sicherheitspuffer.
aktive_tipps     = {}
signal_log       = []
SIGNAL_LOG_DATEI = "signal_log.json"
_wetter_cache    = {}
WETTER_TTL       = 1800
_ecken_avg_cache = {}
ECKEN_AVG_TTL    = 3600
_standings_cache = {}
STANDINGS_TTL    = 1800
streak_aktuell   = 0
streak_beste     = 0

statistik = {
    "ecken":    {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "ecken_over":{"gewonnen":0,"verloren":0,"gewinn":0.0},
    "karten":   {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "torwart":  {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "druck":    {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "comeback": {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "torflut":  {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "rotkarte": {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "hz1tore":  {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "vztore":   {"gewonnen":0,"verloren":0,"gewinn":0.0},
    "hz2tore":  {"gewonnen":0,"verloren":0,"gewinn":0.0},  # v59.22 NEU: eigener Bucket statt fälschlich unter hz1tore
}
wochen_statistik = {k:{"gewonnen":0,"verloren":0,"gewinn":0.0} for k in statistik}
stunden_statistik = {str(h):{"gewonnen":0,"verloren":0} for h in range(24)}
liga_statistik    = {}
tagesbericht_gesendet = None
vip_werbung_gesendet = None  # v60.2 NEU: gleiches Prinzip wie tagesbericht_gesendet
STATISTIK_DATEI = "statistik.json"


# ============================================================
#  STATISTIK SPEICHERN / LADEN
# ============================================================

def statistik_laden():
    global statistik,wochen_statistik,tagesbericht_gesendet,vip_werbung_gesendet,stunden_statistik,liga_statistik,signal_log
    import json, os as _os
    if not _os.path.exists(STATISTIK_DATEI):
        return
    try:
        with open(STATISTIK_DATEI,"r") as f:
            data = json.load(f)
        for k in statistik:
            if k in data.get("statistik",{}):
                statistik[k] = data["statistik"][k]
        for k in wochen_statistik:
            if k in data.get("wochen_statistik",{}):
                wochen_statistik[k] = data["wochen_statistik"][k]
        if data.get("stunden_statistik"):
            stunden_statistik.update(data["stunden_statistik"])
        if data.get("liga_statistik"):
            liga_statistik.update(data["liga_statistik"])
        if data.get("tagesbericht_gesendet"):
            from datetime import date
            tagesbericht_gesendet = date.fromisoformat(data["tagesbericht_gesendet"])
        if data.get("vip_werbung_gesendet"):
            from datetime import date
            vip_werbung_gesendet = date.fromisoformat(data["vip_werbung_gesendet"])
        print(f"  [Statistik] Geladen aus {STATISTIK_DATEI}")
    except Exception as e:
        print(f"  [Statistik] Ladefehler: {e}")
    import os as _os2
    if _os2.path.exists(SIGNAL_LOG_DATEI):
        try:
            with open(SIGNAL_LOG_DATEI,"r") as f:
                signal_log = json.load(f)
            print(f"  [Signal-Log] {len(signal_log)} Einträge geladen")
        except Exception as e:
            print(f"  [Signal-Log] Ladefehler: {e}")

def statistik_speichern():
    import json
    try:
        data = {
            "statistik":statistik,"wochen_statistik":wochen_statistik,
            "stunden_statistik":stunden_statistik,"liga_statistik":liga_statistik,
            "tagesbericht_gesendet":str(tagesbericht_gesendet) if tagesbericht_gesendet else None,
            "vip_werbung_gesendet":str(vip_werbung_gesendet) if vip_werbung_gesendet else None,
        }
        with open(STATISTIK_DATEI,"w") as f:
            json.dump(data,f,indent=2)
    except Exception as e:
        print(f"  [Statistik] Speicherfehler: {e}")

def signal_log_speichern():
    import json
    try:
        with open(SIGNAL_LOG_DATEI,"w") as f:
            json.dump(signal_log[-500:],f,indent=2)
    except Exception as e:
        print(f"  [Signal-Log] Speicherfehler: {e}")

_notified_sets_letzter_push = 0

def notified_sets_speichern():
    import json
    global _notified_sets_letzter_push
    try:
        data = {
            "ecken":           list(notified_ecken),
            "ecken_over":      list(notified_ecken_over),
            "karten":          list(notified_karten),
            "torwart":         list(notified_torwart),
            "druck":           list(notified_druck),
            "comeback":        list(notified_comeback),
            "torflut":         list(notified_torflut),
            "rotkarte":        list(notified_rotkarte),
            "hz1tore":         list(notified_hz1tore),
            "vztore":          list(notified_vztore),
            "auswertung_done": list(auswertung_done),
            "xg":              list(notified_xg),
            "cs2":             list(notified_cs2),
            "sharp":           list(notified_sharp),
            "corner_rush":     list(notified_corner_rush),
            "anomalie":        list(notified_anomalie),
            "early_goal":      list(notified_early_goal),
            "hz2":             list(notified_hz2),
            "rk_ecken":        list(notified_rk_ecken),
            "value":           list(notified_value),
            "gespeichert":     de_now().strftime("%Y-%m-%d %H:%M"),
        }
        with open(NOTIFIED_DATEI,"w") as f:
            json.dump(data,f)
        if time.time() - _notified_sets_letzter_push > 120:
            _notified_sets_letzter_push = time.time()
            threading.Thread(target=_notified_sets_github_push,daemon=True).start()
    except Exception as e:
        print(f"  [Notified] Speicherfehler: {e}")

def _notified_sets_github_push():
    import base64, os as _os
    if not GITHUB_TOKEN or not _os.path.exists(NOTIFIED_DATEI):
        return
    try:
        headers = {"Authorization":f"token {GITHUB_TOKEN}","Accept":"application/vnd.github.v3+json"}
        with open(NOTIFIED_DATEI,"rb") as f:
            inhalt = base64.b64encode(f.read()).decode()
        resp = _github_push_datei(NOTIFIED_DATEI,inhalt,headers)
        if resp.status_code not in (200,201):
            _github_push_datei(NOTIFIED_DATEI,inhalt,headers)
    except Exception:
        pass

def notified_sets_laden():
    """
    v59 FIX (Duplikate): Vorher wurden die gespeicherten "bereits gepostet"-Sets
    verworfen, wenn die letzte Speicherung länger als 30 Minuten her war – auch
    am selben Tag. Nach jedem Neustart (Redeploy/Crash auf Railway) hat der Bot
    dadurch alle bereits getippten Live-Spiele vergessen und teils erneut
    gepostet (gleiches Spiel, gleicher Tipp, kurz nacheinander). Jetzt zählt
    nur noch der Tages-Check.
    """
    import json, os as _os
    global notified_ecken,notified_ecken_over,notified_karten,notified_torwart
    global notified_druck,notified_comeback,notified_torflut,notified_rotkarte
    global notified_hz1tore,notified_vztore,auswertung_done
    global notified_xg,notified_cs2,notified_sharp
    global notified_corner_rush,notified_anomalie,notified_early_goal
    global notified_hz2,notified_rk_ecken,notified_value
    if not _os.path.exists(NOTIFIED_DATEI):
        return
    try:
        with open(NOTIFIED_DATEI) as f:
            data = json.load(f)
        if data.get("gespeichert","")[:10] != de_now().strftime("%Y-%m-%d"):
            print(f"  [Notified] Sets von gestern – nicht geladen")
            return
        notified_ecken      = set(data.get("ecken",     []))
        notified_ecken_over = set(data.get("ecken_over",[]))
        notified_karten     = set(data.get("karten",    []))
        notified_torwart    = set(data.get("torwart",   []))
        notified_druck      = set(data.get("druck",     []))
        notified_comeback   = set(data.get("comeback",  []))
        notified_torflut    = set(data.get("torflut",   []))
        notified_rotkarte   = set(data.get("rotkarte",  []))
        notified_hz1tore    = set(data.get("hz1tore",   []))
        notified_vztore     = set(data.get("vztore",    []))
        auswertung_done     = set(data.get("auswertung_done",[]))
        notified_xg         = set(data.get("xg",          []))
        notified_cs2        = set(data.get("cs2",         []))
        notified_sharp      = set(data.get("sharp",       []))
        notified_corner_rush  = set(data.get("corner_rush",  []))
        notified_anomalie     = set(data.get("anomalie",     []))
        notified_early_goal   = set(data.get("early_goal",   []))
        notified_hz2          = set(data.get("hz2",          []))
        notified_rk_ecken     = set(data.get("rk_ecken",     []))
        notified_value        = set(data.get("value",        []))
        total = sum(len(s) for s in [notified_ecken,notified_ecken_over,notified_karten,
            notified_torwart,notified_druck,notified_comeback,notified_torflut,
            notified_rotkarte,notified_hz1tore,notified_vztore,
            notified_xg,notified_cs2,notified_sharp,
            notified_corner_rush,notified_anomalie,notified_early_goal,
            notified_hz2,notified_rk_ecken,notified_value])
        print(f"  [Notified] {total} Match-IDs geladen ✅")
    except Exception as e:
        print(f"  [Notified] Ladefehler: {e}")

# ============================================================
#  HILFSFUNKTIONEN
# ============================================================

def jetzt():
    return (datetime.now(timezone.utc)+timedelta(hours=2)).strftime("%H:%M")

def heute():
    return (datetime.now(timezone.utc)+timedelta(hours=2)).strftime("%d.%m.%Y")

def de_now():
    return datetime.now(timezone.utc)+timedelta(hours=2)

def parse_score(score_str):
    try:
        parts = score_str.replace(" ","").split("-")
        return int(parts[0]),int(parts[1])
    except Exception:
        return 0,0

def html_zu_discord(text):
    text = text.replace("<b>","**").replace("</b>","**")
    text = re.sub(r"<[^>]+>","",text)
    return text

_signal_stunde_zaehler = {}

def signal_spam_check() -> bool:
    stunde = de_now().strftime("%Y-%m-%d-%H")
    _signal_stunde_zaehler[stunde] = _signal_stunde_zaehler.get(stunde,0)+1
    for k in list(_signal_stunde_zaehler.keys()):
        if k != stunde:
            del _signal_stunde_zaehler[k]
    if _signal_stunde_zaehler[stunde] > 8:
        print(f"  [Spam] Signal unterdrückt – {_signal_stunde_zaehler[stunde]}/8 diese Stunde")
        return False
    return True

def send_telegram(message: str):
    if BOT_PAUSIERT:
        print("  [Telegram] Signal unterdrückt (Bot pausiert)")
        return
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id":TELEGRAM_CHAT_ID,"text":message,"parse_mode":"HTML"}
    resp = requests.post(url,json=payload,timeout=10)
    if resp.status_code != 200:
        print(f"  [Telegram Fehler] {resp.text}")

def send_telegram_gruppe(message: str, chat_id: str = None):
    ziel = chat_id or TELEGRAM_CHAT_PREMATCH
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    payload = {"chat_id":ziel,"text":message,"parse_mode":"HTML"}
    resp = requests.post(url,json=payload,timeout=10)
    if resp.status_code != 200:
        print(f"  [Telegram Gruppe Fehler] {resp.text}")

def send_discord_embed(webhook_url: str, embed: dict, wait_for_message: bool = False):
    """
    v59: Kann jetzt optional auf die Discord-Antwort warten (wait=true) und gibt
    dann {"message_id":..., "channel_id":...} zurück. Damit lässt sich später
    (nach der Auswertung) eine Reaktion direkt auf diese Nachricht setzen.
    """
    if not webhook_url or webhook_url.startswith("DISCORD"):
        return None
    try:
        params = {"wait": "true"} if wait_for_message else None
        resp = requests.post(webhook_url, params=params, json={"embeds":[embed]}, timeout=10)
        if resp.status_code not in (200,204):
            print(f"  [Discord Fehler] {resp.status_code}: {resp.text}")
            return None
        if wait_for_message and resp.status_code == 200:
            data = resp.json()
            return {"message_id": data.get("id"), "channel_id": data.get("channel_id")}
    except Exception as e:
        print(f"  [Discord Fehler] {e}")
    return None

def send_discord(webhook_url: str, message: str):
    if not webhook_url or webhook_url.startswith("DISCORD"):
        return
    discord_msg = html_zu_discord(message)
    resp = requests.post(webhook_url,json={"content":discord_msg},timeout=10)
    if resp.status_code not in (200,204):
        print(f"  [Discord Fehler] {resp.status_code}: {resp.text}")

# ── v57: Tages-Reset-Helper für den API-Monitor ──────────────────────
def _api_monitor_tag_pruefen():
    global _api_hard_stop_alarm_gesendet
    heute_str = de_now().strftime("%Y-%m-%d")
    if _api_monitor["datum"] != heute_str:
        _api_monitor["heute"] = 0
        _api_monitor["datum"] = heute_str
        _api_hard_stop_alarm_gesendet = False

def _nacht_drosselung_greift() -> bool:
    """v60.1: Prüft NUR nachts (vor NACHT_ENDE_STUNDE), ob der Tag beim aktuellen
    Tempo das Limit reißen würde. Rechnet den bisherigen Verbrauch seit Mitternacht
    auf 24h hoch – reißt diese Hochrechnung die NACHT_DROSSEL_MARGE, wird pausiert,
    damit für den Tag (7 Uhr bis Mitternacht) noch genug Kontingent übrig bleibt."""
    global _nacht_drosselung_alarm_gesendet
    now = de_now()
    if now.hour >= NACHT_ENDE_STUNDE:
        _nacht_drosselung_alarm_gesendet = False  # Tag hat begonnen -> Reset für die nächste Nacht
        return False
    stunden_seit_mitternacht = max(now.hour+now.minute/60, 0.05)
    hochrechnung = _api_monitor.get("heute",0) / stunden_seit_mitternacht * 24
    if hochrechnung > API_DAILY_LIMIT*NACHT_DROSSEL_MARGE:
        if not _nacht_drosselung_alarm_gesendet:
            print(f"  [Nacht-Drosselung] Hochrechnung {hochrechnung:,.0f}/{API_DAILY_LIMIT:,} Calls "
                  f"-> neue Signale pausiert bis {NACHT_ENDE_STUNDE}:00 Uhr")
            send_telegram(
                f"🌙 <b>Nacht-Drosselung aktiv</b>\n"
                f"Beim aktuellen Tempo würde das Tageslimit gerissen werden, deshalb "
                f"pausieren neue Signale bis {NACHT_ENDE_STUNDE}:00 Uhr.\n"
                f"📊 Hochrechnung: {hochrechnung:,.0f}/{API_DAILY_LIMIT:,} Calls\n🕐 {jetzt()} Uhr")
            _nacht_drosselung_alarm_gesendet = True
        return True
    return False

def rate_limit_check():
    with _api_calls_lock:
        _api_monitor_tag_pruefen()
        if _api_monitor["heute"] >= API_DAILY_HARD_STOP:
            raise ApiTagesLimitErreicht(
                f"{_api_monitor['heute']}/{API_DAILY_LIMIT} Calls heute – Tageslimit-Schutz aktiv"
            )
        if _nacht_drosselung_greift():
            raise ApiTagesLimitErreicht(
                f"Nacht-Drosselung aktiv (vor {NACHT_ENDE_STUNDE}:00 Uhr) – Kontingent wird fürs Tagesgeschäft geschont"
            )
        now = time.time()
        while _api_calls_log and _api_calls_log[0] < now-60:
            _api_calls_log.pop(0)
        if len(_api_calls_log) >= MAX_API_PER_MIN:
            wait = 60-(now-_api_calls_log[0])+1
            print(f"  [Rate-Limit] {len(_api_calls_log)} Calls/Min – warte {wait:.0f}s")
            time.sleep(max(wait,1))
        _api_calls_log.append(time.time())
        api_monitor_increment()

def bot_fehler_melden(bot_name: str, fehler: Exception):
    if isinstance(fehler, ApiTagesLimitErreicht):
        print(f"  [{bot_name}] Pausiert – {fehler}")
        return
    fehler_zaehler[bot_name] = fehler_zaehler.get(bot_name,0)+1
    count = fehler_zaehler[bot_name]
    print(f"  [{bot_name}] Fehler #{count}: {fehler}")
    if count == FEHLER_ALERT_AB:
        msg = (f"🚨 <b>Bot-Fehler Alert!</b>\nBot: <b>{bot_name}</b>\n"
               f"Fehler: {fehler}\nAufeinanderfolgende Fehler: <b>{count}</b>\n🕐 {jetzt()} Uhr")
        send_telegram(msg)

def bot_fehler_reset(bot_name: str):
    if fehler_zaehler.get(bot_name,0) > 0:
        fehler_zaehler[bot_name] = 0

def tipp_erlaubt(match_id: str, bot_name: str) -> bool:
    if match_id in bereits_getippt:
        erster = bereits_getippt[match_id]
        if erster != bot_name:
            print(f"  [{bot_name}] Doppel-Tipp verhindert für {match_id} (bereits von {erster})")
            return False
    bereits_getippt[match_id] = bot_name
    return True

def karten_emoji(typ: str) -> str:
    t = typ.lower().replace(" ","").replace("_","")
    if "yellowred" in t: return "🟨🟥"
    if "red"       in t: return "🟥"
    if "yellow"    in t: return "🟨"
    return "🃏"

def api_get_with_retry(url: str, params: dict, max_retries: int = 3, headers: dict = None) -> requests.Response:
    # v60: livescore-api.com braucht KEINEN Auth-Header, sondern key+secret als
    # normale URL-Parameter. Die werden hier automatisch an jeden Call angehängt,
    # damit das an keiner der ~15 Aufruf-Stellen im Code einzeln gemacht werden muss.
    params = dict(params or {})
    params.setdefault("key", API_KEY)
    params.setdefault("secret", API_SECRET)
    rate_limit_check()
    for attempt in range(max_retries):
        try:
            resp = requests.get(url,params=params,headers=headers,timeout=10)
            resp.raise_for_status()
            return resp
        except requests.exceptions.HTTPError as e:
            status = e.response.status_code if e.response else 0
            if status in (503,502,504) and attempt < max_retries-1:
                wait = 2**attempt
                print(f"  [API] {status} – Retry {attempt+1}/{max_retries-1} in {wait}s")
                time.sleep(wait)
            else:
                raise
        except (requests.exceptions.ConnectionError,requests.exceptions.Timeout) as e:
            if attempt < max_retries-1:
                wait = 2**attempt
                print(f"  [API] Verbindungsfehler – Retry in {wait}s: {e}")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError("Alle Retry-Versuche fehlgeschlagen")


# ============================================================
#  API-FUNKTIONEN
# ============================================================

def _safe_int(val, default=0):
    try:
        return int(val)
    except (ValueError,TypeError):
        return default

def ls_get_live_matches():
    try:
        resp = api_get_with_retry(f"{LS_BASE}/matches/live.json",{})
        raw  = (resp.json().get("data") or {}).get("match",[]) or []
    except Exception as e:
        print(f"  [API] Live-Spiele Fehler: {e}")
        return []
    alle_matches = [_af_fixture_zu_intern(m) for m in raw]
    print(f"  [API] {len(alle_matches)} Live-Spiele geladen")
    return alle_matches

# v60: livescore-api.com liefert Statistiken als flache Liste [{"type":"corners","home":9,"away":2}, ...]
# statt pro Team getrennt wie API-Football – deshalb hier ein einfacher Type->Wert Lookup.
_LS_STAT_TYP_MAP = {
    "corners_home":       ("corners","home"),
    "corners_away":       ("corners","away"),
    "shots_on_target_home":("shots_on_target","home"),
    "shots_on_target_away":("shots_on_target","away"),
    "saves_home":         ("saves","home"),
    "saves_away":         ("saves","away"),
    "possession_home":    ("possesion","home"),   # Tippfehler "possesion" kommt so von livescore-api.com
    "possession_away":    ("possesion","away"),
}

def ls_get_statistiken(match_id):
    result = {
        "corners_home":0,"corners_away":0,
        "shots_on_target_home":0,"shots_on_target_away":0,
        "saves_home":0,"saves_away":0,
        "possession_home":"0","possession_away":"0",
    }
    try:
        params = {"match_id":match_id}
        resp   = api_get_with_retry(f"{LS_BASE}/matches/statistics.json",params)
        data   = resp.json().get("data",[]) or []
        by_typ = {d.get("type"): d for d in data if isinstance(d,dict)}
        for feld,(typ,seite) in _LS_STAT_TYP_MAP.items():
            eintrag = by_typ.get(typ)
            if not eintrag:
                continue
            wert = eintrag.get(seite,0)
            result[feld] = str(wert) if "possession" in feld else _safe_int(wert)
    except Exception as e:
        print(f"  [API] Statistik Fehler ({match_id}): {e}")
    return result

# v60: livescore-api.com Event-Typen -> unsere internen Event-Namen (siehe KARTEN_TYPEN
# oben, die auf "Yellow Card"/"Red Card"/"Yellow Red Card" prüfen)
_LS_EVENT_MAP = {
    "GOAL":"Goal","GOAL_PENALTY":"Goal","OWN_GOAL":"Goal",
    "MISSED_PENALTY":"Missed Penalty",
    "YELLOW_CARD":"Yellow Card","RED_CARD":"Red Card","YELLOW_RED_CARD":"Yellow Red Card",
    "SUBSTITUTION":"Substitution",
}

def ls_get_events(match_id):
    try:
        params = {"id":match_id}
        resp   = api_get_with_retry(f"{LS_BASE}/matches/events.json",params)
        data   = resp.json().get("data") or {}
        raw    = data.get("event",[]) or []
    except Exception as e:
        print(f"  [API] Events Fehler ({match_id}): {e}")
        return []
    events = []
    for e in raw:
        typ_roh    = (e.get("event") or "").upper()
        event_name = _LS_EVENT_MAP.get(typ_roh, e.get("label") or typ_roh)
        events.append({
            "event":     event_name,
            "time":      e.get("time",0),
            "home_away": "home" if e.get("is_home") else "away",
            "player":    {"name": (e.get("player") or {}).get("name") or "?"},
        })
    return events

def ls_get_single_match(match_id):
    """Holt ein einzelnes Spiel per (Live-)Match-ID. WICHTIG (Bugfix): livescore-api.com
    kennt beim live.json Endpoint KEINEN Filter-Parameter für die Match-ID selbst (nur
    fixture_id/competition_id/team_id/country_id – fixture_id ist aber eine ANDERE ID,
    die nur vor Spielbeginn existiert, siehe deren Doku). Ein Filter mit "fixture_id"
    hätte hier so gut wie nie etwas gefunden. Deshalb wird stattdessen die (eh schon
    gecachte) komplette Live-Liste durchsucht – kostet keinen zusätzlichen API-Call,
    wenn der 20s-Cache aus get_live_matches() noch frisch ist."""
    try:
        for m in get_live_matches():
            if str(m.get("id","")) == str(match_id):
                return m
    except Exception as e:
        print(f"  [API] Single-Match (live-cache) Fehler ({match_id}): {e}")
    # HINWEIS: livescore-api.com dokumentiert für history.json keinen offiziellen
    # "eine ID direkt abrufen"-Filter (nur competition_id/team_id/from/to/page). Als
    # Fallback wird hier trotzdem ein "id"-Parameter probiert (viele ihrer Endpunkte
    # unterstützen das inoffiziell) – falls das nicht greift, kommt einfach {} zurück
    # und der aufrufende Code behandelt das Spiel wie "nicht mehr auffindbar", was in
    # der Praxis fast nie vorkommt (nur bei sehr alten, schon aus dem Live-Feed
    # rausgefallenen Spielen).
    try:
        params = {"id":match_id}
        resp   = api_get_with_retry(f"{LS_BASE}/matches/history.json",params)
        data   = (resp.json().get("data") or {}).get("match",[]) or []
        for m in data:
            if str(m.get("id","")) == str(match_id):
                return _af_fixture_zu_intern(m)
    except Exception as e:
        print(f"  [API] Single-Match (history) Fehler ({match_id}): {e}")
    return {}

def ls_get_match_by_fixture_id(fixture_id):
    """v60.2 NEU: Sucht das (Live- oder frisch beendete) Spiel zu einer FIXTURE-ID
    (z.B. aus fixtures/list.json bzw. den Pre-Match-Tipps) – NICHT zu verwechseln mit
    ls_get_single_match() oben, das eine normale Match-ID braucht. livescore-api.com
    bietet für matches/live.json extra einen fixture_id-Filter dafür an."""
    try:
        params = {"fixture_id":fixture_id}
        resp   = api_get_with_retry(f"{LS_BASE}/matches/live.json",params)
        raw    = (resp.json().get("data") or {}).get("match",[]) or []
        if raw:
            return _af_fixture_zu_intern(raw[0])
    except Exception as e:
        print(f"  [API] Fixture-Lookup Fehler ({fixture_id}): {e}")
    return {}

def get_live_matches():
    global _cache_matches,_cache_timestamp
    with _cache_lock:
        now = time.time()
        if now-_cache_timestamp > CACHE_TTL:
            _cache_matches   = ls_get_live_matches()
            _cache_timestamp = now
            print(f"  [Cache] {len(_cache_matches)} Spiele geladen")
        return list(_cache_matches)

def get_statistiken(match_id: str) -> dict:
    now = time.time()
    with _stats_cache_lock:
        cached = _stats_cache.get(match_id)
        if cached and now-cached["ts"] < STATS_CACHE_TTL:
            return cached["data"]
    data = ls_get_statistiken(match_id)
    with _stats_cache_lock:
        _stats_cache[match_id] = {"data":data,"ts":now}
    return data

def get_events(match_id: str) -> list:
    now = time.time()
    with _events_cache_lock:
        cached = _events_cache.get(match_id)
        if cached and now-cached["ts"] < EVENTS_CACHE_TTL:
            return cached["data"]
    data = ls_get_events(match_id)
    with _events_cache_lock:
        _events_cache[match_id] = {"data":data,"ts":now}
    return data

def cache_invalidieren(match_id: str):
    with _stats_cache_lock:
        _stats_cache.pop(match_id,None)
    with _events_cache_lock:
        _events_cache.pop(match_id,None)

def cache_aufraumen():
    now = time.time()
    with _stats_cache_lock:
        abgelaufen = [k for k,v in _stats_cache.items() if now-v["ts"] > STATS_CACHE_TTL*10]
        for k in abgelaufen:
            del _stats_cache[k]
    with _events_cache_lock:
        abgelaufen = [k for k,v in _events_cache.items() if now-v["ts"] > EVENTS_CACHE_TTL*10]
        for k in abgelaufen:
            del _events_cache[k]

# ============================================================
#  v60: QUOTEN DEAKTIVIERT (livescore-api.com liefert keine Über/Unter-
#  Linien und keine Ecken-Märkte, sondern nur eine einfache 1/X/2
#  Siegquote pro Spiel – das reicht für Über/Unter-Tipps, Ecken-Tipps
#  usw. nicht aus)
# ============================================================
# Auf Wunsch komplett deaktiviert statt umgebaut: alle Funktionen unten
# behalten ihre ursprüngliche Signatur (gleicher Name, gleiche Parameter,
# gleicher Rückgabetyp) und geben aber immer "keine Quote vorhanden"
# zurück. Der Rest vom Bot-Code kennt diesen Fall bereits (er zeigt dann
# einfach keine Quote im Signal an, blockiert aber nichts) – dadurch
# musste an den ~15 Aufruf-Stellen im restlichen Code nichts geändert
# werden. Betroffen sind: Ecken-Verfügbarkeits-Check, Value-Bot,
# Arbitrage-Bot, die Quoten-Anzeige in den Signal-Nachrichten und CLV.
# Value-Bot und Arbitrage-Bot wurden zusätzlich unten aus der
# Bot-Liste rausgenommen, da sie ohne echte Quoten nie mehr etwas finden
# können und sonst nur nutzlos API-Kontingent verbrauchen würden.

def af_get_odds(fixture_id) -> list:
    return []

def af_odds_ist_live(fixture_id) -> bool:
    return False

def af_odds_beste_quote(fixture_id, linie: float = None, richtung: str = None, nur_live: bool = True) -> float:
    return None

def af_odds_details(fixture_id, linie: float = None, richtung: str = None, nur_live: bool = True) -> dict:
    return {"quote":None,"avg_quote":None,"bookmaker_anzahl":0}

def af_odds_vergleich_text(fixture_id) -> str:
    return ""

def af_odds_ecken_verfuegbar(fixture_id) -> bool:
    # True = "nicht blockieren" (wir können es nicht prüfen, also lassen wir das
    # Ecken-Signal durch statt es fälschlich zu unterdrücken)
    return True

def af_odds_fuer_value_bot(fixture_id) -> dict:
    return {}

def get_quote(match_id, typ=None, linie: float = None, richtung: str = None, nur_live: bool = True):
    return None


def kelly_einsatz(quote: float, typ: str) -> float:
    if not quote or quote <= 1.0:
        return EINSATZ
    ges = statistik[typ]["gewonnen"]+statistik[typ]["verloren"]
    if ges < 10:
        return EINSATZ
    p      = statistik[typ]["gewonnen"]/ges
    b      = quote-1.0
    kelly  = (b*p-(1-p))/b
    kelly  = max(0,kelly)*KELLY_FRACTION
    einsatz = round(kelly*100,2)
    return max(KELLY_MIN_EINSATZ,min(einsatz,KELLY_MAX_EINSATZ))

def get_wetter(country: str) -> dict:
    now = time.time()
    if country in _wetter_cache and now-_wetter_cache[country]["ts"] < WETTER_TTL:
        return _wetter_cache[country]
    try:
        resp = requests.get(f"https://wttr.in/{country}?format=j1",timeout=6)
        if resp.status_code != 200:
            return {"wind":0,"regen":0,"ts":now}
        current = resp.json()["current_condition"][0]
        wind    = int(current.get("windspeedKmph",0))
        regen   = float(current.get("precipMM",0))
        result  = {"wind":wind,"regen":regen,"ts":now}
        _wetter_cache[country] = result
        return result
    except Exception:
        return {"wind":0,"regen":0,"ts":now}

def schlechtes_wetter(country: str) -> bool:
    w = get_wetter(country)
    return w["wind"] >= WETTER_WIND_GRENZE or w["regen"] >= WETTER_REGEN_GRENZE

def wetter_analyse(country: str) -> dict:
    w = get_wetter(country)
    wind  = w.get("wind",0)
    regen = w.get("regen",0)
    tipps = []
    info  = []
    if regen >= 5:
        tipps.extend(["unter_ecken","unter_tore"])
        info.append(f"🌧️ Starkregen ({regen}mm) → schwieriger Ball")
    elif regen >= 2:
        tipps.append("unter_ecken")
        info.append(f"🌦️ Regen ({regen}mm) → leicht reduzierte Ecken")
    if wind >= 40:
        tipps.extend(["unter_ecken","mehr_karten"])
        info.append(f"💨 Sturm ({wind}km/h) → Flanken kaum möglich")
    elif wind >= 30:
        tipps.append("unter_ecken")
        info.append(f"🌬️ Wind ({wind}km/h) → Flankenspiel erschwert")
    return {"tipps":tipps,"info":info,"wind":wind,"regen":regen,"schlecht":len(tipps)>0}

def liga_erlaubt(liga: str) -> bool:
    if liga not in liga_statistik:
        return True
    s   = liga_statistik[liga]
    ges = s["gewonnen"]+s["verloren"]
    if ges < LIGA_MIN_TIPPS:
        return True
    quote = s["gewonnen"]/ges
    if quote < LIGA_MIN_TREFFERQUOTE:
        print(f"  [Liga-Filter] {liga} gesperrt: {s['gewonnen']}/{ges} ({round(quote*100)}%)")
        return False
    return True

def _af_flach_match_zu_intern(m: dict) -> dict:
    """Wandelt einen 'flachen' Match-Eintrag um, wie ihn teams/matches.json und
    teams/head2head.json liefern (home_name/away_name/home_id/away_id/score/ht_score
    statt verschachtelter home{}/away{}/scores{} Objekte wie im Live-Feed)."""
    return {
        "id": str(m.get("id","")),
        "status": "FT",  # dies sind immer schon gespielte Spiele (Historie)
        "time": "FT",
        "home": {"id":str(m.get("home_id","")),"name":m.get("home_name","?")},
        "away": {"id":str(m.get("away_id","")),"name":m.get("away_name","?")},
        "competition": {"id":str((m.get("competition") or {}).get("id","")),
                         "name":(m.get("competition") or {}).get("name","?")},
        "country": {"name":"International"},
        "scores": {
            "score":    m.get("score","0 - 0") or "0 - 0",
            "ht_score": m.get("ht_score","") or "",
        },
    }

def _af_team_history(team_id: str, anzahl: int) -> list:
    try:
        params = {"team_id":team_id,"number":anzahl}
        resp   = api_get_with_retry(f"{LS_BASE}/teams/matches.json",params)
        raw    = resp.json().get("data",[]) or []
        return [_af_flach_match_zu_intern(m) for m in raw]
    except Exception as e:
        print(f"  [Team-History] Fehler ({team_id}): {e}")
        return []

def get_team_ecken_avg(team_id: str):
    now = time.time()
    if team_id in _ecken_avg_cache:
        cached = _ecken_avg_cache[team_id]
        if now-cached["ts"] < ECKEN_AVG_TTL:
            return cached["avg"]
    try:
        matches     = _af_team_history(team_id,ECKEN_HISTORY_SPIELE)
        ecken_liste = []
        for m in matches:
            mid = m.get("id","")
            if not mid:
                continue
            try:
                stats = ls_get_statistiken(mid)
                ecken = stats["corners_home"]+stats["corners_away"]
                if ecken > 0:
                    ecken_liste.append(ecken)
            except Exception:
                continue
        if len(ecken_liste) >= 3:
            avg = round(sum(ecken_liste)/len(ecken_liste),1)
            _ecken_avg_cache[team_id] = {"avg":avg,"ts":now}
            return avg
        return None
    except Exception as e:
        print(f"  [Ecken-Avg] Fehler für Team {team_id}: {e}")
        return None

def ecken_tipp_sinnvoll(game: dict, grenze: float) -> bool:
    home_id = str((game.get("home") or {}).get("id",""))
    away_id = str((game.get("away") or {}).get("id",""))
    if not home_id or not away_id:
        return True
    avg_home = get_team_ecken_avg(home_id)
    avg_away = get_team_ecken_avg(away_id)
    if avg_home is None or avg_away is None:
        return True
    erwartet = round((avg_home+avg_away)/2,1)
    sinnvoll = grenze > erwartet+ECKEN_TOLERANZ
    print(f"  [Ecken-Avg] Heim ⌀{avg_home} | Gast ⌀{avg_away} | Erwartet: {erwartet} | Grenze: {grenze} | Sinnvoll: {sinnvoll}")
    return sinnvoll

_h2h_cache = {}
H2H_TTL    = 3600

def get_h2h_daten(team1_id: str, team2_id: str) -> list:
    key = f"{min(team1_id,team2_id)}_{max(team1_id,team2_id)}"
    now = time.time()
    if key in _h2h_cache and now-_h2h_cache[key]["ts"] < H2H_TTL:
        return _h2h_cache[key]["data"]
    try:
        params  = {"team1_id":team1_id,"team2_id":team2_id}
        resp    = api_get_with_retry(f"{LS_BASE}/teams/head2head.json",params)
        raw     = (resp.json().get("data") or {}).get("h2h",[]) or []
        raw     = raw[:H2H_MAX_SPIELE]
        matches = [_af_flach_match_zu_intern(m) for m in raw]
        _h2h_cache[key] = {"data":matches,"ts":now}
        return matches
    except Exception as e:
        print(f"  [H2H] Fehler: {e}")
        return []

def analysiere_h2h_tore(matches: list):
    if len(matches) < H2H_MIN_SPIELE:
        return None
    hz1_liste = []
    vz_liste  = []
    for m in matches:
        score = (m.get("scores") or {}).get("score","")
        ht    = (m.get("scores") or {}).get("ht_score","")
        h,a   = parse_score(score)
        if h+a == 0 and not score:
            continue
        vz_liste.append(h+a)
        if ht:
            hh,ha = parse_score(ht)
            hz1_liste.append(hh+ha)
    if len(vz_liste) < H2H_MIN_SPIELE:
        return None
    return {
        "avg_vz":     round(sum(vz_liste)/len(vz_liste),2),
        "avg_hz1":    round(sum(hz1_liste)/len(hz1_liste),2) if len(hz1_liste) >= H2H_MIN_SPIELE else None,
        "spiele":     len(vz_liste),
        "hz1_spiele": len(hz1_liste),
    }

def tipp_aus_avg(avg: float, ueber_grenze: float, unter_grenze: float):
    if avg > ueber_grenze:
        linie = int(avg-0.5)+0.5
        return ("über",linie)
    elif avg < unter_grenze:
        linie = int(avg)+0.5
        return ("unter",linie)
    return None

def berechne_konfidenz(typ,liga,quote,h2h_spiele=0,wetter_schlecht=False,bookmaker_anzahl=0,form_uebereinstimmung=True):
    score = 5
    if liga in liga_statistik:
        s   = liga_statistik[liga]
        ges = s["gewonnen"]+s["verloren"]
        if ges >= 5:
            hit = s["gewonnen"]/ges
            if hit >= 0.65:   score += 2
            elif hit >= 0.55: score += 1
            elif hit <= 0.35: score -= 2
            elif hit <= 0.45: score -= 1
    if h2h_spiele >= 8:       score += 1
    elif 0 < h2h_spiele < 3:  score -= 1
    if not form_uebereinstimmung: score -= 2
    if wetter_schlecht:           score -= 1
    if quote:
        if quote >= 2.0:   score += 1
        elif quote < 1.35: score -= 1
    if bookmaker_anzahl >= 4:   score += 1
    elif bookmaker_anzahl == 1: score -= 1
    return max(1,min(10,score))

def konfidenz_emoji(score):
    if score >= 8: return "🟢"
    if score >= 6: return "🟡"
    if score >= 4: return "🟠"
    return "🔴"

def get_quote_details(match_id, linie: float = None, richtung: str = None, nur_live: bool = True):
    """v59.1: Quoten jetzt über API-Football (fixture-basiert), nicht mehr the-odds-api.com.
    v59.18 FIX: linie/richtung ergänzt. v59.24 FIX: nur_live siehe get_quote."""
    return af_odds_details(match_id, linie=linie, richtung=richtung, nur_live=nur_live)

def get_odds_vergleich(match_id) -> str:
    """v59.1: Quoten jetzt über API-Football (fixture-basiert), nicht mehr the-odds-api.com."""
    return af_odds_vergleich_text(match_id)

def gegentipp_registrieren(match_id,markt,richtung,bot):
    aktive_tipps.setdefault(match_id,[])
    aktive_tipps[match_id].append({"markt":markt,"richtung":richtung,"bot":bot})

def gegentipp_check(match_id,markt,richtung,bot):
    for tipp in aktive_tipps.get(match_id,[]):
        if tipp["markt"] == markt and tipp["richtung"] != richtung:
            print(f"  [Gegentipp] Widerspruch: {bot} ({richtung}) vs {tipp['bot']} ({tipp['richtung']})")
            return False
    return True

def get_team_saisonform(team_id):
    try:
        matches = _af_team_history(team_id,5)
        tore    = []
        for m in matches:
            score = (m.get("scores") or {}).get("score","")
            h,a   = parse_score(score)
            if h+a > 0 or score:
                tore.append(h+a)
        return round(sum(tore)/len(tore),2) if len(tore) >= 3 else None
    except Exception:
        return None

def form_stimmt_ueberein(home_id,away_id,h2h_avg,richtung):
    form_home = get_team_saisonform(home_id)
    form_away = get_team_saisonform(away_id)
    if form_home is None or form_away is None:
        return True
    form_avg  = (form_home+form_away)/2
    h2h_hoch  = h2h_avg > 2.5
    form_hoch = form_avg > 2.5
    if richtung == "über":
        ok = h2h_hoch and form_hoch
    else:
        ok = (not h2h_hoch) and (not form_hoch)
    if not ok:
        print(f"  [Form] H2H {h2h_avg} vs Form {form_avg:.1f} weichen ab")
    return ok

def clv_auswerten(spiel):
    einstieg = spiel.get("quote")
    if not einstieg:
        return ""
    try:
        details = get_quote_details(spiel.get("match_id"))
        schluss = details.get("avg_quote")
        if not schluss:
            return ""
        diff = round((einstieg-schluss)/schluss,3)
        if diff > CLV_WARNSCHWELLE:
            return f"\n📊 CLV: Guter Einstieg! {einstieg} -> Schluss {schluss} (+{round(diff*100,1)}%)"
        elif diff < -CLV_WARNSCHWELLE:
            return f"\n📊 CLV: Quote gesunken: {einstieg} -> Schluss {schluss} ({round(diff*100,1)}%)"
        return f"\n📊 CLV: Quote stabil: {einstieg} -> Schluss {schluss}"
    except Exception:
        return ""

def get_standings(league_id: str) -> list:
    now = time.time()
    lid = str(league_id)
    if lid in _standings_cache and now-_standings_cache[lid]["ts"] < STANDINGS_TTL:
        return _standings_cache[lid]["data"]
    try:
        params = {"competition_id":lid,"include_form":1}
        resp   = api_get_with_retry(f"{LS_BASE}/competitions/table.json",params)
        resp_data = resp.json().get("data") or {}
        # v60: livescore-api.com gruppiert Tabellen in stages -> groups -> standings
        # (statt einer flachen Liste bei API-Football) – hier alles wieder flach ziehen
        flach = []
        for stage in resp_data.get("stages",[]) or []:
            for gruppe in stage.get("groups",[]) or []:
                flach.extend(gruppe.get("standings",[]) or [])
        standings = []
        for t in flach:
            team  = t.get("team") or {}
            form  = t.get("form") or []
            standings.append({
                "team_id":                  str(team.get("id","")),
                "overall_league_position":  t.get("rank",0),
                "overall_league_PTS":       t.get("points",0),
                "overall_league_payed":     t.get("matches",0),
                "overall_league_W":         t.get("won",0),
                "overall_league_D":         t.get("drawn",0),
                "overall_league_L":         t.get("lost",0),
                "overall_league_GF":        t.get("goals_scored",0) or 0,
                "overall_league_GA":        t.get("goals_conceded",0) or 0,
                "league_team_form":         "".join(form) if isinstance(form,list) else (form or ""),
            })
        _standings_cache[lid] = {"data":standings,"ts":now}
        return standings
    except Exception as e:
        print(f"  [Standings] Fehler ({league_id}): {e}")
        return []

def get_team_standing(league_id: str, team_id: str):
    standings = get_standings(league_id)
    for t in standings:
        if str(t.get("team_id","")) == str(team_id):
            gf = _safe_int(t.get("overall_league_GF",0))
            ga = _safe_int(t.get("overall_league_GA",0))
            gp = _safe_int(t.get("overall_league_payed",t.get("overall_league_played",0)))
            return {
                "position": _safe_int(t.get("overall_league_position",t.get("position",0))),
                "punkte":   _safe_int(t.get("overall_league_PTS",0)),
                "gespielt": gp,"siege":_safe_int(t.get("overall_league_W",0)),
                "unent":    _safe_int(t.get("overall_league_D",0)),
                "niederl":  _safe_int(t.get("overall_league_L",0)),
                "tore_f":   gf,"tore_g":ga,
                "tore_avg": round(gf/gp,1) if gp > 0 else 0,
                "form":     t.get("league_team_form",""),
            }
    return None

def form_zu_emojis(form: str) -> str:
    result = ""
    for c in (form or "")[-5:]:
        if c == "W":   result += "🟢"
        elif c == "D": result += "🟡"
        elif c == "L": result += "🔴"
    return result or "–"

def baue_analyse_text(home,away,home_id,away_id,league_id,extra=None) -> str:
    h_stand = get_team_standing(league_id,home_id)
    a_stand = get_team_standing(league_id,away_id)
    zeilen  = []
    if h_stand:
        hf = form_zu_emojis(h_stand["form"])
        zeilen.append(f"🏠 {home}: Platz {h_stand['position']} | {hf} | ⚽ {h_stand['tore_avg']}/Spiel")
    if a_stand:
        af = form_zu_emojis(a_stand["form"])
        zeilen.append(f"✈️ {away}: Platz {a_stand['position']} | {af} | ⚽ {a_stand['tore_avg']}/Spiel")
    if extra:
        for k,v in extra.items():
            zeilen.append(f"{k}: {v}")
    return "\n".join(zeilen)

TYP_NAMEN = {
    "ecken":"Ecken Unter","ecken_over":"Ecken Über",
    "karten":"Karten Über 5","torwart":"Mind. 1 Tor",
    "druck":"Druck/Ecken Dominanz","comeback":"Beide Teams treffen",
    "torflut":"Torreich","rotkarte":"Überzahl Tor",
    "hz1tore":"HZ1 Tore","vztore":"Vollzeit Tore","hz2tore":"HZ2 Tore",
    "xg":"xG Signal","value":"Value Bet",
}

def claude_tipp_review(home,away,typ,analyse,liga=""):
    if not ANTHROPIC_API_KEY or not ANTHROPIC_API_KEY.strip():
        return True,""
    if not claude_budget_verfuegbar(liga):
        return True,""
    try:
        typ_name = TYP_NAMEN.get(typ,typ)
        prompt = (f"Du bist ein erfahrener Sportwetten-Analyst. Analysiere auf Deutsch:\n\n"
                  f"Spiel: {home} vs {away}\nTipp: {typ_name}\nLive-Daten:\n{analyse}\n\n"
                  f"Antworte NUR in diesem Format:\n"
                  f"EMPFOHLEN: [2-3 präzise Sätze]\noder\nSKEPTISCH: [2-3 Sätze]")
        resp = requests.post("https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":ANTHROPIC_API_KEY,"anthropic-version":"2023-06-01"},
            json={"model":"claude-sonnet-4-20250514","max_tokens":200,"messages":[{"role":"user","content":prompt}]},
            timeout=15)
        if resp.status_code != 200:
            return True,""
        text        = resp.json().get("content",[{}])[0].get("text","").strip()
        empfohlen   = text.startswith("EMPFOHLEN")
        begruendung = text.replace("EMPFOHLEN:","").replace("SKEPTISCH:","").strip()
        claude_budget_erhoehen()
        return empfohlen,begruendung
    except Exception as e:
        print(f"  [Claude] Review Fehler: {e}")
        return True,""

def claude_live_begruendung(home,away,typ,stats,score,minute) -> str:
    typ_name = TYP_NAMEN.get(typ,typ)
    fallbacks = {
        "druck":    f"{home} dominiert mit {stats.get('corners_home',0)} Ecken klar.",
        "comeback": f"Trotz Rückstand dominiert das Team die Statistiken.",
        "torwart":  f"Trotz {stats.get('shots_on_target_home',0)+stats.get('shots_on_target_away',0)} Schüssen ist das Spiel noch torlos.",
        "torflut":  f"Bereits {score} nach der Halbzeit – weiteres Tor sehr wahrscheinlich.",
        "rotkarte": f"Die numerische Überzahl verschafft klaren Vorteil.",
        "ecken":    f"Nur {stats.get('corners_home',0)+stats.get('corners_away',0)} Ecken in der ersten Hälfte.",
        "karten":   f"Bereits {stats.get('corners_home',0)} Karten in Minute {minute}.",
    }
    return fallbacks.get(typ,"")

def dynamischer_sleep(matches=None):
    try:
        for m in (matches or []):
            minute = _safe_int(m.get("time",0))
            if 43 <= minute <= 47 or 87 <= minute <= 93:
                time.sleep(60)
                return
    except Exception:
        pass
    time.sleep(FUSSBALL_INTERVAL*60)

def signal_eintragen(match_id,typ,home,away,liga,hz1_wert,grenze,quote,einsatz):
    beobachtete_spiele_speichern()
    signal_log.append({
        "match_id":match_id,"typ":typ,"home":home,"away":away,"liga":liga,
        "hz1_wert":hz1_wert,"grenze":grenze,"quote":quote,"einsatz":einsatz,
        "zeit":de_now().strftime("%Y-%m-%d %H:%M"),"gewonnen":None
    })
    signal_log_speichern()

def signal_auswertung_aktualisieren(match_id,gewonnen):
    for s in reversed(signal_log):
        if s["match_id"] == match_id and s["gewonnen"] is None:
            s["gewonnen"] = gewonnen
            signal_log_speichern()
            break

def beobachtete_spiele_speichern():
    import json
    try:
        with open(BEOBACHTETE_DATEI,"w") as f:
            json.dump(beobachtete_spiele,f,indent=2)
    except Exception as e:
        print(f"  [Persistenz] Speicherfehler: {e}")

def beobachtete_spiele_laden():
    import json, os as _os
    if not _os.path.exists(BEOBACHTETE_DATEI):
        return
    try:
        with open(BEOBACHTETE_DATEI,"r") as f:
            data = json.load(f)
        beobachtete_spiele.update(data)
        print(f"  [Persistenz] {len(data)} beobachtete Spiele geladen")
    except Exception as e:
        print(f"  [Persistenz] Ladefehler: {e}")

def bankroll_laden() -> float:
    import json, os as _os
    if not _os.path.exists(BANKROLL_DATEI):
        bankroll_speichern(BANKROLL)
        return BANKROLL
    try:
        with open(BANKROLL_DATEI,"r") as f:
            return json.load(f).get("bankroll",BANKROLL)
    except Exception:
        return BANKROLL

def bankroll_speichern(betrag: float):
    import json
    try:
        with open(BANKROLL_DATEI,"w") as f:
            json.dump({"bankroll":round(betrag,2),"aktualisiert":de_now().strftime("%Y-%m-%d %H:%M")},f)
    except Exception as e:
        print(f"  [Bankroll] Speicherfehler: {e}")

def bankroll_aktualisieren(gewonnen: bool, einsatz: float, quote: float = None):
    br = bankroll_laden()
    if gewonnen and quote:
        gewinn = round(einsatz*(quote-1),2)
        br    += gewinn
        print(f"  [Bankroll] +{gewinn}€ | Neue Bankroll: {br}€")
    else:
        br -= einsatz
        print(f"  [Bankroll] -{einsatz}€ | Neue Bankroll: {br}€")
    bankroll_speichern(br)
    return br

def kelly_einsatz_bankroll(quote: float, typ: str) -> float:
    br  = bankroll_laden()
    ges = statistik[typ]["gewonnen"]+statistik[typ]["verloren"]
    if ges < 10 or not quote or quote <= 1.0:
        return min(round(br*0.02,2),KELLY_MAX_EINSATZ)
    p      = statistik[typ]["gewonnen"]/ges
    b      = quote-1.0
    kelly  = max(0,(b*p-(1-p))/b)*KELLY_FRACTION
    einsatz = round(kelly*br,2)
    return max(KELLY_MIN_EINSATZ,min(einsatz,KELLY_MAX_EINSATZ))

_api_monitor = {"heute":0,"datum":""}

def api_monitor_increment():
    _api_monitor_tag_pruefen()
    _api_monitor["heute"] += 1
    if _api_monitor["heute"] % 3000 == 0:
        check_rate_limit_warnung()
    if _api_monitor["heute"] >= API_DAILY_HARD_STOP and not _api_hard_stop_alarm_gesendet:
        globals()["_api_hard_stop_alarm_gesendet"] = True
        send_telegram(
            f"🛑 <b>API Tageslimit-Schutz aktiv!</b>\n"
            f"📊 {_api_monitor['heute']:,}/{API_DAILY_LIMIT:,} Livescore-Calls heute\n"
            f"⛔ Weitere Livescore-Anfragen sind bis Mitternacht blockiert.\n🕐 {jetzt()} Uhr"
        )

def api_monitor_bericht() -> str:
    n   = _api_monitor["heute"]
    pct = round(n/API_DAILY_LIMIT*100,1)
    bar = "█"*int(pct/5)+"░"*(20-int(pct/5))
    status = "🛑 GESPERRT" if n >= API_DAILY_HARD_STOP else ("⚠️" if pct>=80 else "✅")
    return f"📡 API heute: <b>{n:,}</b>/{API_DAILY_LIMIT:,} ({pct}%) {status}\n{bar}"

beobachtete_spiele_multi = {}

def berechne_dynamische_grenzen(h2h_avg: float, form_avg=None, typ: str = "vz") -> tuple:
    if form_avg is not None:
        gewichteter_avg = round(h2h_avg*0.7+form_avg*0.3,2)
    else:
        gewichteter_avg = h2h_avg
    if typ == "hz1":
        ueber_grenze = max(0.9,min(1.4,1.0+(gewichteter_avg*0.1)))
        unter_grenze = max(0.4,min(0.8,0.5+(gewichteter_avg*0.05)))
    else:
        ueber_grenze = max(2.2,min(3.0,2.3+(gewichteter_avg*0.15)))
        unter_grenze = max(1.3,min(2.0,1.5+(gewichteter_avg*0.1)))
    return round(ueber_grenze,2),round(unter_grenze,2)

_lineup_cache = {}
LINEUP_TTL    = 1800

def get_team_lineup(match_id: str) -> dict:
    now = time.time()
    if match_id in _lineup_cache and now-_lineup_cache[match_id]["ts"] < LINEUP_TTL:
        return _lineup_cache[match_id]["data"]
    try:
        params  = {"match_id":match_id}
        resp    = api_get_with_retry(f"{LS_BASE}/matches/lineups.json",params)
        lineup  = (resp.json().get("data") or {}).get("lineup") or {}
        data    = {}
        for seite in ("home","away"):
            spieler = (lineup.get(seite) or {}).get("players",[]) or []
            # substitution "0" = Startelf, "1" = Bank (siehe livescore-api.com Doku)
            startelf = [p for p in spieler if str(p.get("substitution","0")) == "0"]
            data[seite] = {"starting_eleven": startelf}
        _lineup_cache[match_id] = {"data":data,"ts":now}
        return data
    except Exception:
        return {}

def verletzungs_check(match_id: str, home: str, away: str) -> str:
    try:
        lineup      = get_team_lineup(match_id)
        if not lineup:
            return ""
        home_lineup = lineup.get("home",{})
        away_lineup = lineup.get("away",{})
        home_count  = len(home_lineup.get("starting_eleven",[]) or [])
        away_count  = len(away_lineup.get("starting_eleven",[]) or [])
        warnungen   = []
        if home_count > 0 and home_count < 11:
            warnungen.append(f"⚠️ {home}: Nur {home_count}/11 Spieler gelistet")
        if away_count > 0 and away_count < 11:
            warnungen.append(f"⚠️ {away}: Nur {away_count}/11 Spieler gelistet")
        return "\n".join(warnungen)
    except Exception:
        return ""

def multi_signal_check(match_id: str, aktueller_bot: str) -> int:
    if match_id in beobachtete_spiele_multi:
        for typ,s in beobachtete_spiele_multi[match_id].items():
            erster_bot = s.get("bot","")
            if erster_bot and erster_bot != aktueller_bot:
                print(f"  [Multi-Signal] {match_id}: {erster_bot} + {aktueller_bot} → +{MULTI_SIGNAL_BONUS}")
                return MULTI_SIGNAL_BONUS
    return 0

def auswertung_fallback_check():
    jetzt_ts = time.time()
    with _tracker_lock:
        for key,sig in list(_signal_tracker.items()):
            if sig.get("status") != "offen":
                continue
            signal_zeit = sig.get("signal_zeit",jetzt_ts)
            stunden = (jetzt_ts-signal_zeit)/3600
            if stunden > MAX_BEOBACHTUNG_STUNDEN:
                _signal_tracker[key]["status"] = "abgelaufen"
                print(f"  [Fallback] {sig.get('home','?')} vs {sig.get('away','?')} nach {stunden:.1f}h abgelaufen")

def beobachtung_hinzufuegen(match_id: str, spiel: dict):
    typ = spiel.get("typ","unbekannt")
    if match_id not in beobachtete_spiele_multi:
        beobachtete_spiele_multi[match_id] = {}
    beobachtete_spiele_multi[match_id][typ] = spiel
    beobachtete_spiele[f"{match_id}_{typ}"] = spiel
    beobachtete_spiele_speichern()
    tracker_signal_hinzufuegen(match_id,spiel)
    notified_sets_speichern()
    kombi_signal_check(match_id)

def kombi_signal_check(match_id: str):
    if match_id in kombi_gesendet:
        return
    typen = set(beobachtete_spiele_multi.get(match_id,{}).keys())
    if len(typen) < 2:
        return
    for kombi_typen,kombi_name in KOMBI_SIGNAL_TYPEN.items():
        if kombi_typen.issubset(typen):
            spiel     = next(iter(beobachtete_spiele_multi[match_id].values()))
            home      = spiel.get("home","?")
            away      = spiel.get("away","?")
            alle_tipps = []
            for t in kombi_typen:
                s = beobachtete_spiele_multi[match_id].get(t,{})
                if s.get("richtung") and s.get("linie"):
                    alle_tipps.append(f"{t.upper()}: {s['richtung']} {s['linie']}")
                elif s.get("typ"):
                    alle_tipps.append(t.upper())
            msg = (f"⚡ <b>KOMBI-SIGNAL! – {kombi_name}</b>\n"
                   f"━━━━━━━━━━━━━━━━━━━━\n"
                   f"📌 <b>{home} vs {away}</b>\n"
                   f"🤖 Mehrere Bots tippen auf dieses Spiel:\n"
                   f"{'  '.join([f'✅ {t}' for t in alle_tipps])}\n"
                   f"━━━━━━━━━━━━━━━━━━━━\n"
                   f"🎯 Erhöhtes Vertrauen in dieses Spiel!\n🕐 {jetzt()} Uhr")
            send_telegram(msg)
            kombi_gesendet.add(match_id)
            break

def prüfe_ecken_verfuegbar(match_id) -> bool:
    """v59.1: Prüft jetzt über API-Football (fixture-basiert), nicht mehr the-odds-api.com."""
    return af_odds_ecken_verfuegbar(match_id)

def whitelist_check(liga: str, home: str = "", away: str = "") -> bool:
    if not _whitelist.get("aktiv"):
        return True
    ligen = [l.lower() for l in _whitelist.get("ligen",[])]
    teams = [t.lower() for t in _whitelist.get("teams",[])]
    if ligen and liga.lower() not in ligen:
        return False
    if teams and home.lower() not in teams and away.lower() not in teams:
        return False
    return True

def berechne_ev_score(konfidenz: int, quote: float) -> dict:
    if not quote or quote <= 1.0:
        return {"ev":0,"label":"–","empfohlen":False,"ev_pct":0}
    p   = konfidenz/10
    ev  = round((p*quote)-1,3)
    ev_pct = round(ev*100,1)
    if ev >= 0.15:
        label = f"💎 Sehr gut ({ev_pct}%)"; empfohlen = True
    elif ev >= 0.08:
        label = f"✅ Gut ({ev_pct}%)"; empfohlen = True
    elif ev >= 0.0:
        label = f"🟡 Grenzwertig ({ev_pct}%)"; empfohlen = False
    else:
        label = f"❌ Negativ ({ev_pct}%)"; empfohlen = False
    return {"ev":ev,"ev_pct":ev_pct,"label":label,"empfohlen":empfohlen}


# ============================================================
#  DISCORD EMBEDS
# ============================================================

FARBE_ECKEN               = 0xF4A300
FARBE_ECKEN_OVER          = 0x9B59B6
FARBE_KARTEN              = 0xE74C3C
FARBE_TORWART             = 0x1ABC9C
FARBE_DRUCK               = 0x3498DB
FARBE_COMEBACK            = 0xE67E22
FARBE_TORFLUT             = 0xFF6B6B
FARBE_ROTKARTE            = 0xC0392B
FARBE_HZ1TORE             = 0x27AE60
FARBE_VZTORE              = 0x8E44AD
FARBE_AUSWERTUNG_GEWONNEN = 0x2ECC71
FARBE_AUSWERTUNG_VERLOREN = 0xE74C3C

def einsatz_empfehlung_text(einsatz) -> str:
    """v59: Formatiert die empfohlene Einsatzhöhe inkl. Anteil an der aktuellen Bankroll für Discord-Embeds."""
    if not einsatz:
        return "–"
    try:
        br  = bankroll_laden()
        pct = round(einsatz/br*100,1) if br else 0
        return f"**{einsatz}€** (≈{pct}% der Bankroll)"
    except Exception:
        return f"**{einsatz}€**"

def discord_ecken_tipp(home,away,comp,country,score,corners_home,corners_away,corners,grenze,quote,einsatz=None,konfidenz=None):
    qt = f"\n💶 **Quote:** {quote}" if quote else ""
    return {
        "title":"📐 Ecken Tipp","color":FARBE_ECKEN,
        "fields":[
            {"name":"🏆 Liga","value":f"{comp}","inline":True},
            {"name":"🌍 Land","value":f"{country}","inline":True},
            {"name":"📊 Halbzeitstand","value":f"**{score}**","inline":True},
            {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
            {"name":"📐 Ecken zur Halbzeit","value":f"🔵 {home}: **{corners_home}**\n🔴 {away}: **{corners_away}**\n📊 Gesamt: **{corners}**","inline":False},
            {"name":"🎯 Empfehlung","value":f"Unter **{grenze} Ecken** (Gesamtspiel){qt}","inline":False},
            {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":False},
            {"name":"🎯 Konfidenz","value": (f"{konfidenz_emoji(konfidenz)} **{konfidenz}/10**" if konfidenz is not None else "–"),"inline":False},
        ],
        "footer":{"text":f"Ecken-Bot • {heute()} {jetzt()}"},
    }

def discord_torwart_tipp(home,away,comp,country,shots_home,shots_away,saves_h,saves_a,poss_h,poss_a,min_text,quote,einsatz=None,konfidenz=None):
    qt = f"\n💶 **Quote:** {quote}" if quote else ""
    return {
        "title":"🧤 Torwart Alarm","color":FARBE_TORWART,
        "fields":[
            {"name":"🏆 Liga","value":f"{comp}","inline":True},
            {"name":"🌍 Land","value":f"{country}","inline":True},
            {"name":"📊 Stand","value":f"**0:0** | {min_text}","inline":True},
            {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
            {"name":"🎯 Schüsse aufs Tor","value":f"🔵 {home}: **{shots_home}**\n🔴 {away}: **{shots_away}**\n📊 Gesamt: **{shots_home+shots_away}**","inline":True},
            {"name":"🧤 Paraden","value":f"🔵 {saves_h} | 🔴 {saves_a}","inline":True},
            {"name":"⚽ Ballbesitz","value":f"{poss_h}% | {poss_a}%","inline":True},
            {"name":"🎯 Empfehlung","value":f"Mindestens **1 Tor** fällt noch{qt}","inline":False},
            {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":False},
            {"name":"🎯 Konfidenz","value": (f"{konfidenz_emoji(konfidenz)} **{konfidenz}/10**" if konfidenz is not None else "–"),"inline":False},
        ],
        "footer":{"text":f"Torwart-Bot • {heute()} {jetzt()}"},
    }

def discord_druck_tipp(home,away,comp,country,score,minute,druck_team,ecken_stark,ecken_schwach,quote,einsatz=None,konfidenz=None):
    qt = f"\n💶 **Quote:** {quote}" if quote else ""
    return {
        "title":"🔥 Druck Signal","color":FARBE_DRUCK,
        "fields":[
            {"name":"🏆 Liga","value":f"{comp}","inline":True},
            {"name":"🌍 Land","value":f"{country}","inline":True},
            {"name":"📊 Stand","value":f"**{score}** | Min. **{minute}'**","inline":True},
            {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
            {"name":"🔥 Dominantes Team","value":f"**{druck_team}**","inline":False},
            {"name":"📐 Ecken","value":f"**{ecken_stark}** : {ecken_schwach}","inline":True},
            {"name":"🎯 Empfehlung","value":f"Nächste Ecke / Tor für **{druck_team}**{qt}","inline":False},
            {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":False},
            {"name":"🎯 Konfidenz","value": (f"{konfidenz_emoji(konfidenz)} **{konfidenz}/10**" if konfidenz is not None else "–"),"inline":False},
        ],
        "footer":{"text":f"Druck-Bot • {heute()} {jetzt()}"},
    }

def discord_comeback_tipp(home,away,comp,country,score,minute,rueckliegend,fuehrend,shots_r,shots_f,poss_r,quote,einsatz=None,konfidenz=None):
    qt = f"\n💶 **Quote:** {quote}" if quote else ""
    return {
        "title":"🔄 Comeback Signal","color":FARBE_COMEBACK,
        "fields":[
            {"name":"🏆 Liga","value":f"{comp}","inline":True},
            {"name":"🌍 Land","value":f"{country}","inline":True},
            {"name":"📊 Stand","value":f"**{score}** | Min. **{minute}'**","inline":True},
            {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
            {"name":"📉 Rückliegend","value":f"**{rueckliegend}**","inline":True},
            {"name":"📈 Führend","value":f"**{fuehrend}**","inline":True},
            {"name":"🎯 Schüsse Rückliegend","value":f"**{shots_r}** | Gegner: {shots_f}","inline":True},
            {"name":"⚽ Ballbesitz","value":f"**{poss_r}%**","inline":True},
            {"name":"🎯 Empfehlung","value":f"Beide Teams treffen (Comeback){qt}","inline":False},
            {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":False},
            {"name":"🎯 Konfidenz","value": (f"{konfidenz_emoji(konfidenz)} **{konfidenz}/10**" if konfidenz is not None else "–"),"inline":False},
        ],
        "footer":{"text":f"Comeback-Bot • {heute()} {jetzt()}"},
    }

def discord_torflut_tipp(home,away,comp,country,score_hz1,tore_hz1,grenze,quote,shots_ges=0,poss_h="?",poss_a="?",einsatz=None,konfidenz=None):
    qt = f"\n💶 **Quote:** {quote}" if quote else ""
    return {
        "title":"🌊 Torflut Signal","color":FARBE_TORFLUT,
        "fields":[
            {"name":"🏆 Liga","value":f"{comp}","inline":True},
            {"name":"🌍 Land","value":f"{country}","inline":True},
            {"name":"📊 Halbzeitstand","value":f"**{score_hz1}**","inline":True},
            {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
            {"name":"⚽ Tore HZ1","value":f"**{tore_hz1}** Tore","inline":True},
            {"name":"🎯 Schüsse gesamt","value":f"**{shots_ges}**","inline":True},
            {"name":"⚽ Ballbesitz","value":f"{poss_h}% | {poss_a}%","inline":True},
            {"name":"🎯 Empfehlung","value":f"Über **{grenze} Tore** im Gesamtspiel{qt}","inline":False},
            {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":False},
            {"name":"🎯 Konfidenz","value": (f"{konfidenz_emoji(konfidenz)} **{konfidenz}/10**" if konfidenz is not None else "–"),"inline":False},
        ],
        "footer":{"text":f"Torflut-Bot • {heute()} {jetzt()}"},
    }

def discord_hz1tore_tipp(home,away,comp,country,richtung,linie,avg_hz1,spiele,quote,einsatz=None,konfidenz=None):
    qt = f"\n💶 **Quote:** {quote}" if quote else ""
    pfeil = "📈" if richtung == "über" else "📉"
    return {
        "title":f"🥅 HZ1-Tore Signal ({richtung.upper()} {linie})","color":FARBE_HZ1TORE,
        "fields":[
            {"name":"🏆 Liga","value":f"{comp} ({country})","inline":True},
            {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":True},
            {"name":"📊 H2H Basis","value":f"**{spiele}** Spiele","inline":True},
            {"name":f"{pfeil} H2H Ø HZ1-Tore","value":f"**{avg_hz1}** Tore","inline":True},
            {"name":"🎯 Empfehlung","value":f"**{richtung.capitalize()} {linie}** Tore (HZ1){qt}","inline":False},
            {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":False},
            {"name":"🎯 Konfidenz","value": (f"{konfidenz_emoji(konfidenz)} **{konfidenz}/10**" if konfidenz is not None else "–"),"inline":False},
        ],
        "footer":{"text":f"H2H-Analyse • {heute()} {jetzt()}"},
    }

def discord_vztore_tipp(home,away,comp,country,richtung,linie,avg_vz,spiele,quote,einsatz=None,konfidenz=None):
    qt = f"\n💶 **Quote:** {quote}" if quote else ""
    pfeil = "📈" if richtung == "über" else "📉"
    return {
        "title":f"🏆 Vollzeit-Tore Signal ({richtung.upper()} {linie})","color":FARBE_VZTORE,
        "fields":[
            {"name":"🏆 Liga","value":f"{comp} ({country})","inline":True},
            {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":True},
            {"name":"📊 H2H Basis","value":f"**{spiele}** Spiele","inline":True},
            {"name":f"{pfeil} H2H Ø VZ-Tore","value":f"**{avg_vz}** Tore","inline":True},
            {"name":"🎯 Empfehlung","value":f"**{richtung.capitalize()} {linie}** Tore (Vollzeit){qt}","inline":False},
            {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":False},
            {"name":"🎯 Konfidenz","value": (f"{konfidenz_emoji(konfidenz)} **{konfidenz}/10**" if konfidenz is not None else "–"),"inline":False},
        ],
        "footer":{"text":f"H2H-Analyse • {heute()} {jetzt()}"},
    }

def discord_auswertung(typ,home,away,gewonnen,details: dict):
    farbe = FARBE_AUSWERTUNG_GEWONNEN if gewonnen else FARBE_AUSWERTUNG_VERLOREN
    emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
    titel = {
        "ecken":"📐 Auswertung – Ecken Unter","ecken_over":"📐 Auswertung – Ecken Über",
        "karten":"🃏 Auswertung – Karten","torwart":"🧤 Auswertung – Torwart",
        "druck":"🔥 Auswertung – Druck","comeback":"🔄 Auswertung – Comeback",
        "torflut":"🌊 Auswertung – Torflut","rotkarte":"🟥 Auswertung – Rote Karte",
        "hz1tore":"🥅 Auswertung – HZ1 Tore","vztore":"🏆 Auswertung – Vollzeit Tore",
        "hz2tore":"⚡ Auswertung – HZ2 Tore",
    }.get(typ,"📊 Auswertung")
    felder = [{"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False}]
    for k,v in details.items():
        felder.append({"name":k,"value":v,"inline":True})
    felder.append({"name":"Ergebnis","value":f"**{emoji}**","inline":False})
    return {"title":titel,"color":farbe,"fields":felder,"footer":{"text":f"Auswertung • {heute()} {jetzt()}"}}


# ============================================================
#  STATISTIK & BERICHTE
# ============================================================

def update_statistik(typ,gewonnen,quote,liga=None,match_id=None):
    stunde = str(de_now().hour)
    emoji  = "✅" if gewonnen else "❌"
    print(f"  [Statistik] {emoji} {typ.upper()} | Quote: {quote} | Liga: {liga}")
    with _statistik_lock:
        if gewonnen:
            gewinn = round((quote-1)*EINSATZ,2) if quote else round(EINSATZ*0.7,2)
            statistik[typ]["gewonnen"]        += 1
            statistik[typ]["gewinn"]          += gewinn
            wochen_statistik[typ]["gewonnen"] += 1
            wochen_statistik[typ]["gewinn"]   += gewinn
            stunden_statistik[stunde]["gewonnen"] += 1
            if liga:
                liga_statistik.setdefault(liga,{"gewonnen":0,"verloren":0})
                liga_statistik[liga]["gewonnen"] += 1
        else:
            statistik[typ]["verloren"]        += 1
            statistik[typ]["gewinn"]          -= EINSATZ
            wochen_statistik[typ]["verloren"] += 1
            wochen_statistik[typ]["gewinn"]   -= EINSATZ
            stunden_statistik[stunde]["verloren"] += 1
            if liga:
                liga_statistik.setdefault(liga,{"gewonnen":0,"verloren":0})
                liga_statistik[liga]["verloren"] += 1
        if match_id:
            signal_auswertung_aktualisieren(match_id,gewonnen)
        global streak_aktuell,streak_beste
        if gewonnen:
            streak_aktuell = max(1,streak_aktuell+1) if streak_aktuell >= 0 else 1
            streak_beste   = max(streak_beste,streak_aktuell)
            if streak_aktuell >= 3:
                send_telegram(f"🔥 <b>{streak_aktuell} Tipps in Folge gewonnen!</b> 💪")
        else:
            streak_aktuell = min(-1,streak_aktuell-1) if streak_aktuell <= 0 else -1
            if streak_aktuell <= -3:
                send_telegram(f"⚠️ <b>{abs(streak_aktuell)} Tipps in Folge verloren.</b>")
        if quote and quote > 1.0:
            einsatz_wert = EINSATZ
            roi_gewinn   = round((quote-1)*einsatz_wert,2) if gewonnen else -einsatz_wert
            if "roi" not in statistik[typ]:
                statistik[typ]["roi"] = 0.0
            statistik[typ]["roi"] = round(statistik[typ].get("roi",0.0)+roi_gewinn,2)
    statistik_speichern()

def statistik_zeile(name,stat):
    gesamt = stat["gewonnen"]+stat["verloren"]
    if gesamt == 0:
        return f"{name}: Noch keine Tipps"
    pct    = round(stat["gewonnen"]/gesamt*100)
    gewinn = round(stat["gewinn"],2)
    emoji  = "📈" if gewinn >= 0 else "📉"
    return f"{name}: {stat['gewonnen']}/{gesamt} ({pct}%) {emoji} {'+' if gewinn>=0 else ''}{gewinn}€"

def bot_rangliste(discord: bool = False) -> str:
    """
    v59.12 FIX: Die Streak-Zeile nutzte immer HTML <b>-Tags. Da diese Funktion sowohl für
    Telegram (HTML) als auch für Discord-Embeds (Markdown) verwendet wird, zeigten die
    Discord-Embeds die <b>-Tags als reinen Text an (Discord rendert kein HTML). Jetzt steuert
    der discord-Parameter, welches Format genutzt wird.
    """
    bots = {
        "📐 Ecken Unter":  wochen_statistik["ecken"],
        "🧤 Torwart":       wochen_statistik["torwart"],
        "🔥 Druck":         wochen_statistik["druck"],
        "🔄 Comeback":      wochen_statistik["comeback"],
        "🌊 Torflut":       wochen_statistik["torflut"],
        "🥅 HZ1-Tore":     wochen_statistik["hz1tore"],
        "🏆 VZ-Tore":       wochen_statistik["vztore"],
        "⚡ HZ2-Tore":      wochen_statistik["hz2tore"],
    }
    rang = []
    for name,stat in bots.items():
        ges = stat["gewonnen"]+stat["verloren"]
        if ges == 0:
            continue
        pct = round(stat["gewonnen"]/ges*100)
        rang.append((name,stat["gewonnen"],ges,pct))
    rang.sort(key=lambda x:x[3],reverse=True)
    zeilen = []
    for i,(name,gw,ges,pct) in enumerate(rang,1):
        medal = ["🥇","🥈","🥉"][i-1] if i <= 3 else f"{i}."
        zeilen.append(f"{medal} {name}: {gw}/{ges} ({pct}%)")
    streak_emoji = "🔥" if streak_aktuell > 0 else "❄️"
    fett = "**" if discord else "<b>"
    fett_zu = "**" if discord else "</b>"
    streak_text = f"{streak_emoji} Streak: {fett}{abs(streak_aktuell)}x {'Gewinn' if streak_aktuell > 0 else 'Verlust'}{fett_zu} | Beste: {fett}{streak_beste}x{fett_zu}"
    return "\n".join(zeilen)+f"\n━━━━━━━━━━━━━━━━━━━━\n{streak_text}" if zeilen else "Noch keine Daten"

def send_tagesbericht():
    """
    v59 FIX (Tagesbericht):
    - Setzt jetzt den GLOBALEN Marker `tagesbericht_gesendet`, statt sich auf
      eine lokale Variable im Auswertungs-Thread zu verlassen -> überlebt
      Thread-/Prozess-Neustarts (siehe bot_auswertung_und_berichte).
    - Postet den Bericht jetzt zusätzlich als Discord-Embed in den
      Bilanz-Channel (Bankroll, Streak, Bot-Rangliste, Virtuelle Konten) –
      vorher ging der Tagesbericht ausschließlich an Telegram.
    """
    global tagesbericht_gesendet
    gw  = sum(statistik[t]["gewonnen"] for t in statistik)
    vl  = sum(statistik[t]["verloren"] for t in statistik)
    ges = gw+vl
    gn  = round(sum(statistik[t]["gewinn"] for t in statistik),2)
    pct = round(gw/ges*100) if ges else 0
    ei  = "📈" if gn >= 0 else "📉"
    msg = (f"📋 <b>Tagesbericht – {heute()}</b>\n"
           f"━━━━━━━━━━━━━━━━━━━━\n"
           f"✅ Gewonnen: <b>{gw}</b>\n❌ Verloren: <b>{vl}</b>\n"
           f"🎯 Trefferquote: <b>{pct}%</b>\n"
           f"{ei} Simulation: <b>{'+' if gn>=0 else ''}{gn}€</b>\n"
           f"━━━━━━━━━━━━━━━━━━━━\n📊 <b>Nach Wetttyp:</b>\n"
           f"📐 {statistik_zeile('Ecken Unter', statistik['ecken'])}\n"
           f"🧤 {statistik_zeile('Torwart',     statistik['torwart'])}\n"
           f"🔥 {statistik_zeile('Druck',        statistik['druck'])}\n"
           f"🔄 {statistik_zeile('Comeback',     statistik['comeback'])}\n"
           f"🌊 {statistik_zeile('Torflut',      statistik['torflut'])}\n"
           f"🥅 {statistik_zeile('HZ1-Tore',     statistik['hz1tore'])}\n"
           f"🏆 {statistik_zeile('VZ-Tore',       statistik['vztore'])}\n"
           f"⚡ {statistik_zeile('HZ2-Tore',      statistik['hz2tore'])}\n"
           f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
    br   = bankroll_laden()
    diff = round(br-BANKROLL,2)
    br_pfeil = "📈 +" if diff >= 0 else "📉 "
    msg += (f"\n━━━━━━━━━━━━━━━━━━━━\n"
            f"💰 <b>Bankroll:</b> {br}€ ({br_pfeil}{diff}€ seit Start)\n"
            f"{api_monitor_bericht()}")
    send_telegram(msg)

    # ── v59 NEU: Tagesbericht zusätzlich als Discord-Embed in den Bilanz-Channel ──
    try:
        rang  = bot_rangliste(discord=True)
        vk    = vk_status_text()
        farbe = 0x2ECC71 if gn >= 0 else 0xE74C3C
        # v59.9: Diagnose-Zähler – zeigt wie viele offene/nicht auswertbare Signale gerade
        # im Tracker stehen, damit sichtbar ist ob die Auswertung insgesamt "klemmt".
        with _tracker_lock:
            _offen_ct = sum(1 for s in _signal_tracker.values() if s.get("status")=="offen")
            _na_ct    = sum(1 for s in _signal_tracker.values() if s.get("status")=="nicht_auswertbar")
        tages_embed = {
            "title": f"📋 Tagesbericht – {heute()}", "color": farbe,
            "fields": [
                {"name":"✅ Gewonnen","value":f"**{gw}**","inline":True},
                {"name":"❌ Verloren","value":f"**{vl}**","inline":True},
                {"name":"🎯 Trefferquote","value":f"**{pct}%**","inline":True},
                {"name":"💰 Bankroll","value":f"**{br}€** ({br_pfeil}{diff}€)","inline":True},
                {"name":"🔥 Streak","value": (f"**Noch keine Serie** (Beste: {streak_beste}x)" if streak_aktuell==0 else f"**{abs(streak_aktuell)}x {'Gewinn' if streak_aktuell>0 else 'Verlust'}** (Beste: {streak_beste}x)"),"inline":True},
                {"name":"⏳ Offen / ❓ Nicht auswertbar","value":f"**{_offen_ct}** offen | **{_na_ct}** nicht auswertbar","inline":True},
                {"name":"🏆 Bot-Rangliste","value": rang or "Noch keine Daten","inline":False},
                {"name":"🏦 Virtuelle Konten","value": vk or "–","inline":False},
                {"name":"🎁 Willkommensbonus","value":"Zahl **100€** ein und du bekommst **50€ BettingLab-Bonus** obendrauf!","inline":False},
            ],
            "footer": {"text": f"BetlabLIVE • {heute()} {jetzt()}"},
        }
        send_discord_embed(DISCORD_WEBHOOK_BILANZ, tages_embed)
    except Exception as e:
        print(f"  [Tagesbericht] Discord-Post Fehler: {e}")

    tagesbericht_gesendet = de_now().date()
    for t in statistik:
        statistik[t] = {"gewonnen":0,"verloren":0,"gewinn":0.0}
    statistik_speichern()
    print(f"  [Bericht] Tagesbericht gesendet ({heute()})")

def send_wochenbericht():
    gw  = sum(wochen_statistik[t]["gewonnen"] for t in wochen_statistik)
    vl  = sum(wochen_statistik[t]["verloren"] for t in wochen_statistik)
    ges = gw+vl
    gn  = round(sum(wochen_statistik[t]["gewinn"] for t in wochen_statistik),2)
    pct = round(gw/ges*100) if ges else 0
    ei  = "📈" if gn >= 0 else "📉"
    msg = (f"📅 <b>Wochenbericht</b>\n"
           f"━━━━━━━━━━━━━━━━━━━━\n"
           f"✅ Gewonnen: <b>{gw}</b>\n❌ Verloren: <b>{vl}</b>\n"
           f"🎯 Trefferquote: <b>{pct}%</b>\n"
           f"{ei} Simulation: <b>{'+' if gn>=0 else ''}{gn}€</b>\n"
           f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
    send_telegram(msg)
    farbe_w = 0x2ECC71 if gn >= 0 else 0xE74C3C
    rang_w  = bot_rangliste(discord=True)
    woche_embed = {
        "title":f"📅 Wochenbericht – KW {de_now().isocalendar()[1]}","color":farbe_w,
        "fields":[
            {"name":"✅ Gewonnen","value":f"**{gw}**","inline":True},
            {"name":"❌ Verloren","value":f"**{vl}**","inline":True},
            {"name":"🎯 Trefferquote","value":f"**{pct}%**","inline":True},
            {"name":"🏆 Bot-Rangliste (Woche)","value":rang_w or "Keine Daten","inline":False},
        ],
        "footer":{"text":f"BetlabLIVE • KW {de_now().isocalendar()[1]} • {heute()}"},
    }
    send_discord_embed(DISCORD_WEBHOOK_BILANZ,woche_embed)
    for t in wochen_statistik:
        wochen_statistik[t] = {"gewonnen":0,"verloren":0,"gewinn":0.0}
    print(f"  [Bericht] Wochenbericht gesendet")

def claude_verloren_analyse(home: str, away: str, typ: str, details: str):
    if not ANTHROPIC_API_KEY or ANTHROPIC_API_KEY.startswith("ANTHROPIC"):
        return
    try:
        typ_namen = {
            "ecken":"Ecken Unter","torwart":"Mind. 1 Tor",
            "druck":"Druck Signal","comeback":"Comeback",
            "torflut":"Torflut","hz1tore":"HZ1 Tore","vztore":"Vollzeit Tore","hz2tore":"HZ2 Tore",
        }.get(typ,typ)
        prompt = (f"Ein Sportwetten-Tipp hat verloren. Analysiere kurz auf Deutsch warum:\n\n"
                  f"Spiel: {home} vs {away}\nTipp-Typ: {typ_namen}\nDetails: {details}\n\n"
                  f"Antworte in max. 2 Sätzen. Nur die Analyse, kein Kommentar.")
        resp = requests.post("https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":ANTHROPIC_API_KEY,"anthropic-version":"2023-06-01"},
            json={"model":"claude-sonnet-4-20250514","max_tokens":150,"messages":[{"role":"user","content":prompt}]},
            timeout=15)
        if resp.status_code != 200:
            return
        analyse = resp.json().get("content",[{}])[0].get("text","").strip()
        if analyse:
            msg = (f"🔍 <b>Verlust-Analyse ({typ_namen})</b>\n"
                   f"━━━━━━━━━━━━━━━━━━━━\n"
                   f"📌 {home} vs {away}\n🤖 {analyse}\n━━━━━━━━━━━━━━━━━━━━")
            send_telegram(msg)
    except Exception as e:
        print(f"  [Verlust-Analyse] Fehler: {e}")


# ============================================================
#  AUSWERTUNGSFUNKTIONEN
# ============================================================

FT_STATI_SET = {"FT","FINISHED","AET","PEN","FULL TIME","AFTER EXTRA TIME","PENALTIES","ENDED","FT.","AET."}

def ist_spiel_fertig(status: str, time_val: str = "") -> bool:
    s = str(status or "").upper().strip()
    t = str(time_val or "").upper().strip()
    if s in FT_STATI_SET:
        return True
    if t in {"FT","FULL TIME","AET","ENDED","FINISHED"}:
        return True
    return False

FT_STATI = FT_STATI_SET

def _hole_tore_via_events(match_id: str) -> tuple:
    try:
        events = ls_get_events(match_id)
        h = len([e for e in events if e.get("event") in ("Goal","goal","GOAL") and e.get("home_away") == "home"])
        a = len([e for e in events if e.get("event") in ("Goal","goal","GOAL") and e.get("home_away") == "away"])
        return h,a
    except Exception:
        return 0,0

def _hole_hz1_tore_via_events(match_id: str) -> int:
    try:
        events = ls_get_events(match_id)
        hz1 = [e for e in events if e.get("event") in ("Goal","goal","GOAL") and _safe_int(e.get("time",99)) <= 45]
        return len(hz1)
    except Exception:
        return -1

def auswertung_ecken(spiel,ft_result=None):
    match_id  = spiel["match_id"]
    hz1_ecken = spiel["hz1_ecken"]
    grenze    = hz1_ecken*2+1
    home,away,quote = spiel["home"],spiel["away"],spiel.get("quote")
    liga = spiel.get("competition",spiel.get("liga",""))
    try:
        result = ft_result or ls_get_match_result(match_id,home,away,liga)
        if not result:
            return None
        stats       = get_statistiken(match_id)
        total_ecken = stats["corners_home"]+stats["corners_away"]
        if total_ecken == 0 and hz1_ecken > 0:
            total_ecken = hz1_ecken
        gewonnen    = total_ecken < grenze
        update_statistik("ecken",gewonnen,quote,liga=liga,match_id=match_id)  # v59.17: liga ergänzt
        emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
        ql = f"💶 Quote: <b>{quote}</b> → Gewinn: <b>+{round((quote-1)*EINSATZ,2)}€</b>\n" if quote and gewonnen else ""
        return (f"📊 <b>Auswertung – Ecken Unter</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                f"📌 {home} vs {away}\n📐 Ecken HZ1: <b>{hz1_ecken}</b>\n"
                f"🎯 Tipp: Unter <b>{grenze}</b> Ecken\n📈 Tatsächlich: <b>{total_ecken}</b>\n{ql}"
                f"━━━━━━━━━━━━━━━━━━━━\n{emoji}\n🕐 {jetzt()} Uhr")
    except Exception as e:
        print(f"  [Auswertung] Ecken Fehler: {e}")
        return None

def auswertung_torwart(spiel,ft_result=None):
    match_id = spiel["match_id"]
    home,away,quote = spiel["home"],spiel["away"],spiel.get("quote")
    liga = spiel.get("competition",spiel.get("liga",""))
    try:
        result = ft_result or ls_get_match_result(match_id,home,away,liga)
        if not result:
            return None
        score = result.get("score","")
        h,a   = parse_score(score) if score else (0,0)
        if h == 0 and a == 0:
            try:
                events  = ls_get_events(match_id)
                tore_e  = len([e for e in events if e.get("event") in ("Goal","goal")])
                if tore_e > 0:
                    h,a   = 1,0
                    score = f"mind. {tore_e} Tor(e)"
            except Exception:
                pass
        tore     = h+a
        gewonnen = tore >= 1
        update_statistik("torwart",gewonnen,quote,liga=liga,match_id=match_id)  # v59.17: liga ergänzt
        emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
        ql = f"💶 Quote: <b>{quote}</b> → Gewinn: <b>+{round((quote-1)*EINSATZ,2)}€</b>\n" if quote and gewonnen else ""
        return (f"📊 <b>Auswertung – Torwart</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                f"📌 {home} vs {away}\n🎯 Tipp: Mind. 1 Tor\n"
                f"📈 Endstand: <b>{score}</b> ({tore} Tore)\n{ql}"
                f"━━━━━━━━━━━━━━━━━━━━\n{emoji}\n🕐 {jetzt()} Uhr")
    except Exception as e:
        print(f"  [Auswertung] Torwart Fehler: {e}")
        return None

def auswertung_druck(spiel,ft_result=None):
    """
    v59.23 FIX: Der Tipp lautet "Nächste Ecke / Tor für {druck_team}" – die Auswertung hat
    bisher AUSSCHLIESSLICH auf zusätzliche Tore geprüft und Ecken komplett ignoriert. Ein
    Team, das nach dem Signal 5 weitere Ecken aber kein Tor mehr geholt hat, wurde dadurch
    fälschlich als "verloren" gewertet, obwohl der Tipp (der ausdrücklich auch Ecken abdeckt)
    eigentlich gewonnen war. Jetzt zählt: Tor ODER mindestens 1 weitere Ecke für das Team.
    """
    match_id     = spiel["match_id"]
    druck_team   = spiel["druck_team"]
    score_signal = spiel.get("score_signal","")
    home,away,quote = spiel["home"],spiel["away"],spiel.get("quote")
    h_sig,a_sig  = parse_score(score_signal) if score_signal else (0,0)
    ecken_home_sig = spiel.get("ecken_home_signal")
    ecken_away_sig = spiel.get("ecken_away_signal")
    liga = spiel.get("competition",spiel.get("liga",""))
    try:
        result = ft_result or ls_get_match_result(match_id,home,away,liga)
        if not result:
            return None
        score = result.get("score","")
        h,a   = parse_score(score) if score else (0,0)
        if h == 0 and a == 0 and (h_sig > 0 or a_sig > 0):
            try:
                events = ls_get_events(match_id)
                h = len([e for e in events if e.get("event") in ("Goal","goal") and e.get("home_away") == "home"])
                a = len([e for e in events if e.get("event") in ("Goal","goal") and e.get("home_away") == "away"])
                score = f"{h} - {a}" if h+a > 0 else score_signal
            except Exception:
                h,a = h_sig,a_sig; score = score_signal
        if druck_team == home:
            tor_gewonnen = h > h_sig
        else:
            tor_gewonnen = a > a_sig
        if not score_signal:
            tor_gewonnen = (h > a) if druck_team == home else (a > h)

        # v59.23 NEU: Ecken-Zuwachs seit Signal zusätzlich prüfen
        ecken_gewonnen = False
        if ecken_home_sig is not None and ecken_away_sig is not None:
            try:
                stats_ft = get_statistiken(match_id)
                ecken_h_ft = stats_ft["corners_home"]
                ecken_a_ft = stats_ft["corners_away"]
                if druck_team == home:
                    ecken_gewonnen = ecken_h_ft > ecken_home_sig
                else:
                    ecken_gewonnen = ecken_a_ft > ecken_away_sig
            except Exception as e:
                print(f"  [Auswertung] Druck Ecken-Check Fehler: {e}")

        gewonnen = tor_gewonnen or ecken_gewonnen
        update_statistik("druck",gewonnen,quote,liga=liga,match_id=match_id)  # v59.17: liga ergänzt
        emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
        ql = f"💶 Quote: <b>{quote}</b> → Gewinn: <b>+{round((quote-1)*EINSATZ,2)}€</b>\n" if quote and gewonnen else ""
        return (f"📊 <b>Auswertung – Druck</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                f"📌 {home} vs {away}\n🔥 Druck-Team: <b>{druck_team}</b>\n"
                f"📊 Stand bei Signal: <b>{score_signal}</b>\n"
                f"📈 Endstand: <b>{score}</b>\n{ql}"
                f"━━━━━━━━━━━━━━━━━━━━\n{emoji}\n🕐 {jetzt()} Uhr")
    except Exception as e:
        print(f"  [Auswertung] Druck Fehler: {e}")
        return None

def auswertung_comeback(spiel,ft_result=None):
    match_id     = spiel["match_id"]
    rueckliegend = spiel["rueckliegend"]
    home,away,quote = spiel["home"],spiel["away"],spiel.get("quote")
    score_signal = spiel.get("score_signal","")
    h_sig,a_sig  = parse_score(score_signal) if score_signal else (0,0)
    liga = spiel.get("competition",spiel.get("liga",""))
    try:
        result = ft_result or ls_get_match_result(match_id,home,away,liga)
        if not result:
            return None
        score = result.get("score","")
        h,a   = parse_score(score) if score else (0,0)
        if h == 0 and a == 0 and (h_sig > 0 or a_sig > 0):
            try:
                events    = ls_get_events(match_id)
                tore_home = len([e for e in events if e.get("event") in ("Goal","goal") and e.get("home_away") == "home"])
                tore_away = len([e for e in events if e.get("event") in ("Goal","goal") and e.get("home_away") == "away"])
                if tore_home > 0 or tore_away > 0:
                    h,a = tore_home,tore_away; score = f"{h} - {a}"
                else:
                    h,a = h_sig,a_sig; score = score_signal
            except Exception:
                h,a = h_sig,a_sig; score = score_signal
        bereits_beide_getroffen = h_sig >= 1 and a_sig >= 1
        gewonnen = (h >= 1 and a >= 1) or bereits_beide_getroffen
        update_statistik("comeback",gewonnen,quote,liga=liga,match_id=match_id)  # v59.17: liga ergänzt
        emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
        ql = f"💶 Quote: <b>{quote}</b> → Gewinn: <b>+{round((quote-1)*EINSATZ,2)}€</b>\n" if quote and gewonnen else ""
        hinweis = " (zum Signalzeitpunkt bereits erfüllt)" if bereits_beide_getroffen else ""
        return (f"📊 <b>Auswertung – Comeback</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                f"📌 {home} vs {away}\n🔄 Rückliegend: <b>{rueckliegend}</b>\n"
                f"📊 Stand bei Signal: <b>{score_signal}</b>\n"
                f"🎯 Tipp: Beide Teams treffen{hinweis}\n"
                f"📈 Endstand: <b>{score}</b>\n{ql}"
                f"━━━━━━━━━━━━━━━━━━━━\n{emoji}\n🕐 {jetzt()} Uhr")
    except Exception as e:
        print(f"  [Auswertung] Comeback Fehler: {e}")
        return None

def auswertung_torflut(spiel,ft_result=None):
    match_id  = spiel["match_id"]
    grenze    = spiel["grenze"]
    hz1_tore  = spiel["hz1_tore"]
    home,away,quote = spiel["home"],spiel["away"],spiel.get("quote")
    liga = spiel.get("competition",spiel.get("liga",""))
    try:
        result = ft_result or ls_get_match_result(match_id,home,away,liga)
        if not result:
            return None
        score = result.get("score","")
        h,a   = parse_score(score) if score else (0,0)
        tore  = h+a
        if tore == 0 and hz1_tore > 0:
            h_e,a_e = _hole_tore_via_events(match_id)
            if h_e+a_e > 0:
                h,a = h_e,a_e; tore = h+a; score = f"{h} - {a}"
            else:
                tore = hz1_tore; score = f"mind. {hz1_tore} (HZ1)"
        gewonnen = tore > grenze
        update_statistik("torflut",gewonnen,quote,liga=liga,match_id=match_id)  # v59.17: liga ergänzt
        emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
        ql = f"💶 Quote: <b>{quote}</b> → Gewinn: <b>+{round((quote-1)*EINSATZ,2)}€</b>\n" if quote and gewonnen else ""
        return (f"📊 <b>Auswertung – Torflut</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                f"📌 {home} vs {away}\n⚽ Tore HZ1: <b>{hz1_tore}</b>\n"
                f"🎯 Tipp: Über <b>{grenze}</b> Tore gesamt\n"
                f"📈 Endstand: <b>{score}</b> ({tore} Tore)\n{ql}"
                f"━━━━━━━━━━━━━━━━━━━━\n{emoji}\n🕐 {jetzt()} Uhr")
    except Exception as e:
        print(f"  [Auswertung] Torflut Fehler: {e}")
        return None

def auswertung_hz1tore(spiel,ft_result=None):
    match_id  = spiel["match_id"]
    richtung  = spiel["richtung"]
    linie     = spiel["linie"]
    home,away,quote = spiel["home"],spiel["away"],spiel.get("quote")
    liga = spiel.get("competition",spiel.get("liga",""))
    try:
        result = ft_result or ls_get_match_result(match_id,home,away,liga)
        if not result:
            return None
        ht = result.get("ht_score","")
        if not ht:
            import time as _t; _t.sleep(5)
            match = ls_get_single_match(match_id)
            ht    = (match.get("scores") or {}).get("ht_score","")
        if not ht:
            hz1_tore_events = _hole_hz1_tore_via_events(match_id)
            if hz1_tore_events >= 0:
                ht = f"{hz1_tore_events} - 0"
            else:
                return None
        hh,ha    = parse_score(ht)
        hz1_tore = hh+ha
        if richtung == "über":
            gewonnen = hz1_tore > linie
        else:
            gewonnen = hz1_tore < linie
        update_statistik("hz1tore",gewonnen,quote,liga=spiel.get("liga"),match_id=match_id)
        emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
        ql = f"💶 Quote: <b>{quote}</b> → Gewinn: <b>+{round((quote-1)*spiel.get('einsatz',EINSATZ),2)}€</b>\n" if quote and gewonnen else ""
        return (f"📊 <b>Auswertung – HZ1 Tore</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                f"📌 {home} vs {away}\n"
                f"🎯 Tipp: {richtung.capitalize()} <b>{linie}</b> Tore (HZ1)\n"
                f"📈 HZ1-Ergebnis: <b>{ht}</b> ({hz1_tore} Tore)\n{ql}"
                f"━━━━━━━━━━━━━━━━━━━━\n{emoji}\n🕐 {jetzt()} Uhr")
    except Exception as e:
        print(f"  [Auswertung] HZ1-Tore Fehler: {e}")
        return None

def auswertung_vztore(spiel,ft_result=None):
    match_id  = spiel["match_id"]
    richtung  = spiel["richtung"]
    linie     = spiel["linie"]
    home,away,quote = spiel["home"],spiel["away"],spiel.get("quote")
    liga = spiel.get("competition",spiel.get("liga",""))
    try:
        result = ft_result or ls_get_match_result(match_id,home,away,liga)
        if not result:
            return None
        score    = result.get("score","")
        h,a      = parse_score(score) if score else (0,0)
        vz_tore  = h+a
        if vz_tore == 0:
            h_e,a_e = _hole_tore_via_events(match_id)
            if h_e+a_e > 0:
                h,a = h_e,a_e; vz_tore = h+a; score = f"{h} - {a}"
        if richtung == "über":
            gewonnen = vz_tore > linie
        else:
            gewonnen = vz_tore < linie
        update_statistik("vztore",gewonnen,quote,liga=spiel.get("liga"),match_id=match_id)
        emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
        ql = f"💶 Quote: <b>{quote}</b> → Gewinn: <b>+{round((quote-1)*spiel.get('einsatz',EINSATZ),2)}€</b>\n" if quote and gewonnen else ""
        return (f"📊 <b>Auswertung – Vollzeit Tore</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                f"📌 {home} vs {away}\n"
                f"🎯 Tipp: {richtung.capitalize()} <b>{linie}</b> Tore (VZ)\n"
                f"📈 Endstand: <b>{score}</b> ({vz_tore} Tore)\n{ql}"
                f"━━━━━━━━━━━━━━━━━━━━\n{emoji}\n🕐 {jetzt()} Uhr")
    except Exception as e:
        print(f"  [Auswertung] VZ-Tore Fehler: {e}")
        return None

def auswertung_hz2tore(spiel,ft_result=None):
    """
    v59.22 NEU: Der HZ2-Tore-Bot wurde bisher fälschlich mit auswertung_hz1tore()
    ausgewertet (prüft nur den Halbzeitstand). Da dieser Bot nur feuert, wenn HZ1 bereits
    0:0 stand, wäre die Auswertung IMMER "0 Tore -> verloren" gewesen, egal was in der
    2. Halbzeit passierte. Diese Funktion berechnet die HZ2-Tore korrekt als
    (Vollzeit-Tore minus Halbzeit-Tore).
    """
    match_id  = spiel["match_id"]
    richtung  = spiel.get("richtung","ueber")
    linie     = spiel.get("linie",0.5)
    home,away,quote = spiel["home"],spiel["away"],spiel.get("quote")
    liga = spiel.get("competition",spiel.get("liga",""))
    try:
        result = ft_result or ls_get_match_result(match_id,home,away,liga)
        if not result:
            return None
        score = result.get("score","")
        ht    = result.get("ht_score","")
        h,a   = parse_score(score) if score else (0,0)
        vz_tore = h+a
        if vz_tore == 0:
            h_e,a_e = _hole_tore_via_events(match_id)
            if h_e+a_e > 0:
                h,a = h_e,a_e; vz_tore = h+a; score = f"{h} - {a}"
        if ht:
            hh,ha = parse_score(ht)
            hz1_tore = hh+ha
        else:
            hz1_tore = 0  # Bot feuert nur bei bestätigtem 0:0 zur Halbzeit
        hz2_tore = max(0, vz_tore-hz1_tore)
        if richtung in ("über","ueber"):
            gewonnen = hz2_tore > linie
        else:
            gewonnen = hz2_tore < linie
        update_statistik("hz2tore",gewonnen,quote,liga=liga,match_id=match_id)
        emoji = "✅ GEWONNEN" if gewonnen else "❌ VERLOREN"
        ql = f"💶 Quote: <b>{quote}</b> → Gewinn: <b>+{round((quote-1)*spiel.get('einsatz',EINSATZ),2)}€</b>\n" if quote and gewonnen else ""
        return (f"📊 <b>Auswertung – HZ2 Tore</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                f"📌 {home} vs {away}\n"
                f"🎯 Tipp: {richtung.capitalize()} <b>{linie}</b> Tore (HZ2)\n"
                f"📈 Endstand: <b>{score}</b> (HZ1: {hz1_tore} | HZ2: {hz2_tore} Tore)\n{ql}"
                f"━━━━━━━━━━━━━━━━━━━━\n{emoji}\n🕐 {jetzt()} Uhr")
    except Exception as e:
        print(f"  [Auswertung] HZ2-Tore Fehler: {e}")
        return None

# ============================================================
#  AUSWERTUNGS-THREAD
# ============================================================

def send_vip_werbung():
    """v60.2 NEU: Tägliche Werbe-Nachricht für die Free-Gruppe, die auf die VIP-Gruppe
    hinweist. Geht an den gleichen Kanal wie die normalen Signale (TELEGRAM_CHAT_ID)."""
    global vip_werbung_gesendet
    if not VIP_TELEGRAM_HANDLE:
        print("  [VIP-Werbung] Übersprungen – VIP_TELEGRAM_HANDLE ist nicht gesetzt (Railway-Variable fehlt)")
        return
    msg = (
        "🔥 <b>Bist du bereit für den nächsten Schritt?</b> 🔥\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "Die meisten in dieser Gruppe sehen nur einen Bruchteil dessen, was hier "
        "wirklich möglich ist. Die Free-Tipps sind erst der Anfang – ein kleiner "
        "Vorgeschmack auf das, was unsere <b>VIP-Mitglieder</b> jeden Tag bekommen.\n\n"
        "🏆 <b>Als VIP-Mitglied bekommst du:</b>\n"
        "✅ Exklusive Tipps mit deutlich höherer Konfidenz\n"
        "✅ Mehr Signale – früher, öfter & direkt live\n"
        "✅ Volle KI-Analyse zu jedem einzelnen Tipp\n"
        "✅ Persönlichen Support & direkten Austausch mit mir\n"
        "✅ Eine kleine, eng zusammenhaltende Community statt anonymer Masse\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        "💭 <i>Stell dir vor, du hättest die VIP-Tipps schon die ganze Woche gehabt...</i>\n\n"
        "Die Plätze in der VIP-Gruppe sind bewusst begrenzt – nicht jeder, der "
        "anfragt, wird aufgenommen. So bleibt die Qualität hoch und jedes Mitglied "
        "bekommt wirklich die Aufmerksamkeit, die es verdient.\n\n"
        "Viele warten zu lange und ärgern sich dann über die Tipps, die sie verpasst "
        "haben. Warte nicht auf den perfekten Moment – der kommt sonst nie.\n\n"
        "━━━━━━━━━━━━━━━━━━━━\n"
        f"📩 <b>Interesse?</b> Schreib mir einfach privat auf Telegram: <b>@{VIP_TELEGRAM_HANDLE}</b>\n"
        "Ich beantworte dir gerne alle Fragen und wir schauen gemeinsam, ob VIP "
        "das Richtige für dich ist.\n"
        "━━━━━━━━━━━━━━━━━━━━\n\n"
        "⚠️ 18+ | Verantwortungsvoll spielen"
    )
    send_telegram(msg)
    vip_werbung_gesendet = de_now().date()
    print("  [VIP-Werbung] ✅ Werbenachricht gesendet")

def bot_auswertung_und_berichte():
    """
    v59 FIX (Tagesbericht): Vorher wurde eine lokale Variable
    (tagesbericht_gesendet_local) genutzt, die bei jedem Thread-Neustart
    (z.B. durch den Watchdog nach einem Crash) zurückgesetzt wurde und die
    Bedingung `now.hour == TAGESBERICHT_UHRZEIT` nur exakt in Stunde 2
    zutraf. Lief der Bot ausgerechnet um 02:00 Uhr nicht, fiel der ganze
    Tagesbericht aus. Jetzt: globaler, persistierter Marker
    `tagesbericht_gesendet` + Nachhol-Logik (>= statt ==), sodass der
    Bericht auch nachträglich am selben Tag noch verschickt wird.
    """
    print("[Auswertung-Bot] Gestartet | Berichte – Auswertung via bot_nachschau")
    global tagesbericht_gesendet
    letzter_wochenbericht = de_now().isocalendar()[1]
    while True:
        try:
            now = de_now()
            if now.hour >= TAGESBERICHT_UHRZEIT and tagesbericht_gesendet != now.date():
                send_tagesbericht()
            if now.hour >= VIP_WERBUNG_UHRZEIT and vip_werbung_gesendet != now.date():
                send_vip_werbung()
            aktuelle_woche = now.isocalendar()[1]
            if now.weekday() == 0 and aktuelle_woche != letzter_wochenbericht:
                send_wochenbericht()
                letzter_wochenbericht = aktuelle_woche
            if now.day == 1 and now.hour == 9 and now.minute < 3:
                monatsbericht_key = f"{now.year}-{now.month}"
                if not hasattr(bot_auswertung_und_berichte,"_letzter_monat") or \
                        bot_auswertung_und_berichte._letzter_monat != monatsbericht_key:
                    bot_auswertung_und_berichte._letzter_monat = monatsbericht_key
            auswertung_fallback_check()
            if not hasattr(bot_auswertung_und_berichte,"_letzter_cleanup") or \
                    time.time()-bot_auswertung_und_berichte._letzter_cleanup > 3600:
                cache_aufraumen()
                bot_auswertung_und_berichte._letzter_cleanup = time.time()

            if now.hour == 3 and now.minute < 2:
                mem_key = f"mem_{now.strftime('%Y-%m-%d')}"
                if not hasattr(bot_auswertung_und_berichte,"_mem_key") or                         bot_auswertung_und_berichte._mem_key != mem_key:
                    bot_auswertung_und_berichte._mem_key = mem_key
                    cleanup_memory()
        except Exception as e:
            print(f"  [Auswertung-Bot] Fehler: {e}")
        time.sleep(120)


# ============================================================
#  FUSSBALL BOTS
# ============================================================

def bot_ecken():
    print(f"[Ecken-Bot] Gestartet | max. {MAX_CORNERS} Ecken in HZ1")
    while True:
        try:
            matches  = get_live_matches()
            halftime = [m for m in matches if m.get("status") == "HALF TIME BREAK"]
            print(f"[{jetzt()}] [Ecken-Bot] {len(halftime)} Halbzeit-Spiele")
            for game in halftime:
                match_id = str(game.get("id"))
                if match_id in notified_ecken:
                    continue
                stats        = get_statistiken(match_id)
                corners_home = stats["corners_home"]
                corners_away = stats["corners_away"]
                corners      = corners_home+corners_away
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","International")
                score   = game.get("scores",{}).get("score","?")
                grenze  = corners*2+1
                if corners == 0:
                    continue
                if corners > MAX_CORNERS:
                    print(f"  [Ecken-Bot] {home} vs {away} | Zu viele Ecken: {corners} > {MAX_CORNERS}")
                    continue
                if not liga_erlaubt(comp):
                    continue
                if not liga_signal_erlaubt(comp):  # v59.15: Liga-Zuverlässigkeit
                    continue
                if not whitelist_check(comp,home,away):
                    continue
                # v59.18 FIX: Ecken haben keinen passenden Markt im "Goals Over/Under"-Datensatz –
                # vorher wurde trotzdem eine (irreführende) Tore-Quote angezeigt. Jetzt keine
                # Quote mehr für diesen Tipp-Typ, statt eine unpassende Zahl zu raten.
                qd      = {"quote":None,"avg_quote":None,"bookmaker_anzahl":0}
                quote   = None
                bm_anz  = qd["bookmaker_anzahl"]
                if bm_anz > 0 and bm_anz < MIN_BOOKMAKER_ANZAHL:
                    continue
                if quote and quote < MIN_QUOTE:
                    continue
                if not gegentipp_check(match_id,"ecken","unter","Ecken-Bot"):
                    continue
                schlecht = schlechtes_wetter(country)
                wetter_bonus = 1 if schlecht else 0
                grenze = corners*2+1+wetter_bonus
                if not ecken_tipp_sinnvoll(game,grenze):
                    print(f"  [Ecken-Bot] Hist. Durchschnitt passt nicht – übersprungen")
                    continue
                shots_h_live = stats["shots_on_target_home"]
                shots_a_live = stats["shots_on_target_away"]
                shots_ges_live = shots_h_live+shots_a_live
                poss_h_live  = _safe_int(stats["possession_home"])
                poss_a_live  = _safe_int(stats["possession_away"])

                if shots_ges_live >= 7:
                    print(f"  [Ecken-Bot] Zu intensiv: {shots_ges_live} Schüsse aufs Tor – übersprungen")
                    continue

                if poss_h_live > 0 and poss_a_live > 0:
                    if abs(poss_h_live-poss_a_live) >= 25:
                        print(f"  [Ecken-Bot] Einseitiger Ballbesitz ({poss_h_live}:{poss_a_live}) – übersprungen")
                        continue

                score_h_live,score_a_live = parse_score(game.get("scores",{}).get("score","0 - 0"))
                if abs(score_h_live-score_a_live) >= 2:
                    print(f"  [Ecken-Bot] Hohe Tordifferenz ({score_h_live}:{score_a_live}) – übersprungen")
                    continue

                league_id = str((game.get("competition") or {}).get("id",""))
                home_id   = str((game.get("home") or {}).get("id",""))
                away_id   = str((game.get("away") or {}).get("id",""))
                analyse   = baue_analyse_text(home,away,home_id,away_id,league_id,{
                    "📐 Ecken HZ1": f"{corners} ({corners_home}|{corners_away})",
                    "🎯 Grenze":    f"Unter {grenze} gesamt",
                })
                konfidenz = berechne_konfidenz("ecken",comp,quote,wetter_schlecht=schlecht,bookmaker_anzahl=bm_anz)
                cl_ok,cl_text = claude_tipp_review(home,away,"ecken",analyse)
                if not cl_ok:
                    konfidenz = max(1,konfidenz-2)
                if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                einsatz = kelly_einsatz_bankroll(quote,"ecken") if quote else EINSATZ
                ke      = konfidenz_emoji(konfidenz)
                ev_data = berechne_ev_score(konfidenz, quote) if quote else {"label":"–"}
                ql      = f"\n💶 Quote: <b>{quote}</b> | 💰 Einsatz: <b>{einsatz}€</b> | {ev_data['label']}" if quote else ""
                cl_line = f"\n🤖 Claude: <b>{cl_text}</b>" if cl_text else ""
                odds_vgl = get_odds_vergleich(match_id)
                if not prüfe_ecken_verfuegbar(match_id):
                    notified_ecken.add(match_id)
                    continue
                if not tipp_erlaubt(match_id,"Ecken-Bot"):
                    continue
                if not signal_spam_check():
                    continue
                msg = (f"📐 <b>Ecken Tipp!</b> {ke} Konfidenz: <b>{konfidenz}/10</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                       f"📊 Stand: <b>{score}</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"{analyse}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"📐 Ecken: 🔵 {corners_home} | 🔴 {corners_away} | Gesamt: {corners}\n"
                       f"🎯 Tipp: Unter <b>{grenze}</b> Ecken gesamt{ql}{odds_vgl}{cl_line}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                send_telegram(msg)
                # v59: wait_for_message=True -> Message-ID für spätere ✅/❌-Reaktion sichern
                dc_info = send_discord_embed(DISCORD_WEBHOOK_ECKEN,
                    discord_ecken_tipp(home,away,comp,country,score,corners_home,corners_away,corners,grenze,quote,einsatz=einsatz,konfidenz=konfidenz),
                    wait_for_message=True)
                notified_ecken.add(match_id)
                multi_bonus = multi_signal_check(match_id,"Ecken-Bot")
                konfidenz   = min(10,konfidenz+multi_bonus)
                beobachtung_hinzufuegen(match_id,{
                    "typ":"ecken","match_id":match_id,"home":home,"away":away,
                    "hz1_ecken":corners,"quote":quote,"einsatz":einsatz,"liga":comp,"konfidenz":konfidenz,
                    "webhook":DISCORD_WEBHOOK_ECKEN,"signal_zeit":time.time(),"bot":"Ecken-Bot",
                    "discord_message_id": (dc_info or {}).get("message_id"),
                    "discord_channel_id": (dc_info or {}).get("channel_id"),
                })
                signal_eintragen(match_id,"ecken",home,away,comp,corners,grenze,quote,einsatz)
                gegentipp_registrieren(match_id,"ecken","unter","Ecken-Bot")
                print(f"  [Ecken-Bot] ✅ {home} vs {away} | K:{konfidenz}/10")
                time.sleep(0.5)
            bot_fehler_reset("Ecken-Bot")
        except Exception as e:
            bot_fehler_melden("Ecken-Bot",e)
        try:
            dynamischer_sleep(get_live_matches())
        except Exception:
            time.sleep(FUSSBALL_INTERVAL*60)

def bot_torwart():
    print(f"[Torwart-Bot] Gestartet | 0:0 + mind. {MIN_SHOTS_ON_TARGET} Schüsse")
    while True:
        try:
            matches = get_live_matches()
            aktiv   = [m for m in matches if m.get("status") in ("IN PLAY","ADDED TIME","HALF TIME BREAK")
                       and (20 <= _safe_int(m.get("time",0)) <= SPAETESTE_SIGNAL_MINUTE or m.get("status") == "HALF TIME BREAK")]
            print(f"[{jetzt()}] [Torwart-Bot] {len(aktiv)} aktive Spiele (ab Min. 20)")
            for game in aktiv:
                match_id = str(game.get("id"))
                if match_id in notified_torwart:
                    continue
                score = game.get("scores",{}).get("score","")
                if "0 - 0" not in score and "0-0" not in score:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","International")
                if not liga_signal_erlaubt(comp):  # v59.15: Liga-Zuverlässigkeit
                    continue
                if not liga_erlaubt(comp):  # v59.17: Liga-Trefferquote
                    continue
                stats      = get_statistiken(match_id)
                shots_home = stats["shots_on_target_home"]
                shots_away = stats["shots_on_target_away"]
                shots_ges  = shots_home+shots_away
                if shots_ges < MIN_SHOTS_ON_TARGET:
                    continue
                if not signal_spam_check():
                    continue
                if not tipp_erlaubt(match_id,"Torwart-Bot"):
                    continue
                saves_home = stats["saves_home"]
                saves_away = stats["saves_away"]
                poss_home  = stats["possession_home"]
                poss_away  = stats["possession_away"]
                status  = game.get("status","")
                minute  = game.get("time","?")
                min_text = "Halbzeit" if status == "HALF TIME BREAK" else f"{minute}'"
                quote   = get_quote(match_id,"torwart",linie=0.5,richtung="über")  # v59.18: passende Linie
                if quote and quote < MIN_QUOTE:
                    continue
                einsatz   = kelly_einsatz_bankroll(quote,"torwart") if quote else EINSATZ
                konfidenz = berechne_konfidenz("torwart",comp,quote)  # v59.8: Konfidenz ergänzt
                if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke        = konfidenz_emoji(konfidenz)
                ql      = f"\n💶 Quote: <b>{quote}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote else ""
                msg     = (f"🧤 <b>Torwart-Alarm!</b> {ke} Konfidenz: <b>{konfidenz}/10</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                           f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                           f"📊 Stand: <b>0:0</b> | {min_text}\n━━━━━━━━━━━━━━━━━━━━\n"
                           f"🎯 Schüsse: <b>{shots_ges}</b> ({shots_home}|{shots_away})\n"
                           f"🧤 Paraden: <b>{saves_home+saves_away}</b>\n"
                           f"⚽ Ballbesitz: {poss_home}%|{poss_away}%\n"
                           f"🎯 Tipp: Mind. <b>1 Tor</b> fällt noch{ql}\n"
                           f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                send_telegram(msg)
                dc_info = send_discord_embed(DISCORD_WEBHOOK_TORWART,
                    discord_torwart_tipp(home,away,comp,country,shots_home,shots_away,
                        saves_home,saves_away,poss_home,poss_away,min_text,quote,einsatz=einsatz,konfidenz=konfidenz),
                    wait_for_message=True)
                notified_torwart.add(match_id)
                beobachtung_hinzufuegen(match_id,{
                    "typ":"torwart","match_id":match_id,"home":home,"away":away,
                    "quote":quote,"einsatz":einsatz,"liga":comp,"konfidenz":konfidenz,
                    "webhook":DISCORD_WEBHOOK_TORWART,"signal_zeit":time.time(),"bot":"Torwart-Bot",
                    "discord_message_id": (dc_info or {}).get("message_id"),
                    "discord_channel_id": (dc_info or {}).get("channel_id"),
                })
                signal_eintragen(match_id,"torwart",home,away,comp,shots_ges,1,quote,einsatz)
                print(f"  [Torwart-Bot] ✅ {home} vs {away} | {shots_ges} Schüsse")
                time.sleep(0.5)
            bot_fehler_reset("Torwart-Bot")
        except Exception as e:
            bot_fehler_melden("Torwart-Bot",e)
        try:
            dynamischer_sleep(get_live_matches())
        except Exception:
            time.sleep(FUSSBALL_INTERVAL*60)

def bot_druck():
    print(f"[Druck-Bot] Gestartet | Ratio {DRUCK_RATIO}:1 bei mind. {MIN_DRUCK_ECKEN} Ecken")
    while True:
        try:
            matches = get_live_matches()
            laufend = [m for m in matches if m.get("status") in ("IN PLAY","ADDED TIME")
                       and 20 <= _safe_int(m.get("time",0)) <= SPAETESTE_SIGNAL_MINUTE]
            print(f"[{jetzt()}] [Druck-Bot] {len(laufend)} Spiele geprüft")
            for game in laufend:
                match_id = str(game.get("id"))
                if match_id in notified_druck:
                    continue
                stats  = get_statistiken(match_id)
                c_home = stats["corners_home"]
                c_away = stats["corners_away"]
                gesamt_ecken = c_home+c_away
                if gesamt_ecken == 0 or gesamt_ecken < MIN_DRUCK_ECKEN:
                    continue
                druck_team = None
                if c_away > 0 and c_home/c_away >= DRUCK_RATIO:
                    if c_away < 3:
                        continue
                    druck_team    = game.get("home",{}).get("name","?")
                    ecken_stark   = c_home; ecken_schwach = c_away
                elif c_home > 0 and c_away/c_home >= DRUCK_RATIO:
                    if c_home < 3:
                        continue
                    druck_team    = game.get("away",{}).get("name","?")
                    ecken_stark   = c_away; ecken_schwach = c_home
                if not druck_team:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","International")
                if not liga_signal_erlaubt(comp):  # v59.15: Liga-Zuverlässigkeit
                    continue
                if not liga_erlaubt(comp):  # v59.17: Liga-Trefferquote
                    continue
                score   = game.get("scores",{}).get("score","?")
                minute  = game.get("time","?")
                quote   = None  # v59.18 FIX: "Nächste Ecke/Tor" hat keinen passenden Tore-Markt – keine Quote statt geraten
                if quote and quote < MIN_QUOTE:
                    continue
                einsatz   = kelly_einsatz_bankroll(quote,"druck") if quote else EINSATZ
                # v59.8: Konfidenz ergänzt, mit Bonus für besonders starke Dominanz
                konfidenz = berechne_konfidenz("druck",comp,quote)
                if ecken_schwach > 0 and ecken_stark/ecken_schwach >= 4:
                    konfidenz = min(10,konfidenz+1)
                if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke      = konfidenz_emoji(konfidenz)
                ql      = f"\n💶 Quote: <b>{quote}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote else ""
                msg = (f"🔥 <b>Druck Signal!</b> {ke} Konfidenz: <b>{konfidenz}/10</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                       f"📊 Stand: <b>{score}</b> | Minute: <b>{minute}'</b>\n"
                       f"🔥 Dominantes Team: <b>{druck_team}</b>\n"
                       f"📐 Ecken: <b>{ecken_stark}</b> : {ecken_schwach}\n"
                       f"🎯 Tipp: Nächste Ecke / Tor für <b>{druck_team}</b>{ql}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                if not signal_spam_check():
                    continue
                notified_druck.add(match_id)
                notified_sets_speichern()
                send_telegram(msg)
                dc_info = send_discord_embed(DISCORD_WEBHOOK_DRUCK,
                    discord_druck_tipp(home,away,comp,country,score,minute,druck_team,
                        ecken_stark,ecken_schwach,quote,einsatz=einsatz,konfidenz=konfidenz),
                    wait_for_message=True)
                tipp_erlaubt(match_id,"Druck-Bot")
                beobachtung_hinzufuegen(match_id,{
                    "typ":"druck","match_id":match_id,"home":home,"away":away,
                    "druck_team":druck_team,"score_signal":score,
                    # v59.23 FIX: Ecken-Stand bei Signal mitspeichern, da der Tipp
                    # "Nächste Ecke / Tor" auch über Ecken gewinnen kann (siehe auswertung_druck).
                    "ecken_home_signal":c_home,"ecken_away_signal":c_away,
                    "quote":quote,"einsatz":einsatz,"liga":comp,"konfidenz":konfidenz,
                    "webhook":DISCORD_WEBHOOK_DRUCK,"signal_zeit":time.time(),"bot":"Druck-Bot",
                    "discord_message_id": (dc_info or {}).get("message_id"),
                    "discord_channel_id": (dc_info or {}).get("channel_id"),
                })
                signal_eintragen(match_id,"druck",home,away,comp,ecken_stark,ecken_schwach,quote,einsatz)
                print(f"  [Druck-Bot] ✅ {home} vs {away} | {druck_team} dominiert ({ecken_stark}:{ecken_schwach})")
                time.sleep(0.5)
            bot_fehler_reset("Druck-Bot")
        except Exception as e:
            bot_fehler_melden("Druck-Bot",e)
        try:
            dynamischer_sleep(get_live_matches())
        except Exception:
            time.sleep(FUSSBALL_INTERVAL*60)

def bot_comeback():
    print(f"[Comeback-Bot] Gestartet | ab Minute {COMEBACK_AB_MINUTE}")
    while True:
        try:
            matches = get_live_matches()
            laufend = [m for m in matches if m.get("status") in ("IN PLAY","ADDED TIME")
                       and COMEBACK_AB_MINUTE <= _safe_int(m.get("time",0)) <= SPAETESTE_SIGNAL_MINUTE]
            print(f"[{jetzt()}] [Comeback-Bot] {len(laufend)} Spiele geprüft")
            for game in laufend:
                match_id = str(game.get("id"))
                if match_id in notified_comeback:
                    continue
                score_str = game.get("scores",{}).get("score","")
                h_tore,a_tore = parse_score(score_str)
                if abs(h_tore-a_tore) != 1:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                rueckliegend = away if h_tore > a_tore else home
                fuehrend     = home if h_tore > a_tore else away
                stats     = get_statistiken(match_id)
                shots_h   = stats["shots_on_target_home"]
                shots_a   = stats["shots_on_target_away"]
                poss_h    = _safe_int(stats["possession_home"])
                poss_a    = _safe_int(stats["possession_away"])
                if rueckliegend == home:
                    shots_r,shots_f = shots_h,shots_a
                    poss_r          = poss_h
                else:
                    shots_r,shots_f = shots_a,shots_h
                    poss_r          = poss_a
                if shots_r == 0 and shots_f == 0 and poss_r == 0:
                    continue
                druck_ok = shots_r > shots_f
                if not druck_ok or poss_r <= 45:
                    continue
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","International")
                if not liga_signal_erlaubt(comp):  # v59.15: Liga-Zuverlässigkeit
                    continue
                if not liga_erlaubt(comp):  # v59.17: Liga-Trefferquote
                    continue
                minute  = game.get("time","?")
                quote   = None  # v59.18 FIX: "Beide Teams treffen" (BTTS) ist kein Über/Unter-Markt – keine Quote statt geraten
                if quote and quote < MIN_QUOTE:
                    continue
                einsatz   = kelly_einsatz_bankroll(quote,"comeback") if quote else EINSATZ
                # v59.8: Konfidenz ergänzt, mit Bonus für hohen Ballbesitz des rückliegenden Teams
                konfidenz = berechne_konfidenz("comeback",comp,quote)
                if poss_r >= 60:
                    konfidenz = min(10,konfidenz+1)
                if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke      = konfidenz_emoji(konfidenz)
                ql      = f"\n💶 Quote: <b>{quote}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote else ""
                msg = (f"🔄 <b>Comeback Signal!</b> {ke} Konfidenz: <b>{konfidenz}/10</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                       f"📊 Stand: <b>{score_str}</b> | Minute: <b>{minute}'</b>\n"
                       f"📉 Rückliegend: <b>{rueckliegend}</b>\n"
                       f"🎯 Schüsse aufs Tor: <b>{shots_r}</b> (Gegner: {shots_f})\n"
                       f"⚽ Ballbesitz: <b>{poss_r}%</b>\n"
                       f"🎯 Tipp: Beide Teams treffen{ql}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                if not signal_spam_check():
                    continue
                notified_comeback.add(match_id)
                notified_sets_speichern()
                send_telegram(msg)
                dc_info = send_discord_embed(DISCORD_WEBHOOK_COMEBACK,
                    discord_comeback_tipp(home,away,comp,country,score_str,minute,
                        rueckliegend,fuehrend,shots_r,shots_f,poss_r,quote,einsatz=einsatz,konfidenz=konfidenz),
                    wait_for_message=True)
                tipp_erlaubt(match_id,"Comeback-Bot")
                beobachtung_hinzufuegen(match_id,{
                    "typ":"comeback","match_id":match_id,"home":home,"away":away,
                    "rueckliegend":rueckliegend,"score_signal":score_str,
                    "quote":quote,"einsatz":einsatz,"liga":comp,"konfidenz":konfidenz,
                    "webhook":DISCORD_WEBHOOK_COMEBACK,"signal_zeit":time.time(),
                    "bot":"Comeback-Bot","competition":comp,
                    "discord_message_id": (dc_info or {}).get("message_id"),
                    "discord_channel_id": (dc_info or {}).get("channel_id"),
                })
                signal_eintragen(match_id,"comeback",home,away,comp,shots_r,1,quote,einsatz)
                print(f"  [Comeback-Bot] ✅ {home} vs {away} | {rueckliegend} dominiert trotz Rückstand")
                time.sleep(0.5)
            bot_fehler_reset("Comeback-Bot")
        except Exception as e:
            bot_fehler_melden("Comeback-Bot",e)
        try:
            dynamischer_sleep(get_live_matches())
        except Exception:
            time.sleep(FUSSBALL_INTERVAL*60)

def bot_torflut():
    print(f"[Torflut-Bot] Gestartet | mind. {TORFLUT_MIN_TORE} Tore in HZ1")
    while True:
        try:
            matches  = get_live_matches()
            halftime = [m for m in matches if m.get("status") == "HALF TIME BREAK"]
            print(f"[{jetzt()}] [Torflut-Bot] {len(halftime)} Halbzeit-Spiele")
            for game in halftime:
                match_id = str(game.get("id"))
                if match_id in notified_torflut:
                    continue
                score_str = game.get("scores",{}).get("score","0 - 0")
                h,a = parse_score(score_str)
                tore_hz1 = h+a
                if tore_hz1 < TORFLUT_MIN_TORE:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","International")
                if not liga_signal_erlaubt(comp):  # v59.15: Liga-Zuverlässigkeit
                    continue
                if not liga_erlaubt(comp):  # v59.17: Liga-Trefferquote
                    continue
                stats_tf   = get_statistiken(match_id)
                shots_h_tf = stats_tf["shots_on_target_home"]
                shots_a_tf = stats_tf["shots_on_target_away"]
                shots_tf   = shots_h_tf+shots_a_tf
                poss_th    = stats_tf["possession_home"]
                poss_ta    = stats_tf["possession_away"]
                shots_pro_tor = max(1,shots_tf/max(tore_hz1,1))
                erwartete_hz2 = max(1,min(4,round(shots_tf/max(shots_pro_tor,1.5),1)))
                grenze_roh = tore_hz1+erwartete_hz2
                if grenze_roh <= tore_hz1+1.5:   grenze = tore_hz1+1
                elif grenze_roh <= tore_hz1+2.5: grenze = tore_hz1+2
                else:                             grenze = tore_hz1+3
                if tore_hz1 >= 4:
                    grenze = max(grenze,tore_hz1+2)
                quote   = get_quote(match_id,"torflut",linie=grenze,richtung="über")  # v59.18: passende Linie
                if quote and quote < MIN_QUOTE:
                    continue
                einsatz   = kelly_einsatz_bankroll(quote,"torflut") if quote else EINSATZ
                # v59.8: Konfidenz ergänzt, mit Bonus je mehr Tore schon in HZ1 fielen
                konfidenz = berechne_konfidenz("torflut",comp,quote)
                if tore_hz1 >= 4:
                    konfidenz = min(10,konfidenz+1)
                if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke      = konfidenz_emoji(konfidenz)
                ql      = f"\n💶 Quote: <b>{quote}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote else ""
                msg     = (f"🌊 <b>Torflut Signal!</b> {ke} Konfidenz: <b>{konfidenz}/10</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                           f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                           f"📊 Halbzeitstand: <b>{score_str}</b>\n"
                           f"⚽ Tore HZ1: <b>{tore_hz1}</b>\n"
                           f"🎯 Tipp: Über <b>{grenze}</b> Tore im Gesamtspiel{ql}\n"
                           f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                if not signal_spam_check():
                    continue
                if not tipp_erlaubt(match_id,"Torflut-Bot"):
                    continue
                send_telegram(msg)
                dc_info = send_discord_embed(DISCORD_WEBHOOK_TORFLUT,
                    discord_torflut_tipp(home,away,comp,country,score_str,tore_hz1,grenze,quote,shots_tf,poss_th,poss_ta,einsatz=einsatz,konfidenz=konfidenz),
                    wait_for_message=True)
                notified_torflut.add(match_id)
                beobachtung_hinzufuegen(match_id,{
                    "typ":"torflut","match_id":match_id,"home":home,"away":away,
                    "hz1_tore":tore_hz1,"grenze":grenze,"quote":quote,"einsatz":einsatz,"konfidenz":konfidenz,
                    "webhook":DISCORD_WEBHOOK_TORFLUT,"signal_zeit":time.time(),"bot":"Torflut-Bot",
                    "discord_message_id": (dc_info or {}).get("message_id"),
                    "discord_channel_id": (dc_info or {}).get("channel_id"),
                })
                signal_eintragen(match_id,"torflut",home,away,comp,tore_hz1,int(grenze),quote,einsatz)
                print(f"  [Torflut-Bot] ✅ {home} vs {away} | {tore_hz1} Tore in HZ1")
                time.sleep(0.5)
            bot_fehler_reset("Torflut-Bot")
        except Exception as e:
            bot_fehler_melden("Torflut-Bot",e)
        try:
            dynamischer_sleep(get_live_matches())
        except Exception:
            time.sleep(FUSSBALL_INTERVAL*60)


def bot_tore_analyse():
    """Kombinierter HZ1+VZ Tore Bot – analysiert H2H einmal pro Spiel."""
    print(f"[Tore-Bot] Gestartet | HZ1+VZ H2H-Analyse bis Minute {H2H_SIGNAL_BIS_MIN}")
    while True:
        try:
            matches = get_live_matches()
            frisch  = [m for m in matches if m.get("status") == "IN PLAY"
                       and 1 <= _safe_int(m.get("time",0)) <= H2H_SIGNAL_BIS_MIN]
            print(f"[{jetzt()}] [Tore-Bot] {len(frisch)} Spiele in Min. 1-{H2H_SIGNAL_BIS_MIN}")
            for game in frisch:
                match_id = str(game.get("id"))
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","International")
                minute  = game.get("time","?")
                if not liga_erlaubt(comp):
                    continue
                if not liga_signal_erlaubt(comp):  # v59.15: Liga-Zuverlässigkeit
                    continue
                home_id = str((game.get("home") or {}).get("id",""))
                away_id = str((game.get("away") or {}).get("id",""))
                if not home_id or not away_id:
                    continue
                h2h = get_h2h_daten(home_id,away_id)
                ana = analysiere_h2h_tore(h2h)
                if not ana:
                    continue
                qd     = get_quote_details(match_id)
                quote  = qd["quote"]
                bm_anz = qd["bookmaker_anzahl"]
                if quote and quote < MIN_QUOTE:
                    continue
                form_home_avg = get_team_saisonform(home_id)
                form_away_avg = get_team_saisonform(away_id)
                form_avg_val  = round((form_home_avg+form_away_avg)/2,2) if form_home_avg and form_away_avg else None
                # ── HZ1-Tore Tipp ──────────────────────────────────────
                if match_id not in notified_hz1tore and ana.get("avg_hz1") is not None:
                    dyn_ueber_hz1,dyn_unter_hz1 = berechne_dynamische_grenzen(ana["avg_hz1"],form_avg_val,"hz1")
                    tipp_hz1 = tipp_aus_avg(ana["avg_hz1"],dyn_ueber_hz1,dyn_unter_hz1)
                    if tipp_hz1 and gegentipp_check(match_id,"hz1tore",tipp_hz1[0],"Tore-Bot"):
                        richtung,linie = tipp_hz1
                        # v59.18 FIX: eigene, zur HZ1-Linie passende Quote holen statt der
                        # generischen (unspezifischen) "quote" von oben zu verwenden.
                        quote_hz1 = get_quote(match_id,"hz1tore",linie=linie,richtung=richtung)
                        form_ok   = form_stimmt_ueberein(home_id,away_id,ana["avg_hz1"],richtung)
                        einsatz   = kelly_einsatz_bankroll(quote_hz1,"hz1tore") if quote_hz1 else EINSATZ
                        konfidenz = berechne_konfidenz("hz1tore",comp,quote_hz1,
                            h2h_spiele=ana["hz1_spiele"],bookmaker_anzahl=bm_anz,
                            form_uebereinstimmung=form_ok)
                        konfidenz = min(10,konfidenz+multi_signal_check(match_id,"Tore-Bot"))
                        analyse_hz1 = f"H2H Ø HZ1-Tore: {ana['avg_hz1']} ({ana['hz1_spiele']} Spiele)\nTipp: {richtung} {linie}"
                        cl_ok,cl_text = claude_tipp_review(home,away,"hz1tore",analyse_hz1)
                        if not cl_ok: konfidenz = max(1,konfidenz-2)
                        if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                            continue
                        cl_hz1 = f"\n🤖 Claude: <b>{cl_text}</b>" if cl_text else ""
                        ql     = f"\n💶 Quote: <b>{quote_hz1}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote_hz1 else ""
                        msg    = (f"🥅 <b>HZ1-Tore Signal!</b> {konfidenz_emoji(konfidenz)} Konfidenz: <b>{konfidenz}/10</b>\n"
                                  f"━━━━━━━━━━━━━━━━━━━━\n🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                                  f"⏱️ Minute: <b>{minute}'</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                                  f"📊 H2H: <b>{ana['hz1_spiele']}</b> Spiele | Ø HZ1-Tore: <b>{ana['avg_hz1']}</b>\n"
                                  f"🎯 Tipp: {richtung.capitalize()} <b>{linie}</b> Tore (HZ1){ql}{cl_hz1}\n"
                                  f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                        # v59.11 FIX (Duplikate): tipp_erlaubt() MUSS vor dem Versand geprüft werden.
                        # Vorher wurde erst gesendet und danach geprüft – schlug die Prüfung fehl
                        # (z.B. weil für dieses Spiel schon ein anderer Tipp registriert war), wurde
                        # zwar bereits gepostet, match_id aber NIE zu notified_hz1tore hinzugefügt.
                        # Ergebnis: exakt dieselbe Nachricht wurde bei jedem Durchlauf (~alle 2 Min)
                        # erneut verschickt, solange das Spiel im Zeitfenster blieb.
                        if not tipp_erlaubt(match_id,"Tore-Bot-HZ1"):
                            notified_hz1tore.add(match_id)  # Spiel trotzdem sperren, nicht endlos erneut prüfen
                            continue
                        send_telegram(msg)
                        dc_info = send_discord_embed(DISCORD_WEBHOOK_HZ1TORE,
                            discord_hz1tore_tipp(home,away,comp,country,richtung,linie,ana["avg_hz1"],ana["hz1_spiele"],quote_hz1,einsatz=einsatz,konfidenz=konfidenz),
                            wait_for_message=True)
                        notified_hz1tore.add(match_id)
                        beobachtung_hinzufuegen(match_id,{
                            "typ":"hz1tore","match_id":match_id,"home":home,"away":away,"liga":comp,
                            "richtung":richtung,"linie":linie,"quote":quote_hz1,"einsatz":einsatz,"konfidenz":konfidenz,
                            "webhook":DISCORD_WEBHOOK_HZ1TORE,"signal_zeit":time.time(),
                            "discord_message_id": (dc_info or {}).get("message_id"),
                            "discord_channel_id": (dc_info or {}).get("channel_id"),
                        })
                        signal_eintragen(match_id,"hz1tore",home,away,comp,ana["avg_hz1"],linie,quote_hz1,einsatz)
                        gegentipp_registrieren(match_id,"hz1tore",richtung,"Tore-Bot")
                        print(f"  [Tore-Bot] HZ1 ✅ {home} vs {away} | {richtung} {linie}")
                # ── VZ-Tore Tipp ────────────────────────────────────────
                if match_id not in notified_vztore:
                    dyn_ueber_vz,dyn_unter_vz = berechne_dynamische_grenzen(ana["avg_vz"],form_avg_val,"vz")
                    tipp_vz = tipp_aus_avg(ana["avg_vz"],dyn_ueber_vz,dyn_unter_vz)
                    if tipp_vz and gegentipp_check(match_id,"vztore",tipp_vz[0],"Tore-Bot"):
                        richtung,linie = tipp_vz
                        # v59.18 FIX: eigene, zur VZ-Linie passende Quote holen statt der
                        # generischen (unspezifischen) "quote" von oben zu verwenden.
                        quote_vz  = get_quote(match_id,"vztore",linie=linie,richtung=richtung)
                        form_ok   = form_stimmt_ueberein(home_id,away_id,ana["avg_vz"],richtung)
                        einsatz   = kelly_einsatz_bankroll(quote_vz,"vztore") if quote_vz else EINSATZ
                        konfidenz = berechne_konfidenz("vztore",comp,quote_vz,
                            h2h_spiele=ana["spiele"],bookmaker_anzahl=bm_anz,
                            form_uebereinstimmung=form_ok)
                        analyse_vz = f"H2H Ø VZ-Tore: {ana['avg_vz']} ({ana['spiele']} Spiele)\nTipp: {richtung} {linie}"
                        cl_ok,cl_text = claude_tipp_review(home,away,"vztore",analyse_vz)
                        if not cl_ok: konfidenz = max(1,konfidenz-2)
                        if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                            continue
                        cl_vz  = f"\n🤖 Claude: <b>{cl_text}</b>" if cl_text else ""
                        ql     = f"\n💶 Quote: <b>{quote_vz}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote_vz else ""
                        msg    = (f"🏆 <b>Vollzeit-Tore Signal!</b> {konfidenz_emoji(konfidenz)} Konfidenz: <b>{konfidenz}/10</b>\n"
                                  f"━━━━━━━━━━━━━━━━━━━━\n🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                                  f"⏱️ Minute: <b>{minute}'</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                                  f"📊 H2H: <b>{ana['spiele']}</b> Spiele | Ø VZ-Tore: <b>{ana['avg_vz']}</b>\n"
                                  f"🎯 Tipp: {richtung.capitalize()} <b>{linie}</b> Tore (VZ){ql}{cl_vz}\n"
                                  f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                        # v59.11 FIX (Duplikate): tipp_erlaubt() MUSS vor dem Versand geprüft werden
                        # (gleicher Bug wie im HZ1-Block darüber – siehe Kommentar dort).
                        if not tipp_erlaubt(match_id,"Tore-Bot-VZ"):
                            notified_vztore.add(match_id)  # Spiel trotzdem sperren, nicht endlos erneut prüfen
                            continue
                        send_telegram(msg)
                        dc_info = send_discord_embed(DISCORD_WEBHOOK_VZTORE,
                            discord_vztore_tipp(home,away,comp,country,richtung,linie,ana["avg_vz"],ana["spiele"],quote_vz,einsatz=einsatz,konfidenz=konfidenz),
                            wait_for_message=True)
                        notified_vztore.add(match_id)
                        beobachtung_hinzufuegen(match_id,{
                            "typ":"vztore","match_id":match_id,"home":home,"away":away,"liga":comp,
                            "richtung":richtung,"linie":linie,"quote":quote_vz,"einsatz":einsatz,"konfidenz":konfidenz,
                            "webhook":DISCORD_WEBHOOK_VZTORE,"signal_zeit":time.time(),
                            "discord_message_id": (dc_info or {}).get("message_id"),
                            "discord_channel_id": (dc_info or {}).get("channel_id"),
                        })
                        signal_eintragen(match_id,"vztore",home,away,comp,ana["avg_vz"],linie,quote_vz,einsatz)
                        gegentipp_registrieren(match_id,"vztore",richtung,"Tore-Bot")
                        print(f"  [Tore-Bot] VZ ✅ {home} vs {away} | {richtung} {linie}")
                time.sleep(0.5)
            bot_fehler_reset("Tore-Bot")
        except Exception as e:
            bot_fehler_melden("Tore-Bot",e)
        try:
            dynamischer_sleep(get_live_matches())
        except Exception:
            time.sleep(FUSSBALL_INTERVAL*60)

# ============================================================
#  CORNER RUSH BOT
# ============================================================

notified_corner_rush = set()
_corner_history = {}

def bot_corner_rush():
    print("[CornerRush-Bot] Gestartet | 4+ Ecken in 10 Min")
    while True:
        try:
            matches = get_live_matches()
            laufend = [m for m in matches if m.get("status") in ("IN PLAY","ADDED TIME")
                       and 20 <= _safe_int(m.get("time",0)) <= SPAETESTE_SIGNAL_MINUTE]
            live_ids_cr = {str(m.get("id")) for m in laufend}
            for mid in list(_corner_history.keys()):
                if mid not in live_ids_cr:
                    del _corner_history[mid]
            for game in laufend:
                match_id = str(game.get("id"))
                if match_id in notified_corner_rush:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","International")
                score   = game.get("scores",{}).get("score","?")
                minute  = _safe_int(game.get("time",0))
                if not whitelist_check(comp,home,away):
                    continue
                stats     = get_statistiken(match_id)
                corners_h = stats["corners_home"]
                corners_a = stats["corners_away"]
                now_ts    = time.time()
                if match_id not in _corner_history:
                    _corner_history[match_id] = []
                _corner_history[match_id].append({"ts":now_ts,"ch":corners_h,"ca":corners_a})
                hist = [e for e in _corner_history[match_id] if now_ts-e["ts"] <= 12*60]
                _corner_history[match_id] = hist
                if len(hist) < 2:
                    continue
                alt    = hist[0]
                diff_h = corners_h-alt["ch"]
                diff_a = corners_a-alt["ca"]
                rush_team  = home if diff_h >= 4 else (away if diff_a >= 4 else None)
                rush_ecken = diff_h if diff_h >= 4 else (diff_a if diff_a >= 4 else 0)
                if not rush_team:
                    continue
                zeit_diff = round((now_ts-alt["ts"])/60,1)
                # v59.7/8: Quote + Einsatz-Empfehlung + Konfidenz ergänzt
                quote_cr   = None  # v59.18 FIX: "Nächste Ecke/Tor" hat keinen passenden Tore-Markt – keine Quote statt geraten
                einsatz_cr = kelly_einsatz_bankroll(quote_cr,"ecken") if quote_cr else EINSATZ
                konfidenz_cr = min(10,5+rush_ecken)  # mehr Ecken im Rush-Fenster = höhere Konfidenz
                if konfidenz_cr < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke_cr        = konfidenz_emoji(konfidenz_cr)
                ql_cr      = f"\n💶 Quote: <b>{quote_cr}</b> | 💰 Einsatz: <b>{einsatz_cr}€</b>" if quote_cr else f"\n💰 Einsatz: <b>{einsatz_cr}€</b>"
                msg = (f"📐 <b>Corner Rush!</b> {ke_cr} Konfidenz: <b>{konfidenz_cr}/10</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                       f"📊 Stand: <b>{score}</b> | Min. <b>{minute}'</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"⚡ <b>{rush_team}</b>: {rush_ecken} Ecken in {zeit_diff} Min!\n"
                       f"📐 Ecken gesamt: {corners_h}|{corners_a}\n"
                       f"💡 Extremer Druck → Tor oder weitere Ecken sehr wahrscheinlich\n"
                       f"🎯 Tipp: <b>Nächste Ecke / Tor für {rush_team}</b>{ql_cr}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                if not signal_spam_check():
                    continue
                notified_corner_rush.add(match_id)
                notified_sets_speichern()
                send_telegram(msg)
                embed = {
                    "title":f"📐 Corner Rush – {rush_team} eskaliert!","color":0xE67E22,
                    "fields":[
                        {"name":"🏆 Liga","value":comp,"inline":True},
                        {"name":"🌍 Land","value":country,"inline":True},
                        {"name":"⏱️ Minute","value":f"**{minute}'**","inline":True},
                        {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
                        {"name":"📊 Stand","value":f"**{score}**","inline":True},
                        {"name":"⚡ Rush-Team","value":f"**{rush_team}**","inline":True},
                        {"name":"📐 Ecken Rush","value":f"**{rush_ecken}** in {zeit_diff} Min","inline":True},
                        {"name":"📐 Gesamt","value":f"{corners_h}|{corners_a}","inline":True},
                        {"name":"🎯 Tipp","value":f"**Nächste Ecke / Tor {rush_team}**","inline":False},
                        {"name":"💶 Quote","value": (f"**{quote_cr}**" if quote_cr else "⚠️ Keine zuverlässige Quote"),"inline":True},
                        {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz_cr),"inline":True},
                        {"name":"🎯 Konfidenz","value":f"{ke_cr} **{konfidenz_cr}/10**","inline":True},
                    ],
                    "footer":{"text":f"CornerRush-Bot • {heute()} {jetzt()}"},
                }
                send_discord_embed(DISCORD_WEBHOOK_ECKEN,embed)
                print(f"  [CornerRush] ✅ {rush_team}: {rush_ecken} Ecken in {zeit_diff} Min")
                time.sleep(0.5)
            bot_fehler_reset("CornerRush-Bot")
        except Exception as e:
            bot_fehler_melden("CornerRush-Bot",e)
        time.sleep(90)


# ============================================================
#  SIGNAL TRACKER
# ============================================================

_signal_tracker = {}
_tracker_lock   = threading.Lock()

def tracker_laden():
    import json, os as _os
    global _signal_tracker
    if not _os.path.exists(SIGNAL_TRACKER_DATEI):
        return
    try:
        with open(SIGNAL_TRACKER_DATEI,"r") as f:
            data = json.load(f)
        with _tracker_lock:
            _signal_tracker = data
        offen = sum(1 for s in data.values() if s.get("status") == "offen")
        print(f"  [Tracker] {len(data)} Signale geladen, {offen} noch offen")
    except Exception as e:
        print(f"  [Tracker] Ladefehler: {e}")

def tracker_speichern():
    import json
    try:
        with _tracker_lock:
            data = dict(_signal_tracker)
        with open(SIGNAL_TRACKER_DATEI,"w") as f:
            json.dump(data,f,indent=2)
    except Exception as e:
        print(f"  [Tracker] Speicherfehler: {e}")

def tracker_signal_hinzufuegen(match_id: str, spiel: dict):
    key = f"{match_id}_{spiel.get('typ','')}"
    with _tracker_lock:
        _signal_tracker[key] = {
            **spiel,
            "match_id":    match_id,
            "status":      "offen",
            "signal_zeit": time.time(),
            "versuche":    0,
            "letzter_versuch": 0,
        }
    tracker_speichern()
    print(f"  [Tracker] ✅ Registriert: {spiel.get('home','?')} vs {spiel.get('away','?')} ({spiel.get('typ','')})")

def tracker_ausgewertet_markieren(key: str, gewonnen: bool):
    with _tracker_lock:
        if key in _signal_tracker:
            _signal_tracker[key]["status"]       = "ausgewertet"
            _signal_tracker[key]["gewonnen"]      = gewonnen
            _signal_tracker[key]["ausgewertet_um"] = de_now().strftime("%Y-%m-%d %H:%M")
    tracker_speichern()

def tracker_nicht_auswertbar_markieren(key: str):
    """
    v59.9 NEU: Für Signale, deren Ergebnis nach 20 Versuchen (~30 Min) nicht sicher
    ermittelt werden konnte. Bleibt BEWUSST außerhalb der Gewonnen/Verloren-Statistik
    (vorher wurden solche Fälle über tracker_ausgewertet_markieren(key,False) fälschlich
    als "verloren" behandelt, aber nie tatsächlich in update_statistik gezählt – dadurch
    blieben Tagesbericht/Wochenbericht bei 0/0 hängen, wenn die Ergebnis-Quellen für eine
    Liga öfter mal nichts liefern, UND es gab nie eine Reaktion, weil dieser Pfad den
    Discord-Embed-Code nie erreicht hat).
    """
    with _tracker_lock:
        if key in _signal_tracker:
            _signal_tracker[key]["status"] = "nicht_auswertbar"
            _signal_tracker[key]["ausgewertet_um"] = de_now().strftime("%Y-%m-%d %H:%M")
    tracker_speichern()

# ============================================================
#  v59.15 NEU: LIGA-ZUVERLÄSSIGKEIT (Qualität statt Quantität)
# ============================================================
# Statt Ligen von Hand zu raten und zu sperren, merkt sich der Bot pro Liga automatisch,
# wie oft die Auswertung tatsächlich gelingt oder scheitert. Ligen mit dauerhaft schlechter
# Auswertungs-Erfolgsquote werden für NEUE Signale automatisch gesperrt – bis sich die
# Datenlage für diese Liga bessert (Sperre ist nicht endgültig, wird laufend neu bewertet).
LIGA_AUSWERTUNG_DATEI    = "liga_auswertung.json"
LIGA_AUSWERTUNG_MIN_PROBEN = 5     # erst ab so vielen Versuchen wird überhaupt bewertet
LIGA_AUSWERTUNG_MIN_QUOTE  = 0.5   # mind. 50% müssen erfolgreich auswertbar sein
liga_auswertung_stats = {}   # {"Liga Name": {"erfolgreich": int, "nicht_auswertbar": int}}
_liga_auswertung_lock = threading.Lock()

def liga_auswertung_laden():
    import json, os as _os
    global liga_auswertung_stats
    if not _os.path.exists(LIGA_AUSWERTUNG_DATEI):
        return
    try:
        with open(LIGA_AUSWERTUNG_DATEI) as f:
            liga_auswertung_stats = json.load(f)
        print(f"  [Liga-Zuverlässigkeit] {len(liga_auswertung_stats)} Ligen geladen")
    except Exception as e:
        print(f"  [Liga-Zuverlässigkeit] Ladefehler: {e}")

def liga_auswertung_speichern():
    import json
    try:
        with _liga_auswertung_lock:
            data = dict(liga_auswertung_stats)
        with open(LIGA_AUSWERTUNG_DATEI,"w") as f:
            json.dump(data,f,indent=2)
    except Exception as e:
        print(f"  [Liga-Zuverlässigkeit] Speicherfehler: {e}")

def liga_auswertung_update(liga: str, erfolgreich: bool):
    if not liga:
        return
    with _liga_auswertung_lock:
        liga_auswertung_stats.setdefault(liga, {"erfolgreich":0,"nicht_auswertbar":0})
        if erfolgreich:
            liga_auswertung_stats[liga]["erfolgreich"] += 1
        else:
            liga_auswertung_stats[liga]["nicht_auswertbar"] += 1
    liga_auswertung_speichern()

def liga_signal_erlaubt(liga: str) -> bool:
    """
    Prüft VOR dem Senden eines neuen Signals, ob diese Liga bisher zuverlässig genug
    auswertbar war. Erst ab LIGA_AUSWERTUNG_MIN_PROBEN Versuchen wird überhaupt geprüft
    (sonst würde jede neue Liga sofort gesperrt). Unterhalb von LIGA_AUSWERTUNG_MIN_QUOTE
    Erfolgsquote wird die Liga für neue Signale gesperrt.
    """
    if not liga:
        return True
    stat = liga_auswertung_stats.get(liga)
    if not stat:
        return True
    gesamt = stat["erfolgreich"] + stat["nicht_auswertbar"]
    if gesamt < LIGA_AUSWERTUNG_MIN_PROBEN:
        return True
    quote = stat["erfolgreich"] / gesamt
    if quote < LIGA_AUSWERTUNG_MIN_QUOTE:
        print(f"  [Liga-Zuverlässigkeit] '{liga}' gesperrt: nur {stat['erfolgreich']}/{gesamt} ({round(quote*100)}%) auswertbar")
        return False
    return True

def tracker_get_offene() -> list:
    jetzt_ts = time.time()
    with _tracker_lock:
        offene = []
        for key,sig in _signal_tracker.items():
            if sig.get("status") != "offen":
                continue
            alter_h = (jetzt_ts-sig.get("signal_zeit",jetzt_ts))/3600
            if alter_h > 5:
                _signal_tracker[key]["status"] = "abgelaufen"
                continue
            offene.append((key,sig))
    return offene

# ============================================================
#  TRIPLE-VERIFIKATION FÜR SPIELERGEBNISSE
# ============================================================

FD_BASE = "https://api.football-data.org/v4"
FD_LIGA_IDS = {
    "Premier League":2021,"Bundesliga":2002,"La Liga":2014,
    "Serie A":2019,"Ligue 1":2015,"Primeira Liga":2017,
    "Eredivisie":2003,"Championship":2016,"Champions League":2001,
    "Europa League":2146,"World Cup":2000,"European Championship":2018,
}

def fd_suche_spiel(home: str, away: str, liga: str = ""):
    if not FOOTBALLDATA_KEY:
        return None
    try:
        headers  = {"X-Auth-Token":FOOTBALLDATA_KEY}
        heute_str = de_now().strftime("%Y-%m-%d")
        liga_id  = FD_LIGA_IDS.get(liga)
        urls = []
        if liga_id:
            urls.append(f"{FD_BASE}/competitions/{liga_id}/matches?status=FINISHED&dateFrom={heute_str}&dateTo={heute_str}")
        urls.append(f"{FD_BASE}/matches?status=FINISHED&dateFrom={heute_str}&dateTo={heute_str}")
        for url in urls:
            try:
                resp = requests.get(url,headers=headers,timeout=8)
                if resp.status_code != 200:
                    continue
                for match in resp.json().get("matches",[]):
                    h = (match.get("homeTeam") or {}).get("shortName","") or (match.get("homeTeam") or {}).get("name","")
                    a = (match.get("awayTeam") or {}).get("shortName","") or (match.get("awayTeam") or {}).get("name","")
                    if (home[:5].lower() in h.lower() or h.lower()[:5] in home.lower()) and \
                       (away[:5].lower() in a.lower() or a.lower()[:5] in away.lower()):
                        score  = match.get("score",{})
                        full   = score.get("fullTime",{})
                        half   = score.get("halfTime",{})
                        h_ft   = full.get("home")
                        a_ft   = full.get("away")
                        h_ht   = half.get("home")
                        a_ht   = half.get("away")
                        status = match.get("status","")
                        if status == "FINISHED" and h_ft is not None and a_ft is not None:
                            return {
                                "status":"FT","score":f"{h_ft} - {a_ft}",
                                "ht_score":f"{h_ht} - {a_ht}" if h_ht is not None else "",
                                "quelle":"football-data.org",
                            }
            except Exception as e:
                print(f"  [FD] Fehler: {e}")
                continue
    except Exception as e:
        print(f"  [FD] Suche Fehler: {e}")
    return None

def thesportsdb_suche_spiel(home: str, away: str):
    try:
        resp = requests.get("https://www.thesportsdb.com/api/v1/json/3/searchteams.php",
                            params={"t":home},timeout=8)
        if resp.status_code != 200:
            return None
        teams   = (resp.json().get("teams") or [])
        team_id = None
        for t in teams:
            if t.get("strSport","").lower() == "soccer" and home[:5].lower() in t.get("strTeam","").lower():
                team_id = t.get("idTeam")
                break
        if not team_id:
            return None
        resp2 = requests.get("https://www.thesportsdb.com/api/v1/json/3/eventslast.php",
                             params={"id":team_id},timeout=8)
        if resp2.status_code != 200:
            return None
        events = (resp2.json().get("results") or [])
        heute_str = de_now().strftime("%Y-%m-%d")
        for ev in events:
            if ev.get("strSport","").lower() != "soccer":
                continue
            if ev.get("dateEvent","") != heute_str:
                continue
            h_name = ev.get("strHomeTeam","").lower()
            a_name = ev.get("strAwayTeam","").lower()
            if home[:5].lower() in h_name and away[:5].lower() in a_name:
                h_score = ev.get("intHomeScore")
                a_score = ev.get("intAwayScore")
                if h_score is not None and a_score is not None:
                    return {"status":"FT","score":f"{h_score} - {a_score}","ht_score":"","quelle":"thesportsdb"}
    except Exception as e:
        print(f"  [TSDB] Fehler: {e}")
    return None

def hz1_bereit_fuer_auswertung(match_id: str) -> bool:
    """
    v59.20 NEU: Für HZ1-Tore-Wetten reicht eine pauschale Wartezeit nicht – die Auswertung
    braucht das ECHTE Ende der 1. Halbzeit. Prüft direkt den aktuellen Live-Status/die
    Spielminute des Fixtures, statt eine feste Zeitspanne seit Signal-Versand zu raten.
    True = 1. Halbzeit ist vorbei (oder Spiel nicht mehr live, dann auswertbar versuchen).
    False = Spiel läuft erkennbar noch in der 1. Halbzeit, noch nicht versuchen.
    """
    try:
        live = get_live_matches()
        for m in live:
            if str(m.get("id")) == str(match_id):
                status  = m.get("status","")
                elapsed = _safe_int(m.get("time",0))
                if status == "HALF TIME BREAK":
                    return True
                if status == "IN PLAY" and elapsed >= 45:
                    return True
                return False
        # Fixture nicht mehr in der Live-Liste -> Spiel vermutlich vorbei, Auswertung versuchen
        return True
    except Exception as e:
        print(f"  [Nachschau] HZ1-Bereitschaftscheck Fehler: {e}")
        return True  # im Zweifel nicht blockieren, normale Auswertungs-Logik greift danach ohnehin

def match_bereit_fuer_volle_auswertung(match_id: str) -> bool:
    """
    v59.28 NEU: Dasselbe Prinzip wie hz1_bereit_fuer_auswertung, aber für alle Wett-Typen,
    die das GESAMTE Spielende brauchen (Ecken, Torwart, Druck, Comeback, Torflut, VZ-Tore,
    HZ2-Tore). Vorher wartete der Nachschau-Bot pauschal 15 Minuten und hatte dann nur noch
    30 Minuten (20 Versuche à 90s) Zeit – bei einem Signal aus Minute 20 reicht das bei
    Weitem nicht bis zum tatsächlichen Spielende (~90+ Min). Das 20-Versuche-Budget wurde
    dadurch oft schon verbraucht, BEVOR das Spiel überhaupt fertig war -> "nicht auswertbar",
    obwohl eine spätere Auswertung problemlos möglich gewesen wäre. Jetzt: Das Budget startet
    erst zu laufen, sobald das Spiel laut Live-Status wirklich kurz vor Schluss oder bereits
    vorbei ist.
    """
    try:
        live = get_live_matches()
        for m in live:
            if str(m.get("id")) == str(match_id):
                status  = m.get("status","")
                elapsed = _safe_int(m.get("time",0))
                if status == "IN PLAY" and elapsed >= 88:
                    return True  # tief in der Nachspielzeit, Spielende steht unmittelbar bevor
                return False  # läuft erkennbar noch, noch nicht versuchen
        # Fixture nicht mehr in der Live-Liste -> Spiel vermutlich vorbei, Auswertung versuchen
        return True
    except Exception as e:
        print(f"  [Nachschau] Vollzeit-Bereitschaftscheck Fehler: {e}")
        return True  # im Zweifel nicht blockieren, normale Auswertungs-Logik greift danach ohnehin

def ls_get_match_result(match_id: str, home: str = "", away: str = "", liga: str = ""):
    ergebnisse = {}
    ht_score   = ""

    # v59.25 FIX (KRITISCH #2): Der Live-Listen-Check (v59.20) allein reicht nicht – die
    # große Sammel-Liste ALLER Live-Spiele von API-Football kann bei manchen Ligen kurzzeitig
    # Lücken haben (Fixture fehlt in der Liste), OBWOHL das Spiel nachweislich noch läuft.
    # Deshalb jetzt ZUSÄTZLICH ein direkter, gezielter Status-Check NUR für dieses eine
    # Fixture (zuverlässiger als die große Liste, weniger anfällig für Pagination/Timing-
    # Lücken). Dieser direkte Abruf wird gleich als "Quelle 1" wiederverwendet, kostet also
    # keinen zusätzlichen API-Call.
    direct_match = None
    try:
        direct_match  = ls_get_single_match(match_id)
        direct_status = str(direct_match.get("status","") or "")
        if direct_status and direct_status not in FT_STATI_SET:
            print(f"  [Triple] Fixture {match_id}: direkter Status '{direct_status}' (nicht beendet) – kein Ergebnis akzeptiert")
            return None
    except Exception as e:
        print(f"  [Triple] Direkter Status-Check Fehler: {e}")

    # v59.19/v59.20 FIX (KRITISCH #1): Zusätzlich weiterhin die Bulk-Live-Liste prüfen –
    # TheSportsDB (Quelle 3 weiter unten) liefert Spielstände auch für noch LAUFENDE Spiele,
    # ohne das zu kennzeichnen. Bisher wurde jedes Ergebnis, das NUR von einer einzigen
    # Quelle kam, blind als "Spiel beendet" (FT) akzeptiert.
    try:
        live_ids_check = {str(m.get("id")) for m in get_live_matches()}
        if str(match_id) in live_ids_check:
            print(f"  [Triple] Fixture {match_id} läuft laut Live-Liste noch – kein Ergebnis akzeptiert")
            return None
    except Exception as e:
        print(f"  [Triple] Live-Listen-Check Fehler: {e}")

    # v59.12 DIAGNOSE: Zeigt bei jedem Versuch den rohen Status/Score von API-Football sowie
    # den aktuellen Tages-API-Verbrauch – damit im Log sichtbar wird, OB das Problem am
    # Tageslimit liegt (dann wäre _api_monitor hoch) oder daran, dass API-Football für diese
    # Fixture schlicht (noch) keinen "FT"-Status/Score zurückgibt.
    print(f"  [Triple-Debug] Fixture {match_id} | API heute: {_api_monitor.get('heute',0)}/{API_DAILY_LIMIT}")

    try:
        match    = direct_match if direct_match else ls_get_single_match(match_id)  # v59.25: wiederverwendet statt erneut abzurufen
        status   = str(match.get("status","") or "").upper().strip()
        time_val = str(match.get("time","") or "").upper().strip()
        score    = (match.get("scores") or {}).get("score","")
        ht_raw   = (match.get("scores") or {}).get("ht_score","") or ""
        print(f"  [Triple-Debug] Fixture {match_id} roh: status='{status}' time='{time_val}' score='{score}'")
        if time_val in ("FT","FULL TIME","AET"):
            status = "FT"
        if status in FT_STATI_SET and score:
            ergebnisse["livescore"] = score
            if ht_raw:
                ht_score = ht_raw
            print(f"  [Triple] Quelle 1 livescore: {score} ✅")
    except Exception as e:
        print(f"  [Triple] Livescore Fehler: {e}")

    if not ergebnisse:
        try:
            live      = get_live_matches()
            live_ids  = {str(m.get("id")) for m in live}
            if match_id not in live_ids:
                print(f"  [Triple] {match_id} nicht mehr in Live-Liste → Retry single-match")
                match2    = ls_get_single_match(match_id)
                status2   = str(match2.get("status","") or "").upper().strip()
                score2    = (match2.get("scores") or {}).get("score","")
                ht2       = (match2.get("scores") or {}).get("ht_score","") or ""
                if score2:
                    ergebnisse["livescore_retry"] = score2
                    if ht2:
                        ht_score = ht2
                    print(f"  [Triple] Retry-Score: {score2} ✅")
        except Exception as e:
            print(f"  [Triple] Live-Check Fehler: {e}")

    if home and away:
        try:
            fd_result = fd_suche_spiel(home,away,liga)
            if fd_result:
                ergebnisse["football_data"] = fd_result["score"]
                if fd_result.get("ht_score") and not ht_score:
                    ht_score = fd_result["ht_score"]
                print(f"  [Triple] Quelle 2 football-data: {fd_result['score']} ✅")
        except Exception as e:
            print(f"  [Triple] football-data Fehler: {e}")

    if home and away and len(ergebnisse) < 2:
        try:
            tsdb = thesportsdb_suche_spiel(home,away)
            if tsdb:
                ergebnisse["thesportsdb"] = tsdb["score"]
                print(f"  [Triple] Quelle 3 TheSportsDB: {tsdb['score']} ✅")
        except Exception as e:
            print(f"  [Triple] TheSportsDB Fehler: {e}")

    from collections import Counter
    zaehler_bisher = Counter(ergebnisse.values())
    bestaetigt     = zaehler_bisher.most_common(1)[0][1] if zaehler_bisher else 0
    if bestaetigt < 2:
        try:
            live      = get_live_matches()
            live_ids2 = {str(m.get("id")) for m in live}
            if match_id not in live_ids2:
                events  = ls_get_events(match_id)
                tore_h  = len([e for e in events if e.get("event") in ("Goal","goal") and e.get("home_away") == "home"])
                tore_a  = len([e for e in events if e.get("event") in ("Goal","goal") and e.get("home_away") == "away"])
                if events:
                    ev_score = f"{tore_h} - {tore_a}"
                    ergebnisse["events"] = ev_score
                    print(f"  [Triple] Quelle 4 Events: {ev_score} ✅")
        except Exception as e:
            print(f"  [Triple] Events Fehler: {e}")

    if not ergebnisse:
        print(f"  [Triple] ❌ Keine Quelle hat Ergebnis geliefert für {match_id}")
        return None

    zaehler = Counter(ergebnisse.values())
    bestes  = zaehler.most_common(1)[0]
    bester_score,anzahl = bestes

    if anzahl >= 2:
        print(f"  [Triple] ✅ BESTÄTIGT ({anzahl}/{len(ergebnisse)} Quellen): {bester_score}")
        return {"status":"FT","score":bester_score,"ht_score":ht_score,
                "quelle":f"triple_verified ({anzahl}/{len(ergebnisse)})",
                "alle_quellen":ergebnisse}

    if "livescore" in ergebnisse or "livescore_retry" in ergebnisse:
        key_ls  = "livescore" if "livescore" in ergebnisse else "livescore_retry"
        score_ls = ergebnisse[key_ls]
        print(f"  [Triple] ⚠️ Nur Livescore-Quelle, FT-Status bestätigt: {score_ls}")
        return {"status":"FT","score":score_ls,"ht_score":ht_score,
                "quelle":"livescore_single","alle_quellen":ergebnisse}

    if len(ergebnisse) == 1:
        einziger = list(ergebnisse.values())[0]
        quelle   = list(ergebnisse.keys())[0]
        print(f"  [Triple] ⚠️ Nur 1 Quelle ({quelle}): {einziger}")
        return {"status":"FT","score":einziger,"ht_score":ht_score,
                "quelle":f"single_source_{quelle}","alle_quellen":ergebnisse}

    print(f"  [Triple] ⚠️ Quellen uneinig: {ergebnisse} – warte auf mehr Daten")
    return None

# ============================================================
#  NACHSCHAU-BOT (90s Retry, 20 Versuche)
# ============================================================

def bot_nachschau():
    """
    EINZIGER Auswertungs-Bot.
    Retry alle 90s, max. 20 Versuche.
    v59: setzt nach jeder Auswertung zusätzlich eine ✅/❌-Reaktion direkt auf
    die ursprüngliche Signal-Nachricht in Discord (siehe discord_add_reaction).
    """
    print("[Nachschau-Bot] Gestartet | 90s Retry, 20 Versuche max")

    AUSWERTUNG_FNS = {
        "ecken":      auswertung_ecken,
        "ecken_over": None,
        "karten":     None,
        "torwart":    auswertung_torwart,
        "druck":      auswertung_druck,
        "comeback":   auswertung_comeback,
        "torflut":    auswertung_torflut,
        "rotkarte":   None,
        "hz1tore":    auswertung_hz1tore,
        "vztore":     auswertung_vztore,
        "hz2tore":    auswertung_hz2tore,  # v59.22 NEU
    }

    while True:
        try:
            wende_konfidenz_decay_an()
            offene = tracker_get_offene()
            if offene:
                print(f"[{jetzt()}] [Nachschau-Bot] {len(offene)} offene Signale prüfen")

            for key,sig in offene:
                match_id = sig.get("match_id","")
                typ      = sig.get("typ","")
                home     = sig.get("home","?")
                away     = sig.get("away","?")
                webhook  = sig.get("webhook","")

                # v59.28 FIX: Die pauschale "15 Min warten"-Regel reichte auch für alle
                # anderen Typen nicht (nicht nur HZ1) – Ecken/Torwart/Druck/Comeback/Torflut/
                # VZ-Tore/HZ2-Tore brauchen alle das ECHTE (nahezu) Spielende, nicht nur eine
                # feste Wartezeit seit dem Signal. Jetzt wird für JEDEN Typ der tatsächliche
                # Live-Status geprüft, bevor das 20-Versuche-Budget zu laufen beginnt.
                if typ == "hz1tore":
                    if not hz1_bereit_fuer_auswertung(match_id):
                        continue
                else:
                    if not match_bereit_fuer_volle_auswertung(match_id):
                        continue

                letzter = sig.get("letzter_versuch",0)
                if time.time()-letzter < 90:
                    continue

                with _tracker_lock:
                    if key in _signal_tracker:
                        _signal_tracker[key]["versuche"]       += 1
                        _signal_tracker[key]["letzter_versuch"] = time.time()

                versuche = sig.get("versuche",0)+1
                print(f"  [Nachschau] Prüfe: {home} vs {away} ({typ}) | Versuch #{versuche}")

                liga_name = sig.get("competition",sig.get("liga",""))
                result = ls_get_match_result(
                    match_id, home=home, away=away,
                    liga=liga_name
                )
                if not result:
                    if versuche >= 20:
                        print(f"  [Nachschau] ❓ Aufgegeben nach {versuche} Versuchen (kein Ergebnis gefunden): {home} vs {away} ({typ})")
                        tracker_nicht_auswertbar_markieren(key)
                        liga_auswertung_update(liga_name, False)  # v59.15
                        msg_id_na  = sig.get("discord_message_id")
                        chan_id_na = sig.get("discord_channel_id")
                        if msg_id_na and chan_id_na:
                            discord_add_reaction(chan_id_na, msg_id_na, "%E2%9D%93")  # ❓
                        send_telegram(f"❓ <b>Nicht auswertbar</b>\n📌 {home} vs {away} ({typ})\n"
                                      f"Ergebnis konnte nach 30 Min über keine Quelle sicher bestätigt werden.\n"
                                      f"🕐 {jetzt()} Uhr")
                    else:
                        print(f"  [Nachschau] {home} vs {away} | Noch kein Ergebnis (Versuch {versuche}/20)")
                    continue

                cache_invalidieren(match_id)
                time.sleep(8)

                auswert_fn = AUSWERTUNG_FNS.get(typ)
                if not auswert_fn:
                    print(f"  [Nachschau] Typ '{typ}' hat keine Auswertungsfunktion – übersprungen")
                    tracker_ausgewertet_markieren(key,False)
                    continue

                msg = None
                try:
                    msg = auswert_fn(sig,ft_result=result)
                except Exception as e:
                    print(f"  [Nachschau] Auswertungs-Fehler {home} vs {away} ({typ}): {e}")

                if not msg:
                    if versuche >= 20:
                        print(f"  [Nachschau] ❓ Kein Auswertungs-Ergebnis nach {versuche} Versuchen: {home} vs {away} ({typ})")
                        tracker_nicht_auswertbar_markieren(key)
                        liga_auswertung_update(liga_name, False)  # v59.15
                        msg_id_na  = sig.get("discord_message_id")
                        chan_id_na = sig.get("discord_channel_id")
                        if msg_id_na and chan_id_na:
                            discord_add_reaction(chan_id_na, msg_id_na, "%E2%9D%93")  # ❓
                        send_telegram(f"❓ <b>Nicht auswertbar</b>\n📌 {home} vs {away} ({typ})\n"
                                      f"Ergebnis lag vor, konnte aber nicht zuverlässig zu einem Gewonnen/Verloren ausgewertet werden.\n"
                                      f"🕐 {jetzt()} Uhr")
                    continue

                gewonnen = "GEWONNEN" in msg
                emoji    = "✅" if gewonnen else "❌"
                send_telegram(msg)

                def _extrahiere_feld(label: str) -> str:
                    m = re.search(rf"{label}:\s*<b>([^<]+)</b>",msg)
                    return m.group(1).strip() if m else "?"

                def _quote_gewinn_feld():
                    quote = sig.get("quote")
                    if not quote:
                        return None
                    einsatz = sig.get("einsatz",EINSATZ)
                    if gewonnen:
                        gewinn = round((quote-1)*einsatz,2)
                        return f"**{quote}** | Einsatz {einsatz}€ → +{gewinn}€"
                    return f"**{quote}** | Einsatz {einsatz}€ → -{einsatz}€"

                if typ in ("ecken","ecken_over"):
                    hz1    = sig.get("hz1_ecken",0)
                    grenze = hz1*2+1 if typ == "ecken" else 14
                    total  = _extrahiere_feld("Tatsächlich")
                    details = {"📐 Ecken bei Signal":f"**{hz1}**","🎯 Tipp":f"{'Unter' if typ=='ecken' else 'Über'} **{grenze}** Ecken","📈 Endstand":f"**{total}** Ecken"}
                elif typ == "torwart":
                    details = {"🎯 Tipp":"Mind. **1 Tor**","📈 Endstand":f"**{_extrahiere_feld('Endstand')}**"}
                elif typ == "druck":
                    details = {"🔥 Druck-Team":f"**{sig.get('druck_team','?')}**","📊 Stand bei Signal":f"**{sig.get('score_signal','?')}**","📈 Ergebnis":f"{emoji} Tor erzielt" if gewonnen else "❌ Kein Tor"}
                elif typ == "comeback":
                    details = {"🔄 Rückliegend":f"**{sig.get('rueckliegend','?')}**","🎯 Tipp":"Beide Teams treffen","📈 Endstand":f"**{_extrahiere_feld('Endstand')}**"}
                elif typ == "torflut":
                    details = {"⚽ Tore HZ1":f"**{sig.get('hz1_tore','?')}**","🎯 Tipp":f"Über **{sig.get('grenze','?')}** Tore","📈 Endstand":f"**{_extrahiere_feld('Endstand')}**"}
                elif typ in ("hz1tore","vztore","hz2tore"):
                    label = "HZ1" if typ == "hz1tore" else ("HZ2" if typ == "hz2tore" else "Vollzeit")
                    details = {"🎯 Tipp":f"**{sig.get('richtung','?').capitalize()} {sig.get('linie','?')}** Tore ({label})","📈 Ergebnis":f"{emoji} {'Gewonnen' if gewonnen else 'Verloren'}"}
                else:
                    details = {"📊 Typ":f"**{typ.upper()}**"}

                qg_feld = _quote_gewinn_feld()
                if qg_feld:
                    details["💶 Quote / Gewinn"] = qg_feld

                embed = discord_auswertung(typ,home,away,gewonnen,details)
                send_discord_embed(webhook,embed)

                # ── v59 NEU (Fix 1): Reaktion direkt auf die Original-Signal-Nachricht ──
                msg_id  = sig.get("discord_message_id")
                chan_id = sig.get("discord_channel_id")
                if msg_id and chan_id:
                    reaction_emoji = "%E2%9C%85" if gewonnen else "%E2%9D%8C"   # ✅ bzw. ❌
                    discord_add_reaction(chan_id, msg_id, reaction_emoji)

                try:
                    einsatz = sig.get("einsatz",EINSATZ)
                    quote   = sig.get("quote")
                    bankroll_aktualisieren(gewonnen,einsatz,quote)
                except Exception as e:
                    print(f"  [Nachschau] Bankroll-Fehler: {e}")

                tracker_ausgewertet_markieren(key,gewonnen)
                liga_auswertung_update(liga_name, True)  # v59.15: erfolgreich ausgewertet (unabhängig von gewonnen/verloren)
                check_streak_alarm()

                if not gewonnen:
                    threading.Thread(target=claude_verloren_analyse,args=(home,away,typ,msg),daemon=True).start()

                print(f"  [Nachschau] {emoji} Ausgewertet: {home} vs {away} ({typ}) → {'GEWONNEN' if gewonnen else 'VERLOREN'}")
                time.sleep(1)

            tracker_speichern()
            bot_fehler_reset("Nachschau-Bot")
        except Exception as e:
            bot_fehler_melden("Nachschau-Bot",e)
        time.sleep(90)

def wende_konfidenz_decay_an():
    now = time.time()
    with _tracker_lock:
        for key,sig in _signal_tracker.items():
            if sig.get("status") != "offen":
                continue
            alter_min = (now-sig.get("signal_zeit",now))/60
            if alter_min < 30:
                continue
            decay_stufen = int((alter_min-30)/15)
            orig_konfidenz = sig.get("konfidenz_original",sig.get("konfidenz",6))
            if "konfidenz_original" not in sig:
                _signal_tracker[key]["konfidenz_original"] = orig_konfidenz
            _signal_tracker[key]["konfidenz"] = max(1,orig_konfidenz-decay_stufen)

def check_streak_alarm():
    global _streak_alarm_gesendet
    if abs(streak_aktuell) >= 5:
        typ = "positiv" if streak_aktuell > 0 else "negativ"
        if _streak_alarm_gesendet.get(typ) == streak_aktuell:
            return
        _streak_alarm_gesendet[typ] = streak_aktuell
        if streak_aktuell >= 5:
            emoji = "🔥"; titel = f"Hot Streak! {streak_aktuell} Tipps in Folge GEWONNEN!"; farbe = 0x2ECC71
        else:
            emoji = "❄️"; titel = f"Cold Streak! {abs(streak_aktuell)} Tipps in Folge VERLOREN!"; farbe = 0xE74C3C
        msg = (f"{emoji} <b>{titel}</b>\n━━━━━━━━━━━━━━━━━━━━\n"
               f"Streak: <b>{streak_aktuell}</b> | Bester: <b>{streak_beste}</b>\n"
               f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
        send_telegram(msg)

_streak_alarm_gesendet = {}


# ============================================================
#  xG BOT
# ============================================================

notified_xg = set()

def bot_xg():
    print("[xG-Bot] Gestartet | Expected Goals Analyse")
    while True:
        try:
            matches = get_live_matches()
            laufend = [m for m in matches if m.get("status") in ("IN PLAY","ADDED TIME")
                       and 30 <= _safe_int(m.get("time",0)) <= SPAETESTE_SIGNAL_MINUTE]
            print(f"[{jetzt()}] [xG-Bot] {len(laufend)} Spiele geprüft")
            for game in laufend:
                match_id = str(game.get("id"))
                if match_id in notified_xg:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","?")
                score   = game.get("scores",{}).get("score","?")
                minute  = _safe_int(game.get("time",0))
                h_tore,a_tore = parse_score(score)
                tore_ges = h_tore+a_tore
                stats    = get_statistiken(match_id)
                shots_h  = stats["shots_on_target_home"]
                shots_a  = stats["shots_on_target_away"]
                if shots_h+shots_a == 0:
                    continue
                xg_h    = round(shots_h*0.33,2)
                xg_a    = round(shots_a*0.33,2)
                xg_ges  = round(xg_h+xg_a,2)
                xg_diff = round(xg_ges-tore_ges,2)
                if xg_diff < 1.5:
                    continue
                notified_xg.add(match_id)
                notified_sets_speichern()
                # v59.7/8: Quote + Einsatz-Empfehlung + Konfidenz ergänzt
                quote_xg   = get_quote(match_id,"xg",linie=tore_ges+0.5,richtung="über")  # v59.18: passende Linie
                einsatz_xg = kelly_einsatz_bankroll(quote_xg,"torflut") if quote_xg else EINSATZ
                konfidenz_xg = min(10,5+round(xg_diff))  # höhere xG-Differenz = höhere Konfidenz
                if konfidenz_xg < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke_xg        = konfidenz_emoji(konfidenz_xg)
                ql_xg      = f"\n💶 Quote: <b>{quote_xg}</b> | 💰 Einsatz: <b>{einsatz_xg}€</b>" if quote_xg else f"\n💰 Einsatz: <b>{einsatz_xg}€</b>"
                msg = (f"📊 <b>xG Signal!</b> {ke_xg} Konfidenz: <b>{konfidenz_xg}/10</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                       f"📊 Stand: <b>{score}</b> | Min. <b>{minute}'</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"📈 xG Gesamt: <b>{xg_ges}</b> | Tore: <b>{tore_ges}</b>\n"
                       f"🎯 xG Differenz: <b>+{xg_diff}</b> ungenutzte Chancen\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"💡 Tipp: Noch mehr Tore wahrscheinlich{ql_xg}\n"
                       f"🕐 {jetzt()} Uhr")
                send_telegram(msg)
                embed = {
                    "title":"📊 xG Signal – Mehr Tore erwartet!","color":0x1ABC9C,
                    "fields":[
                        {"name":"🏆 Liga","value":comp,"inline":True},
                        {"name":"⏱️ Minute","value":f"**{minute}'**","inline":True},
                        {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
                        {"name":"📈 xG gesamt","value":f"**{xg_ges}**","inline":True},
                        {"name":"⚽ Echte Tore","value":f"**{tore_ges}**","inline":True},
                        {"name":"💎 xG Diff","value":f"**+{xg_diff}**","inline":True},
                        {"name":"💡 Tipp","value":"Mehr Tore wahrscheinlich","inline":False},
                        {"name":"💶 Quote","value": (f"**{quote_xg}**" if quote_xg else "⚠️ Keine zuverlässige Quote"),"inline":True},
                        {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz_xg),"inline":True},
                        {"name":"🎯 Konfidenz","value":f"{ke_xg} **{konfidenz_xg}/10**","inline":True},
                    ],
                    "footer":{"text":f"xG-Bot • {heute()} {jetzt()}"},
                }
                send_discord_embed(DISCORD_WEBHOOK_TORWART,embed)
                print(f"  [xG-Bot] ✅ {home} vs {away} | xG {xg_ges} vs {tore_ges} Tore")
                time.sleep(0.5)
            bot_fehler_reset("xG-Bot")
        except Exception as e:
            bot_fehler_melden("xG-Bot",e)
        time.sleep(FUSSBALL_INTERVAL*60)

# ============================================================
#  EARLY GOAL BOT
# ============================================================

notified_early_goal = set()

def bot_early_goal():
    print("[EarlyGoal-Bot] Gestartet | Frühtore Min. 1-10")
    while True:
        try:
            matches = get_live_matches()
            frueh   = [m for m in matches if m.get("status") == "IN PLAY"
                       and 1 <= _safe_int(m.get("time",0)) <= 15]
            for game in frueh:
                match_id = str(game.get("id"))
                if match_id in notified_early_goal:
                    continue
                score_str = game.get("scores",{}).get("score","")
                h,a = parse_score(score_str)
                if h+a == 0:
                    continue
                try:
                    events = ls_get_events(match_id)
                    tore   = [e for e in events if e.get("event") in ("Goal","goal","GOAL")
                              and _safe_int(e.get("time",99)) <= 10]
                    if not tore:
                        continue
                except Exception:
                    if h+a == 0:
                        continue
                    tore = [{"time":"?"}]
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","?")
                minute  = game.get("time","?")
                if not whitelist_check(comp,home,away):
                    continue
                notified_early_goal.add(match_id)
                notified_sets_speichern()
                tor_min = tore[0].get("time","?") if tore else "?"
                quote   = get_quote(match_id,"tore",linie=2.5,richtung="über")  # v59.18: passende Linie
                einsatz = kelly_einsatz_bankroll(quote,"torflut") if quote else EINSATZ
                # v59.8: Konfidenz ergänzt – je früher das Tor, desto höher
                konfidenz = min(10,10-_safe_int(tor_min,5)//2)
                if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke        = konfidenz_emoji(konfidenz)
                ql      = f"\n💶 Quote: <b>{quote}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote else f"\n💰 Einsatz: <b>{einsatz}€</b>"
                msg = (f"⚡ <b>Early Goal!</b> {ke} Konfidenz: <b>{konfidenz}/10</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                       f"📊 Stand: <b>{score_str}</b> | Min. <b>{minute}'</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"⚡ Tor in Minute <b>{tor_min}'</b>!\n"
                       f"📈 Frühtore → 70%+ Chance auf 3+ Tore\n"
                       f"🎯 Tipp: <b>Über 2.5 Tore</b>{ql}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                send_telegram(msg)
                embed = {
                    "title":"⚡ Early Goal – Mehr Tore erwartet!","color":0xF39C12,
                    "fields":[
                        {"name":"🏆 Liga","value":comp,"inline":True},
                        {"name":"⏱️ Minute","value":f"**{minute}'**","inline":True},
                        {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
                        {"name":"📊 Stand","value":f"**{score_str}**","inline":True},
                        {"name":"⚡ Tor in","value":f"Min. **{tor_min}'**","inline":True},
                        {"name":"🎯 Tipp","value":"**Über 2.5 Tore**","inline":False},
                        {"name":"💶 Quote","value": (f"**{quote}**" if quote else "⚠️ Keine zuverlässige Quote"),"inline":True},
                        {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":True},
                        {"name":"🎯 Konfidenz","value":f"{ke} **{konfidenz}/10**","inline":True},
                    ],
                    "footer":{"text":f"EarlyGoal-Bot • {heute()} {jetzt()}"},
                }
                send_discord_embed(DISCORD_WEBHOOK_TORFLUT,embed)
                print(f"  [EarlyGoal-Bot] ✅ {home} vs {away} | Tor in Min. {tor_min}")
                time.sleep(0.5)
            bot_fehler_reset("EarlyGoal-Bot")
        except Exception as e:
            bot_fehler_melden("EarlyGoal-Bot",e)
        time.sleep(60)

# ============================================================
#  ANOMALIE BOT
# ============================================================

notified_anomalie = set()

def bot_anomalie_erkennung():
    print("[Anomalie-Bot] Gestartet | Statistische Ausreißer erkennen")
    while True:
        try:
            matches = get_live_matches()
            laufend = [m for m in matches if m.get("status") == "IN PLAY"
                       and _safe_int(m.get("time",0)) >= 15]
            for game in laufend:
                match_id = str(game.get("id"))
                if match_id in notified_anomalie:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                minute  = _safe_int(game.get("time",0))
                score   = game.get("scores",{}).get("score","?")
                stats   = get_statistiken(match_id)
                ecken_ges    = stats["corners_home"]+stats["corners_away"]
                schuesse_ges = stats["shots_on_target_home"]+stats["shots_on_target_away"]
                anomalien = []
                if minute <= 35 and ecken_ges >= 12:
                    anomalien.append(f"📐 {ecken_ges} Ecken in Minute {minute} (extrem hoch!)")
                if minute <= 25 and schuesse_ges >= 10:
                    anomalien.append(f"🎯 {schuesse_ges} Schüsse aufs Tor in Minute {minute}!")
                if not anomalien:
                    continue
                notified_anomalie.add(match_id)
                notified_sets_speichern()
                msg = (f"🚨 <b>Anomalie erkannt!</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp}\n📌 {home} vs {away}\n"
                       f"📊 Stand: <b>{score}</b> | Min. <b>{minute}'</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"+"\n".join(anomalien)+
                       f"\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"💡 Statistisch außergewöhnlich!\n🕐 {jetzt()} Uhr")
                send_telegram(msg)
                embed = {
                    "title":"🚨 Statistische Anomalie!","color":0xFF0000,
                    "fields":[
                        {"name":"🏆 Liga","value":comp,"inline":True},
                        {"name":"⏱️ Minute","value":f"**{minute}'**","inline":True},
                        {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
                        {"name":"📊 Stand","value":f"**{score}**","inline":True},
                        {"name":"🚨 Anomalie","value":"\n".join(anomalien),"inline":False},
                    ],
                    "footer":{"text":f"Anomalie-Bot • {heute()} {jetzt()}"},
                }
                send_discord_embed(DISCORD_WEBHOOK_VALUE,embed)
                print(f"  [Anomalie] 🚨 {home} vs {away}")
            bot_fehler_reset("Anomalie-Bot")
        except Exception as e:
            bot_fehler_melden("Anomalie-Bot",e)
        time.sleep(90)

# ============================================================
#  VALUE BET BOT
# ============================================================

notified_value = set()

# v59.3 FIX: Jede Regel bekommt jetzt einen festen Ziel-Markt ("markt"). Vorher wurde
# beim Quoten-Matching nur nach Label ("Over"/"Under" + Linie) gesucht, ohne zu prüfen,
# aus welchem Wettmarkt die Quote überhaupt stammt. Dadurch konnten artfremde Märkte
# (z.B. "nur Team X trifft über X.5 Mal") mit zufällig gleichem Label reinrutschen und
# absurd hohe/falsche Quoten liefern (z.B. 34.0 für Tore Über 3.5 gesamt). Die Markt-Namen
# sind API-Football-Standardnamen; falls euer Plan andere Bezeichnungen liefert, bitte im
# API-Football-Dashboard unter "Odds" nachschauen und hier anpassen.
VALUE_MAX_PLAUSIBLE_QUOTE = 10.0  # Sicherheitsnetz: Quoten über diesem Wert werden verworfen

VALUE_REGELN = [
    {"typ":"karten_unter","name":"Karten Unter 4.5","markt":("Cards Over/Under",),
     "check":lambda s,ev,st:(60<=_safe_int(s.get("time",0))<=80 and len([e for e in ev if e.get("event") in KARTEN_TYPEN])<=1),
     "prob":lambda s,ev,st:0.82,"linie":4.5,"richtung":"unter"},
    {"typ":"ecken_unter","name":"Ecken Unter 9.5","markt":("Corners Over/Under","Corners Over Under"),
     "check":lambda s,ev,st:(_safe_int(s.get("time",0))>=70 and st["corners_home"]+st["corners_away"]<=6),
     "prob":lambda s,ev,st:0.80,"linie":9.5,"richtung":"unter"},
    {"typ":"tore_unter","name":"Tore Unter 1.5","markt":("Goals Over/Under","Over/Under"),
     "check":lambda s,ev,st:(_safe_int(s.get("time",0))>=70 and s.get("scores",{}).get("score","") in ("0 - 0","0-0")),
     "prob":lambda s,ev,st:0.75,"linie":1.5,"richtung":"unter"},
    # v59.13 FIX: Beide "über"-Regeln unten prüften vorher nur "mindestens X", nicht "genau
    # einen Schritt unter der Linie" – dadurch wurde z.B. bei 4 bereits gefallenen Toren immer
    # noch "Tore Über 3.5" empfohlen, obwohl der Tipp zu diesem Zeitpunkt schon gewonnen (also
    # gar kein echter Tipp mehr) war. Jetzt wird nur noch signalisiert, wenn die Grenze noch
    # NICHT erreicht ist (genau einen Treffer/Ecke drunter), der Ausgang also noch offen ist.
    {"typ":"tore_ueber","name":"Tore Über 3.5","markt":("Goals Over/Under","Over/Under"),
     "check":lambda s,ev,st:(_safe_int(s.get("time",0))>=60 and sum(parse_score(s.get("scores",{}).get("score","0-0")))==3),
     "prob":lambda s,ev,st:0.78,"linie":3.5,"richtung":"über"},
    {"typ":"ecken_ueber_live","name":"Ecken Über 11.5","markt":("Corners Over/Under","Corners Over Under"),
     "check":lambda s,ev,st:(_safe_int(s.get("time",0))>=60 and 9<=st["corners_home"]+st["corners_away"]<=10),
     "prob":lambda s,ev,st:0.72,"linie":11.5,"richtung":"über"},
]

def get_live_odds_fuer_spiel(match_id) -> dict:
    """v59.1: Quoten jetzt über API-Football (fixture-basiert), nicht mehr the-odds-api.com."""
    return af_odds_fuer_value_bot(match_id)

def berechne_value(prob: float, quote: float) -> float:
    return round(prob*quote-1,3)

def kelly_einsatz_value(prob: float, quote: float) -> float:
    """v59.4: Einsatz-Empfehlung für Value-Bets nach Kelly-Kriterium, basierend auf der
    geschätzten Wahrscheinlichkeit der Regel (statt auf der Trefferquoten-Historie wie
    kelly_einsatz_bankroll, die Value-Bets bisher nicht genutzt hat)."""
    br = bankroll_laden()
    b  = quote - 1.0
    if b <= 0:
        return EINSATZ
    kelly   = max(0,(b*prob-(1-prob))/b) * KELLY_FRACTION
    einsatz = round(kelly*br,2)
    return max(KELLY_MIN_EINSATZ,min(einsatz,KELLY_MAX_EINSATZ))

def bot_value_bet():
    print(f"[Value-Bot] Gestartet | Min. Quote {VALUE_BET_MIN_QUOTE} | Min. Edge {VALUE_BET_MIN_VALUE*100:.0f}%")
    while True:
        try:
            matches = get_live_matches()
            laufend = [m for m in matches if m.get("status") in ("IN PLAY","ADDED TIME","HALF TIME BREAK")]
            print(f"[{jetzt()}] [Value-Bot] {len(laufend)} Spiele geprüft")
            for game in laufend:
                match_id = str(game.get("id"))
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","?")
                minute  = _safe_int(game.get("time",0))
                score   = game.get("scores",{}).get("score","?")
                try:
                    events = get_events(match_id)
                    stats  = get_statistiken(match_id)
                except Exception:
                    continue
                for regel in VALUE_REGELN:
                    regel_key = f"{match_id}_{regel['typ']}"
                    if regel_key in notified_value:
                        continue
                    try:
                        if not regel["check"](game,events,stats):
                            continue
                    except Exception:
                        continue
                    prob  = regel["prob"](game,events,stats)
                    linie = regel["linie"]
                    rich  = regel["richtung"]
                    quoten = get_live_odds_fuer_spiel(match_id)
                    if not quoten:
                        continue
                    ziel_maerkte = regel.get("markt",())
                    beste_quote = None; bester_bm = ""
                    for k,v in quoten.items():
                        # v59.3 FIX: nur noch der zur Regel passende Wettmarkt wird durchsucht
                        # (vorher wurden artfremde Märkte mit zufällig gleichem Label mitgezählt)
                        if v.get("market") not in ziel_maerkte:
                            continue
                        try:
                            point_val = float(v.get("point",0))
                        except (TypeError, ValueError):
                            continue
                        if rich == "unter" and "Under" in v["name"] and abs(point_val-linie)<0.1:
                            if beste_quote is None or v["quote"] > beste_quote:
                                beste_quote = v["quote"]; bester_bm = v["bookmaker"]
                        elif rich == "über" and "Over" in v["name"] and abs(point_val-linie)<0.1:
                            if beste_quote is None or v["quote"] > beste_quote:
                                beste_quote = v["quote"]; bester_bm = v["bookmaker"]
                    if not beste_quote or beste_quote < VALUE_BET_MIN_QUOTE:
                        continue
                    # v59.3 FIX: Sicherheitsnetz gegen unrealistische/falsch zugeordnete Quoten
                    if beste_quote > VALUE_MAX_PLAUSIBLE_QUOTE:
                        print(f"  [Value-Bot] ⚠️ Unplausible Quote {beste_quote} für {regel['name']} ({home} vs {away}) – verworfen")
                        continue
                    edge = berechne_value(prob,beste_quote)
                    if edge < VALUE_BET_MIN_VALUE:
                        continue
                    notified_value.add(regel_key)
                    edge_pct = round(edge*100,1)
                    einsatz  = kelly_einsatz_value(prob,beste_quote)
                    msg = (f"💎 <b>VALUE BET GEFUNDEN!</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                           f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                           f"📊 Stand: <b>{score}</b> | Minute: <b>{minute}'</b>\n"
                           f"━━━━━━━━━━━━━━━━━━━━\n"
                           f"🎯 Tipp: <b>{regel['name']}</b>\n"
                           f"📈 Wahrscheinlichkeit: <b>{round(prob*100)}%</b>\n"
                           f"💶 Beste Quote: <b>{beste_quote}</b> ({bester_bm})\n"
                           f"💎 Value Edge: <b>+{edge_pct}%</b>\n"
                           f"💰 Einsatz: <b>{einsatz}€</b>\n"
                           f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                    send_telegram(msg)
                    embed = {
                        "title":"💎 Value Bet – Fehlerhafte Quote!","color":0xF1C40F,
                        "fields":[
                            {"name":"🏆 Liga","value":f"{comp}","inline":True},
                            {"name":"⏱️ Minute","value":f"**{minute}'**","inline":True},
                            {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
                            {"name":"🎯 Tipp","value":f"**{regel['name']}**","inline":False},
                            {"name":"📈 Wahrsch.","value":f"**{round(prob*100)}%**","inline":True},
                            {"name":"💶 Quote","value":f"**{beste_quote}** ({bester_bm})","inline":True},
                            {"name":"💎 Edge","value":f"**+{edge_pct}%**","inline":True},
                            {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":False},
                            {"name":"🎯 Konfidenz","value":f"{konfidenz_emoji(round(prob*10))} **{round(prob*10)}/10**","inline":True},
                        ],
                        "footer":{"text":f"Value-Bot • {heute()} {jetzt()}"},
                    }
                    send_discord_embed(DISCORD_WEBHOOK_VALUE,embed)
                    print(f"  [Value-Bot] ✅ {home} vs {away} | {regel['name']} | Edge +{edge_pct}%")
                time.sleep(0.5)
            bot_fehler_reset("Value-Bot")
        except Exception as e:
            bot_fehler_melden("Value-Bot",e)
        time.sleep(FUSSBALL_INTERVAL*60)

# ============================================================
#  ARBITRAGE BOT
# ============================================================

notified_arbitrage = set()

def finde_arbitrage() -> list:
    """
    v59.1: Sucht Arbitrage jetzt über die aktuell laufenden Spiele (API-Football
    Odds pro Fixture), statt wie vorher über einen Bulk-Abruf aller Spiele
    weltweit bei the-odds-api.com. Dadurch werden nur LIVE-Spiele erfasst,
    keine Pre-Match-Quoten mehr (API-Football liefert Odds fixture-basiert).
    """
    try:
        matches = get_live_matches()
    except Exception as e:
        print(f"  [Arbitrage] Fehler: {e}")
        return []
    arbs = []
    for m in matches:
        fixture_id = str(m.get("id"))
        home_t = m.get("home",{}).get("name","?")
        away_t = m.get("away",{}).get("name","?")
        werte = _af_odds_werte(af_get_odds(fixture_id), ("Goals Over/Under","Over/Under"))
        if not werte:
            continue
        beste_over  = {}
        beste_under = {}
        for w in werte:
            teile = w["label"].split(" ")
            if len(teile) < 2:
                continue
            richtung, linie = teile[0], teile[1]
            if richtung == "Over":
                if linie not in beste_over or w["odd"] > beste_over[linie]["q"]:
                    beste_over[linie] = {"q":w["odd"],"bm":w["bookmaker"]}
            elif richtung == "Under":
                if linie not in beste_under or w["odd"] > beste_under[linie]["q"]:
                    beste_under[linie] = {"q":w["odd"],"bm":w["bookmaker"]}
        for linie in beste_over:
            if linie not in beste_under:
                continue
            q_over  = beste_over[linie]["q"]
            q_under = beste_under[linie]["q"]
            margin  = round(1/q_over+1/q_under,4)
            if margin < 0.98:
                profit_pct = round((1-margin)*100,2)
                arbs.append({"home":home_t,"away":away_t,"linie":linie,
                             "q_over":q_over,"bm_over":beste_over[linie]["bm"],
                             "q_under":q_under,"bm_under":beste_under[linie]["bm"],
                             "margin":margin,"profit_pct":profit_pct})
    return sorted(arbs,key=lambda x:x["profit_pct"],reverse=True)

def bot_arbitrage():
    print("[Arbitrage-Bot] Gestartet | Suche Arbitrage-Möglichkeiten")
    while True:
        try:
            arbs = finde_arbitrage()
            print(f"[{jetzt()}] [Arbitrage-Bot] {len(arbs)} Arbitragen gefunden")
            for arb in arbs[:5]:
                key = f"{arb['home']}_{arb['away']}_{arb['linie']}"
                if key in notified_arbitrage:
                    continue
                notified_arbitrage.add(key)
                # v59.7: Einsatz-Split jetzt auf die aktuelle Bankroll skaliert (statt fix 100€)
                # und tatsächlich angezeigt (wurde vorher nur berechnet, nie ausgegeben).
                br            = bankroll_laden()
                einsatz_total = round(min(br*0.1,100),2)
                einsatz_over  = round(einsatz_total/arb["q_over"]/(1/arb["q_over"]+1/arb["q_under"]),2)
                einsatz_under = round(einsatz_total-einsatz_over,2)
                gewinn_sicher = round(einsatz_over*arb["q_over"]-einsatz_total,2)
                msg = (f"💰 <b>ARBITRAGE GEFUNDEN!</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"📌 {arb['home']} vs {arb['away']}\n"
                       f"📈 Over {arb['linie']}: <b>{arb['q_over']}</b> @ {arb['bm_over']}\n"
                       f"📉 Under {arb['linie']}: <b>{arb['q_under']}</b> @ {arb['bm_under']}\n"
                       f"💎 Risikoloser Gewinn: <b>+{arb['profit_pct']}%</b>\n"
                       f"💰 Einsatz-Split: <b>{einsatz_over}€</b> auf Over, <b>{einsatz_under}€</b> auf Under (Gesamt {einsatz_total}€)\n"
                       f"✅ Garantierter Gewinn: <b>+{gewinn_sicher}€</b>\n"
                       f"🕐 {jetzt()} Uhr")
                send_telegram(msg)
                embed = {
                    "title":"💰 Arbitrage – Risikoloser Gewinn!","color":0xF1C40F,
                    "fields":[
                        {"name":"⚽ Spiel","value":f"**{arb['home']}** vs **{arb['away']}**","inline":False},
                        {"name":"📈 Over","value":f"**{arb['q_over']}** @ {arb['bm_over']}","inline":True},
                        {"name":"📉 Under","value":f"**{arb['q_under']}** @ {arb['bm_under']}","inline":True},
                        {"name":"💎 Profit","value":f"**+{arb['profit_pct']}%**","inline":True},
                        {"name":"💰 Einsatz-Empfehlung","value":f"**{einsatz_over}€** Over / **{einsatz_under}€** Under (Σ {einsatz_total}€)","inline":False},
                        {"name":"✅ Garantiert","value":f"**+{gewinn_sicher}€**","inline":True},
                        {"name":"🎯 Konfidenz","value":"🔒 **10/10** (mathematisch risikofrei)","inline":True},
                    ],
                    "footer":{"text":f"Arbitrage-Bot • {heute()} {jetzt()}"},
                }
                send_discord_embed(DISCORD_WEBHOOK_VALUE,embed)
                print(f"  [Arbitrage] ✅ {arb['home']} vs {arb['away']} | +{arb['profit_pct']}%")
            bot_fehler_reset("Arbitrage-Bot")
        except Exception as e:
            bot_fehler_melden("Arbitrage-Bot",e)
        time.sleep(10*60)

# ============================================================
#  SHARP MONEY BOT
# ============================================================

notified_sharp = set()
_sharp_history = {}

def bot_sharp_money():
    """
    v59.1: Beobachtet Quotenbewegungen jetzt über API-Football-Odds (Markt
    'Match Winner') pro laufendem Spiel, statt wie vorher über einen
    Bulk-Abruf bei the-odds-api.com.
    """
    print("[Sharp-Money-Bot] Gestartet | Pro-Wetter Bewegungen")
    while True:
        try:
            matches = get_live_matches()
            now = time.time()
            for m in matches:
                fixture_id = str(m.get("id"))
                home_t = m.get("home",{}).get("name","?")
                away_t = m.get("away",{}).get("name","?")
                key    = fixture_id
                werte  = _af_odds_werte(af_get_odds(fixture_id), ("Match Winner",))
                if not werte:
                    continue
                bm_quoten = {}
                for w in werte:
                    bm_quoten.setdefault(w["label"],[]).append({"q":w["odd"],"bookmaker":w["bookmaker"]})
                if key not in _sharp_history:
                    _sharp_history[key] = {"ts":now,"quoten":bm_quoten}
                    continue
                alt = _sharp_history[key]
                if now-alt["ts"] < 8*60:
                    continue
                sharp_signals = []
                for k,quoten_liste in bm_quoten.items():
                    if k not in alt["quoten"]:
                        continue
                    q_neu = sum(v["q"] for v in quoten_liste)/max(len(quoten_liste),1)
                    q_alt = sum(v["q"] for v in alt["quoten"][k])/max(len(alt["quoten"][k]),1)
                    if q_alt <= 1.1 or q_neu <= 1.1:
                        continue
                    bewegung = round((q_neu-q_alt)/q_alt*100,1)
                    if abs(bewegung) >= 8 and len(quoten_liste) >= 3:
                        richtung = "📉 gefallen" if bewegung < 0 else "📈 gestiegen"
                        sharp_signals.append(f"{k}: {q_alt:.2f}→{q_neu:.2f} ({bewegung:+.1f}%) {richtung}")
                _sharp_history[key] = {"ts":now,"quoten":bm_quoten}
                if not sharp_signals or key in notified_sharp:
                    continue
                notified_sharp.add(key)
                notified_sets_speichern()
                msg = (f"💼 <b>Sharp Money Signal!</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"📌 {home_t} vs {away_t}\n━━━━━━━━━━━━━━━━━━━━\n"+
                       "\n".join(sharp_signals)+
                       f"\n━━━━━━━━━━━━━━━━━━━━\n💡 Professionelle Wetter aktiv!\n🕐 {jetzt()} Uhr")
                send_telegram(msg)
                embed = {
                    "title":"💼 Sharp Money – Pro-Wetter aktiv!","color":0x8E44AD,
                    "fields":[
                        {"name":"⚽ Spiel","value":f"{home_t} vs {away_t}","inline":False},
                        {"name":"📊 Bewegungen","value":"\n".join(sharp_signals),"inline":False},
                    ],
                    "footer":{"text":f"Sharp-Bot • {heute()} {jetzt()}"},
                }
                send_discord_embed(DISCORD_WEBHOOK_VALUE,embed)
                print(f"  [Sharp] 💼 {home_t} vs {away_t}")
            bot_fehler_reset("Sharp-Money-Bot")
        except Exception as e:
            bot_fehler_melden("Sharp-Money-Bot",e)
        time.sleep(8*60)


# ============================================================
#  HZ2 TORE BOT
# ============================================================

notified_hz2 = set()

def bot_hz2_tore():
    print("[HZ2-Tore-Bot] Gestartet | 0:0 HZ + Druck → Tor in HZ2")
    while True:
        try:
            matches = get_live_matches()
            pausen  = [m for m in matches if m.get("status") in
                       ("HALF TIME","HT","Half Time","half time","HALFTIME","Half-Time","HALF TIME BREAK")]
            print(f"[{jetzt()}] [HZ2-Tore-Bot] {len(pausen)} Halbzeit-Spiele")
            for game in pausen:
                match_id = str(game.get("id"))
                if match_id in notified_hz2:
                    continue
                score_str = game.get("scores",{}).get("score","")
                h,a = parse_score(score_str)
                if h+a != 0:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","?")
                if not liga_signal_erlaubt(comp):  # v59.15: Liga-Zuverlässigkeit
                    continue
                if not liga_erlaubt(comp):  # v59.17: Liga-Trefferquote
                    continue
                if not whitelist_check(comp,home,away):
                    continue
                stats     = get_statistiken(match_id)
                shots_h   = stats["shots_on_target_home"]
                shots_a   = stats["shots_on_target_away"]
                corners_h = stats["corners_home"]
                corners_a = stats["corners_away"]
                druck_ges = shots_h+shots_a+corners_h+corners_a
                if druck_ges < 8:
                    continue
                notified_hz2.add(match_id)
                notified_sets_speichern()
                # v59.7 FIX: Quote/Einsatz ergänzt + Message-ID wird jetzt erfasst, damit dieser
                # Bot (typ="hz1tore" -> wird vom Nachschau-Bot ausgewertet) auch eine ✅/❌-Reaktion
                # bekommt, wie die anderen 6 auswertbaren Signal-Typen.
                quote     = get_quote(match_id,"hz2",linie=0.5,richtung="über")  # v59.18: passende Linie
                einsatz   = kelly_einsatz_bankroll(quote,"hz2tore") if quote else EINSATZ  # v59.22 FIX: war fälschlich "hz1tore"
                ql        = f"\n💶 Quote: <b>{quote}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote else f"\n💰 Einsatz: <b>{einsatz}€</b>"
                konfidenz = min(10,6+(1 if druck_ges >= 12 else 0)+(1 if druck_ges >= 16 else 0))
                if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke        = konfidenz_emoji(konfidenz)
                msg = (f"⚡ <b>Über 0.5 HZ2 Tore!</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                       f"📊 HZ1: <b>0:0</b> – Halbzeit!\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"🎯 Schüsse: {shots_h}|{shots_a} | Ecken: {corners_h}|{corners_a}\n"
                       f"⚡ Druck-Score: <b>{druck_ges}</b>\n"
                       f"📈 >75% Chance auf mind. 1 Tor in HZ2\n"
                       f"🎯 Tipp: <b>Über 0.5 HZ2 Tore</b>{ql}\n"
                       f"{ke} Konfidenz: <b>{konfidenz}/10</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                send_telegram(msg)
                embed = {
                    "title":"⚡ 0:0 HZ + Druck → Tor in HZ2!","color":0x9B59B6,
                    "fields":[
                        {"name":"🏆 Liga","value":comp,"inline":True},
                        {"name":"🌍 Land","value":country,"inline":True},
                        {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
                        {"name":"📊 HZ1","value":"**0:0**","inline":True},
                        {"name":"🎯 Schüsse","value":f"{shots_h}|{shots_a}","inline":True},
                        {"name":"📐 Ecken","value":f"{corners_h}|{corners_a}","inline":True},
                        {"name":"⚡ Druck","value":f"**{druck_ges}**","inline":True},
                        {"name":ke+" Konfidenz","value":f"**{konfidenz}/10**","inline":True},
                        {"name":"🎯 Tipp","value":"**Über 0.5 HZ2 Tore**","inline":False},
                        {"name":"💶 Quote","value": (f"**{quote}**" if quote else "⚠️ Keine zuverlässige Quote"),"inline":True},
                        {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":True},
                    ],
                    "footer":{"text":f"HZ2-Tore-Bot • {heute()} {jetzt()}"},
                }
                dc_info = send_discord_embed(DISCORD_WEBHOOK_TORE,embed,wait_for_message=True)
                beobachtung_hinzufuegen(match_id,{
                    "typ":"hz2tore","match_id":match_id,"home":home,"away":away,  # v59.22 FIX: war fälschlich "hz1tore"
                    "richtung":"ueber","linie":0.5,"score_signal":"0 - 0",
                    "quote":quote,"einsatz":einsatz,"konfidenz":konfidenz,"webhook":DISCORD_WEBHOOK_TORE,
                    "signal_zeit":time.time(),"bot":"HZ2-Tore-Bot",
                    "discord_message_id": (dc_info or {}).get("message_id"),
                    "discord_channel_id": (dc_info or {}).get("channel_id"),
                })
                print(f"  [HZ2-Tore-Bot] ✅ {home} vs {away} | Druck {druck_ges}")
                time.sleep(0.5)
            bot_fehler_reset("HZ2-Tore-Bot")
        except Exception as e:
            bot_fehler_melden("HZ2-Tore-Bot",e)
        time.sleep(60)

# ============================================================
#  ROTE KARTE ECKEN BOT
# ============================================================

notified_rk_ecken = set()

def bot_rotkarte_ecken():
    print("[RotkarteEcken-Bot] Gestartet | Ecken nach Roter Karte")
    while True:
        try:
            matches = get_live_matches()
            laufend = [m for m in matches if m.get("status") in ("IN PLAY","ADDED TIME")
                       and _safe_int(m.get("time",0)) <= 75]
            for game in laufend:
                match_id = str(game.get("id"))
                if match_id in notified_rk_ecken:
                    continue
                try:
                    events = get_events(match_id)
                    rote   = [e for e in events if e.get("event") in ROTKARTE_TYPEN]
                except Exception:
                    continue
                if not rote:
                    continue
                minute    = _safe_int(game.get("time",0))
                letzte    = rote[-1]
                karte_min = _safe_int(letzte.get("time",99))
                if minute-karte_min > 8:
                    continue
                home    = game.get("home",{}).get("name","?")
                away    = game.get("away",{}).get("name","?")
                comp    = game.get("competition",{}).get("name","?")
                country = (game.get("country") or {}).get("name","?")
                score   = game.get("scores",{}).get("score","?")
                if not whitelist_check(comp,home,away):
                    continue
                karte_fuer    = letzte.get("home_away","")
                geschwaeches  = home if karte_fuer == "home" else away
                staerkeres    = away if karte_fuer == "home" else home
                geschw_id     = str((game.get("home") if karte_fuer=="home" else game.get("away") or {}).get("id",""))
                restminuten   = 90-minute
                if restminuten > 30:
                    continue
                stats = get_statistiken(match_id)
                ecken_schwach = stats["corners_home"] if karte_fuer == "home" else stats["corners_away"]
                ecken_stark   = stats["corners_away"] if karte_fuer == "home" else stats["corners_home"]
                team_avg = get_team_ecken_avg(geschw_id) if geschw_id else None
                if team_avg is not None:
                    erwartete_weitere = team_avg*(restminuten/90)*0.6
                    puffer = max(1,round(erwartete_weitere))
                else:
                    puffer = 1 if restminuten <= 15 else 2
                grenze_ecken  = ecken_schwach+puffer
                notified_rk_ecken.add(match_id)
                notified_sets_speichern()
                spieler = (letzte.get("player") or {}).get("name") or "?"
                # v59.6: Quote + Einsatz-Empfehlung ergänzt (fehlten bisher bei diesem Bot,
                # anders als bei Ecken/Torwart/Druck/Comeback/Torflut/HZ1-VZ-Tore).
                quote   = None  # v59.18 FIX: Ecken haben keinen passenden Tore-Markt – keine Quote statt geraten
                einsatz = kelly_einsatz_bankroll(quote,"rotkarte") if quote else EINSATZ
                # v59.8: Konfidenz ergänzt – je weniger Restzeit, desto sicherer bleibt's unter der Grenze
                konfidenz = min(10,max(1,4+(30-restminuten)//5))
                if konfidenz < MIN_KONFIDENZ_VERSAND:  # v59.17: Mindest-Konfidenz
                    continue
                ke        = konfidenz_emoji(konfidenz)
                ql      = f"\n💶 Quote: <b>{quote}</b> | 💰 Einsatz: <b>{einsatz}€</b>" if quote else f"\n💰 Einsatz: <b>{einsatz}€</b>"
                msg = (f"📐 <b>Rotkarte Ecken-Signal!</b> {ke} Konfidenz: <b>{konfidenz}/10</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🏆 {comp} ({country})\n📌 {home} vs {away}\n"
                       f"📊 Stand: <b>{score}</b> | Min. <b>{minute}'</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"🟥 {spieler} ({geschwaeches}) Rote Karte Min. {karte_min}'\n"
                       f"⏱️ Noch <b>{restminuten} Min</b>\n"
                       f"🎯 Tipp: <b>{geschwaeches} unter {grenze_ecken} Ecken gesamt</b>{ql}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                send_telegram(msg)
                embed = {
                    "title":"📐 Rotkarte → Unter Ecken!","color":0xC0392B,
                    "fields":[
                        {"name":"🏆 Liga","value":comp,"inline":True},
                        {"name":"📊 Stand","value":f"**{score}** | Min. {minute}'","inline":True},
                        {"name":"⚽ Spiel","value":f"{home} vs {away}","inline":False},
                        {"name":"🟥 Rote Karte","value":f"**{spieler}** ({geschwaeches})","inline":False},
                        {"name":"⏱️ Rest","value":f"**{restminuten} Min**","inline":True},
                        {"name":"🎯 Tipp","value":f"**{geschwaeches} unter {grenze_ecken} Ecken**","inline":False},
                        {"name":"💶 Quote","value": (f"**{quote}**" if quote else "⚠️ Keine zuverlässige Quote"),"inline":True},
                        {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz),"inline":True},
                        {"name":"🎯 Konfidenz","value":f"{ke} **{konfidenz}/10**","inline":True},
                    ],
                    "footer":{"text":f"RotkarteEcken-Bot • {heute()} {jetzt()}"},
                }
                send_discord_embed(DISCORD_WEBHOOK_ECKEN,embed)
                print(f"  [RotkarteEcken] ✅ {geschwaeches} in Unterzahl")
                time.sleep(0.5)
            bot_fehler_reset("RotkarteEcken-Bot")
        except Exception as e:
            bot_fehler_melden("RotkarteEcken-Bot",e)
        time.sleep(FUSSBALL_INTERVAL*60)


# ============================================================
#  HILFSFUNKTIONEN: WHITELIST, GITHUB, API-MONITOR, WETTER
# ============================================================

WHITELIST_DATEI = "whitelist.json"
MANUELL_DATEI   = "manuell_tipps.json"
_whitelist      = {"ligen":[],"teams":[],"aktiv":False}
_manuell_tipps  = []

def whitelist_laden():
    import json, os as _os
    global _whitelist
    if not _os.path.exists(WHITELIST_DATEI):
        return
    try:
        with open(WHITELIST_DATEI) as f:
            _whitelist = json.load(f)
        if _whitelist.get("aktiv"):
            print(f"  [Whitelist] Aktiv: {len(_whitelist.get('ligen',[]))} Ligen")
    except Exception as e:
        print(f"  [Whitelist] Fehler: {e}")

def whitelist_speichern():
    import json
    try:
        with open(WHITELIST_DATEI,"w") as f:
            json.dump(_whitelist,f,indent=2)
    except Exception as e:
        print(f"  [Whitelist] Speicherfehler: {e}")

def manuell_tipps_laden():
    import json, os as _os
    global _manuell_tipps
    if not _os.path.exists(MANUELL_DATEI):
        return
    try:
        with open(MANUELL_DATEI) as f:
            _manuell_tipps = json.load(f)
    except Exception as e:
        print(f"  [Manuell] Fehler: {e}")

def manuell_tipps_speichern():
    import json
    try:
        with open(MANUELL_DATEI,"w") as f:
            json.dump(_manuell_tipps,f,indent=2)
    except Exception as e:
        print(f"  [Manuell] Fehler: {e}")

def check_rate_limit_warnung():
    calls_heute = _api_monitor.get("heute",0)
    pct         = round(calls_heute/API_DAILY_LIMIT*100,1)
    if pct >= 80:
        msg = (f"⚠️ <b>API Rate-Limit Warnung!</b>\n"
               f"📊 Heute: <b>{calls_heute:,}</b>/{API_DAILY_LIMIT:,} ({pct}%)\n"
               f"{'🔴 KRITISCH!' if pct >= 95 else '⚠️ Hohe Auslastung'}\n🕐 {jetzt()} Uhr")
        send_telegram(msg)

PERSISTENZ_DATEIEN = [
    "statistik.json","signal_tracker.json","beobachtete_spiele.json",
    "notified_sets.json","bankroll.json","dynamische_filter.json",
    "whitelist.json","admins.json","bekannte_user.json","manuell_tipps.json",
    "liga_auswertung.json",
]
GITHUB_DATA_PFAD   = "data/latest"
GITHUB_DATA_BRANCH = os.environ.get("GITHUB_DATA_BRANCH","data-backup")

def _github_push_datei(datei: str, inhalt_b64: str, headers: dict, sha: str = None):
    pfad    = f"{GITHUB_DATA_PFAD}/{datei}"
    api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{pfad}"
    if sha is None:
        r   = requests.get(api_url,headers=headers,params={"ref":GITHUB_DATA_BRANCH},timeout=10)
        sha = r.json().get("sha") if r.status_code == 200 else None
    payload = {"message":f"backup {de_now().strftime('%H:%M')}","content":inhalt_b64,"branch":GITHUB_DATA_BRANCH}
    if sha:
        payload["sha"] = sha
    return requests.put(api_url,headers=headers,json=payload,timeout=15)

def github_backup():
    import base64, os as _os
    if not GITHUB_TOKEN or GITHUB_TOKEN.startswith("GITHUB"):
        print("  [Backup] Kein GitHub Token")
        return
    headers = {"Authorization":f"token {GITHUB_TOKEN}","Accept":"application/vnd.github.v3+json"}
    try:
        ref_url = f"https://api.github.com/repos/{GITHUB_REPO}/git/refs/heads/{GITHUB_DATA_BRANCH}"
        if requests.get(ref_url,headers=headers,timeout=8).status_code == 404:
            main = requests.get(f"https://api.github.com/repos/{GITHUB_REPO}/git/refs/heads/main",
                                headers=headers,timeout=8).json()
            sha_main = main.get("object",{}).get("sha","")
            if sha_main:
                requests.post(f"https://api.github.com/repos/{GITHUB_REPO}/git/refs",
                    headers=headers,json={"ref":f"refs/heads/{GITHUB_DATA_BRANCH}","sha":sha_main},timeout=10)
    except Exception as e:
        print(f"  [Backup] Branch-Check Fehler: {e}")
    gesichert = 0
    for datei in PERSISTENZ_DATEIEN:
        if not _os.path.exists(datei):
            continue
        try:
            with open(datei,"rb") as f:
                inhalt = base64.b64encode(f.read()).decode()
            resp = _github_push_datei(datei,inhalt,headers)
            if resp.status_code in (200,201):
                gesichert += 1
        except Exception as e:
            print(f"  [Backup] Fehler bei {datei}: {e}")
    return gesichert

def github_restore():
    import base64, os as _os
    if not GITHUB_TOKEN or GITHUB_TOKEN.startswith("GITHUB"):
        print("  [Restore] Kein GitHub Token – übersprungen")
        return 0
    headers = {"Authorization":f"token {GITHUB_TOKEN}","Accept":"application/vnd.github.v3+json"}
    wiederhergestellt = 0
    for datei in PERSISTENZ_DATEIEN:
        if _os.path.exists(datei):
            continue
        try:
            pfad    = f"{GITHUB_DATA_PFAD}/{datei}"
            api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{pfad}"
            resp    = requests.get(api_url,headers=headers,params={"ref":GITHUB_DATA_BRANCH},timeout=10)
            if resp.status_code != 200:
                continue
            inhalt = base64.b64decode(resp.json().get("content",""))
            with open(datei,"wb") as f:
                f.write(inhalt)
            wiederhergestellt += 1
            print(f"  [Restore] ✅ {datei}")
        except Exception as e:
            print(f"  [Restore] Fehler bei {datei}: {e}")
    if wiederhergestellt > 0:
        print(f"  [Restore] {wiederhergestellt} Dateien wiederhergestellt ✅")
    return wiederhergestellt

def bot_github_backup():
    print(f"[Backup-Bot] Gestartet | Stündlich + täglich {GITHUB_BACKUP_UHRZEIT}:00 Uhr")
    backup_gesendet    = None
    letzter_std_backup = 0
    while True:
        try:
            now    = de_now()
            now_ts = time.time()
            if now_ts-letzter_std_backup >= 3600:
                letzter_std_backup = now_ts
                github_backup()
            if now.hour == GITHUB_BACKUP_UHRZEIT and backup_gesendet != now.date():
                github_backup()
                backup_gesendet = now.date()
                send_telegram(f"💾 <b>Tages-Backup abgeschlossen</b>\n✅ GitHub\n🛡️ Daten sicher\n🕐 {jetzt()} Uhr")
        except Exception as e:
            print(f"  [Backup-Bot] Fehler: {e}")
        time.sleep(60)

def bot_wetter_tipp():
    print("[Wetter-Bot] Gestartet | Wetterbasierte Tipps")
    gesendet_wetter = set()
    while True:
        try:
            now   = de_now()
            datum = now.strftime("%Y-%m-%d")
            fixtures = ls_get_fixtures(datum)
            top = filtere_top_spiele(fixtures)
            laender = {}
            for f in top:
                land = (f.get("country") or {}).get("name","")
                liga = f.get("competition",{}).get("name","?")
                if land and land not in laender:
                    laender[land] = liga
            for land,liga in laender.items():
                key = f"{datum}_{land}"
                if key in gesendet_wetter:
                    continue
                analyse = wetter_analyse(land)
                if not analyse["schlecht"]:
                    continue
                gesendet_wetter.add(key)
                tipps_text = []
                for tipp in analyse["tipps"]:
                    if tipp == "unter_ecken": tipps_text.append("📐 Unter Ecken")
                    elif tipp == "unter_tore": tipps_text.append("⚽ Unter Tore")
                    elif tipp == "mehr_karten": tipps_text.append("🃏 Über Karten")
                if not tipps_text:
                    continue
                info_text = "\n".join(analyse["info"])
                msg = (f"🌦️ <b>Wetter-Tipp!</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                       f"🌍 Land: <b>{land}</b>\n🏆 Liga: {liga}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n{info_text}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"🎯 Tipps: {'  '.join(tipps_text)}\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                send_telegram(msg)
                print(f"  [Wetter-Bot] ✅ {land}: {', '.join(tipps_text)}")
            bot_fehler_reset("Wetter-Bot")
        except Exception as e:
            bot_fehler_melden("Wetter-Bot",e)
        time.sleep(30*60)

# ============================================================
#  PREMATCH BOT
# ============================================================

PREMATCH_TOP_LIGEN = {
    "Premier League","Bundesliga","La Liga","Serie A","Ligue 1",
    "Champions League","Europa League","Conference League",
    "Eredivisie","Primeira Liga","Super Lig","Russian Premier League",
    "Belgian Pro League","Scottish Premiership","Championship",
    "Bundesliga 2","Serie B","2. Bundesliga","Ligue 2",
    "DFB-Pokal","FA Cup","Copa del Rey","Coppa Italia",
    "World Cup","European Championship","Nations League",
    "MLS","Brasileirão","Argentine Primera Division",
}

def ls_get_fixtures(date_str: str) -> list:
    # v60: livescore-api.com kennt keinen "timezone"-Parameter (alle Zeiten kommen in
    # UTC zurück) und gibt maximal 30 Fixtures pro Seite zurück – bei einem ganzen
    # Tag mit Spielen aus aller Welt reicht 1 Seite nicht, deshalb wird hier über
    # "page" weitergeblättert, bis next_page leer ist (Sicherheitsdeckel: max. 15
    # Seiten = 450 Fixtures, damit ein Bug in der API nicht in eine Endlosschleife
    # oder ins Tageslimit läuft).
    alle_raw = []
    seite    = 1
    try:
        while seite <= 15:
            params = {"date":date_str,"page":seite}
            resp   = api_get_with_retry(f"{LS_BASE}/fixtures/list.json",params)
            data   = resp.json().get("data") or {}
            raw    = data.get("fixtures",[]) or []
            alle_raw.extend(raw)
            if not data.get("next_page"):
                break
            seite += 1
        matches = [_af_fixture_zu_prematch(fx) for fx in alle_raw]
        print(f"  [PreMatch] {len(matches)} Fixtures für {date_str} geladen ({seite} Seite(n))")
        return matches
    except Exception as e:
        print(f"  [PreMatch] Fixtures Fehler: {e}")
        return []

PREMATCH_NUR_TOP_LIGEN = False  # v60.2: auf Wunsch wieder AUS – jetzt auch kleinere Ligen erlaubt
                                 # (war früher wegen Datenproblemen bei der alten API auf True gesetzt)

def filtere_top_spiele(fixtures: list) -> list:
    if not PREMATCH_NUR_TOP_LIGEN:
        return fixtures
    top = []
    for f in fixtures:
        liga    = f.get("competition",{}).get("name","")
        country = (f.get("country") or {}).get("name","")
        if ist_top_liga(liga,country):
            top.append(f)
    return top

def claude_prematch_analyse(home: str, away: str, liga: str, anstoß: str, bereits_tipps=None) -> dict:
    bereits = ", ".join(bereits_tipps) if bereits_tipps else "keine"
    if not ANTHROPIC_API_KEY or not ANTHROPIC_API_KEY.strip() or not claude_budget_verfuegbar(liga):
        import random
        fallbacks = [
            {"tipp":"Über 2.5 Tore","analyse":"Beide Teams sind offensivstark.","konfidenz":6},
            {"tipp":"Heimsieg","analyse":"Der Heimvorteil spielt eine entscheidende Rolle.","konfidenz":6},
            {"tipp":"Unter 2.5 Tore","analyse":"Beide Defensivreihen sind stabil.","konfidenz":5},
            {"tipp":"Beide Teams treffen","analyse":"Beide Teams kommen zu Torabschlüssen.","konfidenz":6},
        ]
        verfuegbar = [f for f in fallbacks if f["tipp"] not in (bereits_tipps or [])]
        return random.choice(verfuegbar) if verfuegbar else random.choice(fallbacks)
    try:
        prompt = (f"Sportwetten-Analyst. Analysiere auf Deutsch:\n\n"
                  f"Spiel: {home} vs {away}\nLiga: {liga}\nAnstoß: {anstoß} Uhr\n\n"
                  f"Bereits gewählt (NICHT wiederholen): {bereits}\n\n"
                  f"Wähle den wahrscheinlichsten Tipp:\n"
                  f"- Über/Unter 2.5 Tore\n- Über/Unter 1.5 Tore\n- Beide Teams treffen\n"
                  f"- Heimsieg/Auswärtssieg\n- Doppelte Chance 1X/X2\n\n"
                  f"Antworte NUR:\nTIPP: [Typ]\nKONFIDENZ: [1-10]\nANALYSE: [2-3 Sätze]")
        resp = requests.post("https://api.anthropic.com/v1/messages",
            headers={"Content-Type":"application/json","x-api-key":ANTHROPIC_API_KEY,"anthropic-version":"2023-06-01"},
            json={"model":"claude-sonnet-4-20250514","max_tokens":200,"messages":[{"role":"user","content":prompt}]},
            timeout=20)
        if resp.status_code != 200:
            return None
        text         = resp.json().get("content",[{}])[0].get("text","").strip()
        tipp         = ""
        analyse_text = ""
        konfidenz    = 6
        for line in text.split("\n"):
            if line.startswith("TIPP:"):
                tipp = line.replace("TIPP:","").strip()
            elif line.startswith("ANALYSE:"):
                analyse_text = line.replace("ANALYSE:","").strip()
            elif line.startswith("KONFIDENZ:"):
                try: konfidenz = int(line.replace("KONFIDENZ:","").strip())
                except: konfidenz = 6
        if not tipp:
            return None
        return {"tipp":tipp,"analyse":analyse_text,"konfidenz":konfidenz}
    except Exception as e:
        print(f"  [PreMatch] Claude Fehler: {e}")
        return None

def bot_prematch():
    import random
    print(f"[PreMatch-Bot] Gestartet | Posts um {PREMATCH_UHRZEITEN} Uhr")
    gesendet = set()
    while True:
        try:
            now   = de_now()
            key   = f"{now.strftime('%Y-%m-%d')}_{now.hour}"
            datum = now.strftime("%Y-%m-%d")
            if now.hour in PREMATCH_UHRZEITEN and now.minute < 5 and key not in gesendet:
                print(f"  [PreMatch-Bot] Starte Post um {now.hour}:00 Uhr")
                fixtures = ls_get_fixtures(datum)
                top      = filtere_top_spiele(fixtures)
                def anstoß_stunde(spiel):
                    try: return int(spiel.get("time","0:0").split(":")[0])
                    except: return 0
                if now.hour == 10:   top = [s for s in top if anstoß_stunde(s) < 16]
                elif now.hour == 16: top = [s for s in top if 16 <= anstoß_stunde(s) < 20]
                elif now.hour == 20: top = [s for s in top if anstoß_stunde(s) >= 20]
                if not top:
                    gesendet.add(key)
                    time.sleep(60)
                    continue
                auswahl  = random.sample(top,min(PREMATCH_MAX_TIPPS,len(top)))
                analysen = []
                bereits_tipps = []
                for spiel in auswahl:
                    home   = (spiel.get("home_name") or spiel.get("home",{}).get("name","?"))
                    away   = (spiel.get("away_name") or spiel.get("away",{}).get("name","?"))
                    liga   = spiel.get("competition",{}).get("name","?")
                    country = (spiel.get("country") or {}).get("name","")
                    anstoß = spiel.get("time","?")
                    result = claude_prematch_analyse(home,away,liga,anstoß,bereits_tipps)
                    if not result:
                        continue
                    bereits_tipps.append(result["tipp"])
                    analysen.append({
                        "home":home,"away":away,"liga":liga,"country":country,
                        "anstoß":anstoß,"tipp":result["tipp"],
                        "analyse":result["analyse"],"konfidenz":result.get("konfidenz",6),
                        "fixture_id":spiel.get("id",""),"datum":datum,
                    })
                    time.sleep(1)
                if analysen:
                    uhr_emoji = "🌅" if now.hour == 10 else ("🌆" if now.hour == 16 else "🌙")
                    msg = (f"{uhr_emoji} <b>Pre-Match Tipps – {now.strftime('%d.%m.%Y')}</b>\n"
                           f"━━━━━━━━━━━━━━━━━━━━\n"
                           f"🤖 KI-Analyse powered by BetlabLIVE\n"
                           f"━━━━━━━━━━━━━━━━━━━━\n\n")
                    for i,a in enumerate(analysen,1):
                        liga_str = f"{a['liga']}"+(f" ({a['country']})" if a['country'] else "")
                        ke       = konfidenz_emoji(a['konfidenz'])
                        msg += (f"🏆 <b>{liga_str}</b>\n"
                                f"⚽ <b>{a['home']} vs {a['away']}</b>\n"
                                f"🕐 Anstoß: <b>{a['anstoß']} Uhr</b>\n"
                                f"🎯 Tipp: <b>{a['tipp']}</b>\n"
                                f"{ke} Konfidenz: <b>{a['konfidenz']}/10</b>\n"
                                f"📊 {a['analyse']}\n")
                        if i < len(analysen):
                            msg += "\n━━━━━━━━━━━━━━━━━━━━\n\n"
                    msg += (f"\n━━━━━━━━━━━━━━━━━━━━\n"
                            f"💬 Community: discord.gg/G6dt3Kpf\n"
                            f"⚠️ 18+ | Verantwortungsvoll spielen")
                    send_telegram_gruppe(msg)
                    print(f"  [PreMatch-Bot] ✅ {len(analysen)} Tipps um {now.hour}:00 Uhr")
                    for a in analysen:
                        prematch_tipp_registrieren(a)  # v60.2 NEU: für automatische Auswertung
                gesendet.add(key)
                heute_str = now.strftime('%Y-%m-%d')
                gesendet  = {k for k in gesendet if k.startswith(heute_str)}
        except Exception as e:
            print(f"  [PreMatch-Bot] Fehler: {e}")
        time.sleep(60)

VIP_MIN_KONFIDENZ = 8

# ============================================================
#  v60.2 NEU: AUTOMATISCHE AUSWERTUNG DER PREMATCH-TIPPS
# ============================================================
# Bisher wurden Pre-Match-Tipps nur gepostet, aber nie automatisch nachgeprüft
# (anders als die Live-Signale, die schon einen eigenen Nachschau-Bot haben).
# Läuft parallel dazu, komplett unabhängig vom Live-Signal-Tracker.
PREMATCH_TRACKER_DATEI = "prematch_tracker.json"
_prematch_tracker      = {}
_prematch_tracker_lock = threading.Lock()

def prematch_tracker_laden():
    import json, os as _os
    global _prematch_tracker
    if not _os.path.exists(PREMATCH_TRACKER_DATEI):
        return
    try:
        with open(PREMATCH_TRACKER_DATEI,"r") as f:
            data = json.load(f)
        with _prematch_tracker_lock:
            _prematch_tracker = data
        offen = sum(1 for s in data.values() if s.get("status") == "offen")
        print(f"  [PreMatch-Tracker] {len(data)} Tipps geladen, {offen} noch offen")
    except Exception as e:
        print(f"  [PreMatch-Tracker] Ladefehler: {e}")

def prematch_tracker_speichern():
    import json
    try:
        with _prematch_tracker_lock:
            data = dict(_prematch_tracker)
        with open(PREMATCH_TRACKER_DATEI,"w") as f:
            json.dump(data,f,indent=2)
    except Exception as e:
        print(f"  [PreMatch-Tracker] Speicherfehler: {e}")

def prematch_tipp_registrieren(a: dict):
    fixture_id = str(a.get("fixture_id") or "")
    if not fixture_id:
        return  # ohne Fixture-ID kann später kein Ergebnis mehr zugeordnet werden
    with _prematch_tracker_lock:
        _prematch_tracker[fixture_id] = {
            "home":a.get("home"),"away":a.get("away"),"liga":a.get("liga"),
            "tipp":a.get("tipp"),"konfidenz":a.get("konfidenz"),
            "anstoß":a.get("anstoß"),"datum":a.get("datum"),
            "registriert":de_now().isoformat(),"status":"offen",
        }
    prematch_tracker_speichern()

def pruefe_prematch_tipp(tipp: str, score: str):
    """Wertet einen der festen Tipp-Typen aus claude_prematch_analyse() gegen den
    Endstand aus. Gibt True/False zurück, oder None wenn der Tipp-Text nicht zu
    einem der bekannten Muster passt (z.B. weil Claude doch mal frei formuliert hat)."""
    h,a = parse_score(score)
    t   = (tipp or "").lower()
    gesamt = h+a
    linie_match = re.search(r"(\d+(?:[.,]\d+)?)", t)
    linie = float(linie_match.group(1).replace(",",".")) if linie_match else None
    if "beide teams treffen" in t:
        return h > 0 and a > 0
    if "über" in t and linie is not None:
        return gesamt > linie
    if "unter" in t and linie is not None:
        return gesamt < linie
    if "heimsieg" in t:
        return h > a
    if "auswärtssieg" in t or "auswaertssieg" in t:
        return h < a
    kompakt = t.replace(" ","")
    if "1x" in kompakt:
        return h >= a
    if "x2" in kompakt:
        return a >= h
    return None

def bot_prematch_auswertung():
    """v60.2 NEU: Prüft alle 10 Minuten die offenen Pre-Match-Tipps. Ein Tipp wird
    frühestens 110 Minuten nach Anstoß geprüft (Spiel + Halbzeitpause sollten dann
    fast immer vorbei sein) und spätestens 5 Stunden nach Anstoß aufgegeben (dann ist
    das Spiel aus dem Live-Feed von livescore-api.com rausgefallen und nicht mehr
    zuverlässig auffindbar – siehe ls_get_match_by_fixture_id())."""
    print("[PreMatch-Auswertung-Bot] Gestartet")
    while True:
        try:
            now = de_now()
            with _prematch_tracker_lock:
                offene = {k:v for k,v in _prematch_tracker.items() if v.get("status") == "offen"}
            for fixture_id, tipp_info in offene.items():
                try:
                    datum_str = tipp_info.get("datum","")
                    anstoss_str = tipp_info.get("anstoß","")
                    h,m = map(int, anstoss_str.split(":")[:2])
                    kickoff = datetime.strptime(datum_str,"%Y-%m-%d").replace(
                        hour=h,minute=m,second=0,microsecond=0,tzinfo=now.tzinfo)
                except Exception:
                    continue
                minuten_seit_anstoss = (now-kickoff).total_seconds()/60
                if minuten_seit_anstoss < 110:
                    continue  # Spiel läuft vermutlich noch
                match = ls_get_match_by_fixture_id(fixture_id)
                fertig = match and match.get("status") == "FT"
                if not fertig:
                    if minuten_seit_anstoss > 300:
                        with _prematch_tracker_lock:
                            _prematch_tracker[fixture_id]["status"] = "nicht_auswertbar"
                        prematch_tracker_speichern()
                        print(f"  [PreMatch-Auswertung] {tipp_info.get('home')} vs {tipp_info.get('away')}: "
                              f"nicht mehr auffindbar, aufgegeben")
                    continue
                score = (match.get("scores") or {}).get("score","")
                treffer = pruefe_prematch_tipp(tipp_info.get("tipp",""), score)
                home = tipp_info.get("home","?"); away = tipp_info.get("away","?")
                if treffer is None:
                    with _prematch_tracker_lock:
                        _prematch_tracker[fixture_id]["status"] = "nicht_auswertbar"
                    prematch_tracker_speichern()
                    continue
                emoji = "✅" if treffer else "❌"
                ergebnis_wort = "GEWONNEN" if treffer else "VERLOREN"
                msg = (f"{emoji} <b>Pre-Match Auswertung – {ergebnis_wort}</b>\n"
                       f"⚽ {home} vs {away}\n"
                       f"🎯 Tipp: {tipp_info.get('tipp')}\n"
                       f"📌 Endstand: {score}\n🕐 {jetzt()} Uhr")
                send_telegram_gruppe(msg)
                with _prematch_tracker_lock:
                    _prematch_tracker[fixture_id]["status"] = "gewonnen" if treffer else "verloren"
                    _prematch_tracker[fixture_id]["endstand"] = score
                prematch_tracker_speichern()
                print(f"  [PreMatch-Auswertung] {home} vs {away}: {ergebnis_wort} ({score})")
        except Exception as e:
            print(f"  [PreMatch-Auswertung-Bot] Fehler: {e}")
        time.sleep(600)

def bot_prematch_erinnerung():
    print(f"[Erinnerungs-Bot] Gestartet | 15-Min Erinnerungen, nur Tipps ab Konfidenz {VIP_MIN_KONFIDENZ}/10")
    erinnert = set()
    while True:
        try:
            now      = de_now()
            datum    = now.strftime("%Y-%m-%d")
            fixtures = ls_get_fixtures(datum)
            for spiel in fixtures:
                match_id = str(spiel.get("id",""))
                if not match_id or match_id in erinnert:
                    continue
                home    = spiel.get("home_name",spiel.get("home",{}).get("name","?"))
                away    = spiel.get("away_name",spiel.get("away",{}).get("name","?"))
                liga    = spiel.get("competition",{}).get("name","?")
                country = (spiel.get("country") or {}).get("name","")
                anstoß  = spiel.get("time","")
                if not anstoß or ":" not in anstoß:
                    continue
                try:
                    h,m     = map(int,anstoß.split(":"))
                    kickoff = now.replace(hour=h,minute=m,second=0,microsecond=0)
                    diff    = (kickoff-now).total_seconds()/60
                    if 13 <= diff <= 18:
                        if PREMATCH_NUR_TOP_LIGEN and not ist_top_liga(liga,country):
                            continue
                        erinnert.add(match_id)
                        result = claude_prematch_analyse(home,away,liga,anstoß,[])
                        if not result or result.get("konfidenz",0) < VIP_MIN_KONFIDENZ:
                            print(f"  [Erinnerungs-Bot] {home} vs {away} übersprungen (Konfidenz < {VIP_MIN_KONFIDENZ})")
                            continue
                        ke  = konfidenz_emoji(result["konfidenz"])
                        msg = (f"⏰ <b>Anstoß in ~15 Minuten!</b>\n"
                               f"━━━━━━━━━━━━━━━━━━━━\n"
                               f"🏆 {liga}\n⚽ <b>{home} vs {away}</b>\n"
                               f"🕐 Anstoß: <b>{anstoß} Uhr</b>\n"
                               f"━━━━━━━━━━━━━━━━━━━━\n"
                               f"🎯 Tipp: <b>{result['tipp']}</b>\n"
                               f"{ke} Konfidenz: <b>{result['konfidenz']}/10</b>\n"
                               f"📊 {result.get('analyse','')}\n"
                               f"━━━━━━━━━━━━━━━━━━━━\n💬 discord.gg/G6dt3Kpf")
                        send_telegram_gruppe(msg)
                except Exception:
                    continue
        except Exception as e:
            print(f"  [Erinnerungs-Bot] Fehler: {e}")
        time.sleep(60)

def bot_morgen_uebersicht():
    print("[Morgen-Bot] Gestartet | Täglich 08:00 Uhr")
    gesendet_morgen = set()
    while True:
        try:
            now = de_now()
            key = now.strftime("%Y-%m-%d")
            if now.hour == 8 and now.minute < 5 and key not in gesendet_morgen:
                gesendet_morgen.add(key)
                datum    = now.strftime("%Y-%m-%d")
                fixtures = ls_get_fixtures(datum)
                top      = filtere_top_spiele(fixtures)
                if not top:
                    time.sleep(60)
                    continue
                ligen = {}
                for f in top:
                    liga   = f.get("competition",{}).get("name","?")
                    anstoß = f.get("time","?")
                    home   = f.get("home_name") or (f.get("home") or {}).get("name","?")
                    away   = f.get("away_name") or (f.get("away") or {}).get("name","?")
                    if liga not in ligen:
                        ligen[liga] = []
                    ligen[liga].append(f"  {anstoß}: {home} vs {away}")
                liga_text = ""
                for liga,spiele in list(ligen.items())[:6]:
                    liga_text += f"\n🏆 <b>{liga}</b>\n"+"\n".join(spiele[:3])+"\n"
                msg = (f"🌅 <b>Guten Morgen! Tages-Übersicht {now.strftime('%d.%m.%Y')}</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"📅 <b>Heutige Top-Spiele:</b>{liga_text}"
                       f"━━━━━━━━━━━━━━━━━━━━\n"
                       f"🤖 Der Bot ist aktiv und analysiert alle Spiele!")
                send_telegram(msg)
                send_telegram_gruppe(msg)
                print(f"  [Morgen-Bot] ✅ {len(top)} Spiele heute")
        except Exception as e:
            print(f"  [Morgen-Bot] Fehler: {e}")
        time.sleep(60)

def bot_selbstlernend():
    print("[Selbstlern-Bot] Gestartet | Tägliche Filter-Optimierung")
    letzter_check = None
    while True:
        try:
            now = de_now()
            if now.hour == 6 and letzter_check != now.date():
                letzter_check = now.date()
                print("  [Selbstlern] Starte tägliche Analyse...")
                aenderungen = []
                for typ,filter_dict in DYNAMISCHE_FILTER.items():
                    perf = analysiere_bot_performance(typ)
                    if not perf.get("ausreichend"):
                        continue
                    quote = perf["quote"]
                    if quote < 0.40:
                        if typ == "comeback" and filter_dict.get("COMEBACK_AB_MINUTE",30) < 45:
                            DYNAMISCHE_FILTER[typ]["COMEBACK_AB_MINUTE"] = min(45,filter_dict["COMEBACK_AB_MINUTE"]+5)
                            aenderungen.append(f"🔴 Comeback: Minute → {DYNAMISCHE_FILTER[typ]['COMEBACK_AB_MINUTE']}")
                        elif typ == "druck" and filter_dict.get("DRUCK_RATIO",2.5) < 3.5:
                            DYNAMISCHE_FILTER[typ]["DRUCK_RATIO"] = round(filter_dict["DRUCK_RATIO"]+0.25,2)
                            aenderungen.append(f"🔴 Druck: Ratio → {DYNAMISCHE_FILTER[typ]['DRUCK_RATIO']}")
                        elif typ == "torwart" and filter_dict.get("MIN_SHOTS_ON_TARGET",3) < 6:
                            DYNAMISCHE_FILTER[typ]["MIN_SHOTS_ON_TARGET"] = filter_dict["MIN_SHOTS_ON_TARGET"]+1
                            aenderungen.append(f"🔴 Torwart: Min. Schüsse → {DYNAMISCHE_FILTER[typ]['MIN_SHOTS_ON_TARGET']}")
                    elif quote > 0.65:
                        if typ == "torwart" and filter_dict.get("MIN_SHOTS_ON_TARGET",3) > 2:
                            DYNAMISCHE_FILTER[typ]["MIN_SHOTS_ON_TARGET"] = filter_dict["MIN_SHOTS_ON_TARGET"]-1
                            aenderungen.append(f"🟢 Torwart: Min. Schüsse → {DYNAMISCHE_FILTER[typ]['MIN_SHOTS_ON_TARGET']}")
                dynamische_filter_speichern()
                if aenderungen:
                    msg = (f"🧠 <b>Selbstlern-Update!</b>\n━━━━━━━━━━━━━━━━━━━━\n"+
                           "\n".join(aenderungen)+f"\n━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
                    send_telegram(msg)
                # v59.17 NEU: Konfidenz-Kalibrierung täglich mitlaufen lassen (vorher nie
                # aufgerufen). Prüft ob z.B. "8/10"-Signale auch wirklich ~80% treffen.
                try:
                    kalibriere_konfidenz()
                except Exception as e:
                    print(f"  [Selbstlern] Kalibrierung Fehler: {e}")
        except Exception as e:
            print(f"  [Selbstlern] Fehler: {e}")
        time.sleep(60)

def analysiere_bot_performance(typ: str) -> dict:
    with _tracker_lock:
        tipps = [s for s in _signal_tracker.values()
                 if s.get("typ") == typ and s.get("status") == "ausgewertet"]
    if len(tipps) < 8:  # v59.17: von 15 auf 8 gesenkt, damit der Selbstlern-Bot bei
                        # aktuell reduziertem Signal-Volumen schneller reagieren kann
        return {"ausreichend":False,"tipps":len(tipps)}
    gw = sum(1 for t in tipps if t.get("gewonnen"))
    return {"ausreichend":True,"tipps":len(tipps),"gewonnen":gw,"quote":gw/len(tipps)}

def bot_telegram_gruppe():
    print("[Gruppen-Bot] Gestartet | Integriert in Haupt-Bot")
    while True:
        time.sleep(3600)


# ============================================================
#  TELEGRAM BEFEHLE
# ============================================================

ADMIN_IDS  = [TELEGRAM_CHAT_ID]
ADMIN_DATEI = "admins.json"
BEKANNTE_USER = set()
BEKANNTE_DATEI = "bekannte_user.json"

def admins_laden():
    import json, os as _os
    global ADMIN_IDS
    if not _os.path.exists(ADMIN_DATEI):
        return
    try:
        with open(ADMIN_DATEI) as f:
            data = json.load(f)
        ADMIN_IDS = data.get("ids",[TELEGRAM_CHAT_ID])
    except Exception as e:
        print(f"  [Admin] Ladefehler: {e}")

def admins_speichern():
    import json
    try:
        with open(ADMIN_DATEI,"w") as f:
            json.dump({"ids":ADMIN_IDS},f)
    except Exception as e:
        print(f"  [Admin] Speicherfehler: {e}")

def ist_admin(chat_id: str) -> bool:
    return str(chat_id) in [str(a) for a in ADMIN_IDS]

def bekannte_user_laden():
    import json, os as _os
    global BEKANNTE_USER
    if not _os.path.exists(BEKANNTE_DATEI):
        return
    try:
        with open(BEKANNTE_DATEI) as f:
            BEKANNTE_USER = set(json.load(f))
    except Exception:
        pass

def bekannte_user_speichern():
    import json
    try:
        with open(BEKANNTE_DATEI,"w") as f:
            json.dump(list(BEKANNTE_USER),f)
    except Exception:
        pass

def bot_telegram_befehle():
    global BOT_PAUSIERT
    print("[Telegram-Befehle] Gestartet | Lausche auf Befehle")
    letzter_update_id = 0
    while True:
        try:
            url  = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/getUpdates"
            resp = requests.get(url,params={"offset":letzter_update_id+1,"timeout":30},timeout=35)
            if resp.status_code != 200:
                time.sleep(5)
                continue
            updates = resp.json().get("result",[])
            for update in updates:
                letzter_update_id = update["update_id"]
                msg_obj  = update.get("message",{}) or {}
                chat_id  = str(msg_obj.get("chat",{}).get("id",""))
                user_id  = str(msg_obj.get("from",{}).get("id",""))
                username = msg_obj.get("from",{}).get("first_name","Anonym")
                text     = msg_obj.get("text","").strip()
                if not text:
                    continue
                erlaubte_chats = {str(TELEGRAM_CHAT_ID), str(TELEGRAM_CHAT_PREMATCH)}
                erlaubte_chats.update(str(a) for a in ADMIN_IDS)
                if chat_id not in erlaubte_chats:
                    print(f"  [Telegram] Unbekannter Chat {chat_id} von {username} – ignoriert")
                    continue

                def antworten(msg):
                    requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage",
                                  json={"chat_id":chat_id,"text":msg,"parse_mode":"HTML"},timeout=10)

                if text == "/status":
                    aktive = {t.name for t in threading.enumerate()}
                    bots   = ["Ecken-Bot","Torwart-Bot","Druck-Bot","Comeback-Bot",
                              "Torflut-Bot","Tore-Bot","PreMatch-Bot","Nachschau-Bot"]
                    zeilen = "\n".join([f"{'✅' if b in aktive else '❌'} {b}" for b in bots])
                    pause  = "⏸ PAUSIERT" if BOT_PAUSIERT else "▶️ AKTIV"
                    antworten(f"🤖 <b>Bot Status</b> – {pause}\n"
                              f"📦 Build: <b>{BOT_VERSION}</b>\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n{zeilen}\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n{api_monitor_bericht()}\n🕐 {jetzt()} Uhr")

                elif text == "/pause":
                    BOT_PAUSIERT = True
                    antworten("⏸ <b>Alle Signale pausiert.</b>\nMit /start wieder aktivieren.")

                elif text == "/start":
                    BOT_PAUSIERT = False
                    antworten("▶️ <b>Signale wieder aktiv!</b>")

                elif text == "/statistik":
                    gw  = sum(statistik[t]["gewonnen"] for t in statistik)
                    vl  = sum(statistik[t]["verloren"] for t in statistik)
                    ges = gw+vl
                    pct = round(gw/ges*100) if ges else 0
                    gn  = round(sum(statistik[t]["gewinn"] for t in statistik),2)
                    antworten(f"📊 <b>Statistik heute</b>\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n"
                              f"✅ Gewonnen: <b>{gw}</b>\n❌ Verloren: <b>{vl}</b>\n"
                              f"🎯 Trefferquote: <b>{pct}%</b>\n"
                              f"📈 ROI: <b>{'+' if gn>=0 else ''}{gn}€</b>\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")

                elif text == "/bankroll":
                    br   = bankroll_laden()
                    diff = round(br-BANKROLL,2)
                    emoji = "📈" if diff >= 0 else "📉"
                    antworten(f"💰 <b>Bankroll</b>\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n"
                              f"Start: <b>{BANKROLL}€</b>\n"
                              f"Aktuell: <b>{br}€</b>\n"
                              f"{emoji} Differenz: <b>{'+' if diff>=0 else ''}{diff}€</b>\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")

                elif text == "/live":
                    try:
                        matches = get_live_matches()
                        if not matches:
                            antworten("⚽ Gerade keine Live-Spiele.")
                        else:
                            zeilen = [f"⚽ <b>Live Spiele</b> ({len(matches)} gesamt)\n━━━━━━━━━━━━━━━━━━━━"]
                            for m in matches[:15]:
                                home   = m.get("home",{}).get("name","?")
                                away   = m.get("away",{}).get("name","?")
                                score  = m.get("scores",{}).get("score","? - ?")
                                minute = m.get("time","?")
                                status = m.get("status","")
                                min_str = "HZ" if status == "HALF TIME BREAK" else f"{minute}'"
                                zeilen.append(f"🔴 {home} <b>{score}</b> {away} | {min_str}")
                            antworten("\n".join(zeilen))
                    except Exception as e:
                        antworten(f"❌ Fehler: {e}")

                elif text == "/rangliste":
                    rang = bot_rangliste()
                    antworten(f"🏆 <b>Bot-Rangliste (Woche)</b>\n━━━━━━━━━━━━━━━━━━━━\n{rang}\n🕐 {jetzt()} Uhr")

                elif text == "/auswertung":
                    offene     = tracker_get_offene()
                    alle       = list(_signal_tracker.values())
                    ausgewertet = [s for s in alle if s.get("status") == "ausgewertet"]
                    gew        = [s for s in ausgewertet if s.get("gewonnen")]
                    pct = round(len(gew)/len(ausgewertet)*100) if ausgewertet else 0
                    offen_liste = "\n".join([f"  • {s.get('home','?')} vs {s.get('away','?')} ({s.get('typ','')})"
                                            for _,s in offene[:5]]) or "  Keine offenen Signale"
                    antworten(f"📋 <b>Signal-Tracker Status</b>\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n"
                              f"📨 Gesamt: <b>{len(alle)}</b>\n"
                              f"✅ Ausgewertet: <b>{len(ausgewertet)}</b>\n"
                              f"⏳ Offen: <b>{len(offene)}</b>\n"
                              f"🎯 Quote: <b>{pct}%</b>\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n"
                              f"⏳ <b>Offene Signale:</b>\n{offen_liste}\n🕐 {jetzt()} Uhr")

                elif text == "/api":
                    antworten(f"📡 <b>API Monitor</b>\n━━━━━━━━━━━━━━━━━━━━\n{api_monitor_bericht()}\n🕐 {jetzt()} Uhr")

                elif text == "/liga_status":
                    # v59.15: Übersicht der Liga-Zuverlässigkeit für Auswertungen
                    zeilen = []
                    for liga_n, st in sorted(liga_auswertung_stats.items(),
                                              key=lambda x: x[1]["erfolgreich"]+x[1]["nicht_auswertbar"], reverse=True):
                        gesamt = st["erfolgreich"]+st["nicht_auswertbar"]
                        if gesamt < 1:
                            continue
                        quote = round(st["erfolgreich"]/gesamt*100)
                        gesperrt = gesamt >= LIGA_AUSWERTUNG_MIN_PROBEN and (st["erfolgreich"]/gesamt) < LIGA_AUSWERTUNG_MIN_QUOTE
                        icon = "🔴" if gesperrt else ("🟡" if gesamt < LIGA_AUSWERTUNG_MIN_PROBEN else "🟢")
                        zeilen.append(f"{icon} {liga_n}: {st['erfolgreich']}/{gesamt} auswertbar ({quote}%){' – GESPERRT' if gesperrt else ''}")
                    text_out = "\n".join(zeilen[:25]) if zeilen else "Noch keine Daten"
                    antworten(f"📋 <b>Liga-Zuverlässigkeit</b>\n━━━━━━━━━━━━━━━━━━━━\n{text_out}\n"
                              f"━━━━━━━━━━━━━━━━━━━━\n🟢 OK | 🟡 noch zu wenig Daten (&lt;{LIGA_AUSWERTUNG_MIN_PROBEN}) | 🔴 gesperrt (&lt;{int(LIGA_AUSWERTUNG_MIN_QUOTE*100)}%)\n🕐 {jetzt()} Uhr")

                elif text.startswith("/whitelist "):
                    teile = text.split(" ",2)
                    if teile[1] == "on":
                        _whitelist["aktiv"] = True; whitelist_speichern(); antworten("✅ Whitelist aktiviert")
                    elif teile[1] == "off":
                        _whitelist["aktiv"] = False; whitelist_speichern(); antworten("❌ Whitelist deaktiviert")
                    elif teile[1] == "reset":
                        _whitelist["ligen"] = []; _whitelist["teams"] = []; whitelist_speichern(); antworten("🗑️ Whitelist geleert")
                    elif teile[1] == "liga" and len(teile) > 2:
                        _whitelist.setdefault("ligen",[]).append(teile[2]); whitelist_speichern(); antworten(f"✅ Liga hinzugefügt: {teile[2]}")
                    elif teile[1] == "team" and len(teile) > 2:
                        _whitelist.setdefault("teams",[]).append(teile[2]); whitelist_speichern(); antworten(f"✅ Team hinzugefügt: {teile[2]}")

                elif text == "/filter":
                    an  = [f"✅ {n}  →  /filter_off {k}" for k,n in TELEGRAM_BOT_NAMEN.items() if k not in _telegram_deaktiviert]
                    aus = [f"❌ {n}  →  /filter_on {k}"  for k,n in TELEGRAM_BOT_NAMEN.items() if k in _telegram_deaktiviert]
                    antworten("📱 <b>Signal-Filter</b>\n━━━━━━━━━━━━━━━━━━━━\n"
                              "🟢 <b>Aktiv:</b>\n"+"\n".join(an)+
                              (("\n\n🔴 <b>Deaktiviert:</b>\n"+"\n".join(aus)) if aus else ""))

                elif text.startswith("/filter_off "):
                    key = text.replace("/filter_off ","").strip().lower()
                    if key == "alle":
                        _telegram_deaktiviert.update(TELEGRAM_BOT_NAMEN.keys()); telegram_filter_speichern()
                        antworten("❌ Alle Telegram-Signale deaktiviert")
                    elif key in TELEGRAM_BOT_NAMEN:
                        _telegram_deaktiviert.add(key); telegram_filter_speichern()
                        antworten(f"❌ {TELEGRAM_BOT_NAMEN[key]} deaktiviert")

                elif text.startswith("/filter_on "):
                    key = text.replace("/filter_on ","").strip().lower()
                    if key == "alle":
                        _telegram_deaktiviert.clear(); telegram_filter_speichern()
                        antworten("✅ Alle Telegram-Signale aktiviert")
                    elif key in TELEGRAM_BOT_NAMEN:
                        _telegram_deaktiviert.discard(key); telegram_filter_speichern()
                        antworten(f"✅ {TELEGRAM_BOT_NAMEN[key]} aktiviert")

                elif text.startswith("/addadmin ") and ist_admin(chat_id):
                    neuer_id = text.replace("/addadmin ","").strip()
                    if neuer_id not in [str(a) for a in ADMIN_IDS]:
                        ADMIN_IDS.append(neuer_id); admins_speichern()
                        antworten(f"✅ Admin hinzugefügt: {neuer_id}")
                    else:
                        antworten("⚠️ Bereits Admin")

                elif text.startswith("/tipp "):
                    teile_t = text.strip().split(" ",3)
                    if len(teile_t) >= 3:
                        spiel_t = teile_t[1]; bet_t = teile_t[2]
                        quote_t = teile_t[3] if len(teile_t) > 3 else None
                        tipp_obj = {"typ":"manuell","home":spiel_t,"away":"","tipp":bet_t,
                                    "quote":float(quote_t) if quote_t else None,"status":"offen","signal_zeit":time.time()}
                        _manuell_tipps.append(tipp_obj); manuell_tipps_speichern()
                        antworten(f"✅ <b>Manueller Tipp gespeichert!</b>\n"
                                  f"⚽ Spiel: <b>{spiel_t}</b>\n🎯 Tipp: <b>{bet_t}</b>\n"
                                  f"💶 Quote: <b>{quote_t or 'keine'}</b>")
                    else:
                        antworten("Benutzung: /tipp [Spiel] [Bet] [Quote]\nBeispiel: /tipp ManCity Über2.5 1.85")

                elif text in ("/gewonnen","/verloren"):
                    if _manuell_tipps:
                        letzter = _manuell_tipps[-1]
                        letzter["status"]   = "ausgewertet"
                        letzter["gewonnen"] = text == "/gewonnen"
                        manuell_tipps_speichern()
                        antworten(f"{'✅ GEWONNEN' if letzter['gewonnen'] else '❌ VERLOREN'} – {letzter['tipp']}")
                    else:
                        antworten("Kein offener Tipp gefunden")

        except Exception as e:
            print(f"  [Telegram-Befehle] Fehler: {e}")
        time.sleep(2)


# ============================================================
#  WEB-DASHBOARD & HEALTH CHECK
# ============================================================

BOT_START_ZEIT = time.time()

def bot_web_dashboard():
    try:
        from http.server import HTTPServer, BaseHTTPRequestHandler
        import json

        class DashboardHandler(BaseHTTPRequestHandler):
            def log_message(self,format,*args): pass
            def do_GET(self):
                if self.path in ("/health", "/health.json"):
                    uptime_min = round((time.time()-BOT_START_ZEIT)/60)
                    gw = sum(statistik[t]["gewonnen"] for t in statistik)
                    vl = sum(statistik[t]["verloren"] for t in statistik)
                    data = {"status":"ok","uptime_min":uptime_min,"gewonnen":gw,"verloren":vl,
                            "offene_signale":len(tracker_get_offene()),
                            "api_calls_heute":_api_monitor.get("heute",0),
                            "api_limit":API_DAILY_LIMIT}
                    self.send_response(200)
                    self.send_header("Content-Type","application/json")
                    self.send_header("Access-Control-Allow-Origin","*")
                    self.end_headers()
                    self.wfile.write(json.dumps(data).encode())
                elif self.path == "/api/stats":
                    gw  = sum(statistik[t]["gewonnen"] for t in statistik)
                    vl  = sum(statistik[t]["verloren"] for t in statistik)
                    ges = gw+vl
                    pct = round(gw/ges*100) if ges else 0
                    br  = bankroll_laden()
                    data = {
                        "gewonnen":gw,"verloren":vl,"trefferquote":pct,
                        "bankroll":br,"offene_signale":len(tracker_get_offene()),
                        "streak":streak_aktuell,"streak_beste":streak_beste,
                        "nach_typ":{t:statistik[t] for t in statistik},
                        "api_calls":_api_monitor.get("heute",0),
                        "api_limit":API_DAILY_LIMIT,
                        "zeit":jetzt(),
                    }
                    self.send_response(200)
                    self.send_header("Content-Type","application/json")
                    self.send_header("Access-Control-Allow-Origin","*")
                    self.end_headers()
                    self.wfile.write(json.dumps(data).encode())
                elif self.path in ("/","/index.html","/radar","/health"):
                    html = f"""<!DOCTYPE html>
<html lang="de"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>BetlabLIVE v59</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{background:#0d1117;color:#e6edf3;font-family:system-ui;padding:20px}}
h1{{color:#58a6ff;margin-bottom:20px}}
.grid{{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:16px;margin-bottom:24px}}
.card{{background:#161b22;border:1px solid #30363d;border-radius:12px;padding:20px;text-align:center}}
.card .val{{font-size:36px;font-weight:700;margin:8px 0}}
.card .lbl{{font-size:13px;color:#8b949e}}
.green{{color:#3fb950}}.red{{color:#f85149}}.blue{{color:#58a6ff}}.yellow{{color:#d29922}}
.badge{{padding:2px 8px;border-radius:20px;font-size:12px}}
.badge-g{{background:#1f3a2a;color:#3fb950}}.badge-r{{background:#3a1f1f;color:#f85149}}
table{{width:100%;border-collapse:collapse;background:#161b22;border-radius:12px;overflow:hidden}}
th,td{{padding:12px 16px;text-align:left;border-bottom:1px solid #21262d}}
th{{background:#21262d;font-size:12px;color:#8b949e;text-transform:uppercase}}
.footer{{margin-top:20px;color:#8b949e;font-size:12px}}
.v57-badge{{background:rgba(88,166,255,0.1);border:1px solid #58a6ff;padding:4px 12px;border-radius:20px;color:#58a6ff;font-size:12px;margin-bottom:16px;display:inline-block}}
</style></head><body>
<h1>⚽ BetlabLIVE Dashboard</h1>
<div class="v57-badge">v59 – Reaktionen + Duplikat-/Tagesbericht-Fix ✅</div>
<div class="grid" id="cards"></div>
<table><thead><tr><th>Bot</th><th>Gewonnen</th><th>Verloren</th><th>Quote</th></tr></thead>
<tbody id="bots"></tbody></table>
<div class="footer" id="footer"></div>
<script>
async function update(){{
  const r=await fetch('/api/stats');const d=await r.json();
  const pct=d.trefferquote;
  const apiPct=Math.round(d.api_calls/d.api_limit*100);
  document.getElementById('cards').innerHTML=`
    <div class="card"><div class="val green">${{d.gewonnen}}</div><div class="lbl">✅ Gewonnen</div></div>
    <div class="card"><div class="val red">${{d.verloren}}</div><div class="lbl">❌ Verloren</div></div>
    <div class="card"><div class="val ${{pct>=55?'green':pct>=45?'yellow':'red'}}">${{pct}}%</div><div class="lbl">🎯 Trefferquote</div></div>
    <div class="card"><div class="val blue">${{d.bankroll}}€</div><div class="lbl">💰 Bankroll</div></div>
    <div class="card"><div class="val yellow">${{d.offene_signale}}</div><div class="lbl">⏳ Offen</div></div>
    <div class="card"><div class="val ${{apiPct>=90?'red':apiPct>=75?'yellow':'green'}}">${{apiPct}}%</div><div class="lbl">📡 API-Limit</div></div>
  `;
  const namen={{ecken:'📐 Ecken U',torwart:'🧤 Torwart',druck:'🔥 Druck',comeback:'🔄 Comeback',
    torflut:'🌊 Torflut',hz1tore:'🥅 HZ1-Tore',vztore:'🏆 VZ-Tore'}};
  let rows='';
  for(const[k,v]of Object.entries(d.nach_typ)){{
    const ges=v.gewonnen+v.verloren;if(ges===0)continue;
    const q=Math.round(v.gewonnen/ges*100);
    const badge=q>=55?`<span class="badge badge-g">${{q}}%</span>`:`<span class="badge badge-r">${{q}}%</span>`;
    rows+=`<tr><td>${{namen[k]||k}}</td><td class="green">${{v.gewonnen}}</td><td class="red">${{v.verloren}}</td><td>${{badge}}</td></tr>`;
  }}
  document.getElementById('bots').innerHTML=rows||'<tr><td colspan="4" style="text-align:center;color:#8b949e;padding:20px">Noch keine Daten</td></tr>';
  document.getElementById('footer').textContent=`Letzte Aktualisierung: ${{d.zeit}} Uhr | API: ${{d.api_calls.toLocaleString()}}/${{d.api_limit.toLocaleString()}} Calls`;
}}
update();setInterval(update,30000);
</script></body></html>"""
                    self.send_response(200)
                    self.send_header("Content-Type","text/html; charset=utf-8")
                    self.end_headers()
                    self.wfile.write(html.encode())
                else:
                    self.send_response(404)
                    self.end_headers()

        server = HTTPServer(("0.0.0.0",8080),DashboardHandler)
        print("[Dashboard] Gestartet auf Port 8080")
        server.serve_forever()
    except Exception as e:
        print(f"[Dashboard] Fehler: {e}")

def bot_health_check_server():
    print("[Health-Check] Health-Check auf Port 8080 /health integriert (kein separater Server)")
    while True:
        time.sleep(3600)

# ============================================================
#  AB TESTING
# ============================================================

AB_FILTER = {
    "ecken_unter": {
        "A": {"max_hz1_corners": 5, "label": "≤5 HZ1"},
        "B": {"max_hz1_corners": 4, "label": "≤4 HZ1"},
        "ergebnisse": {"A": {"g":0,"total":0}, "B": {"g":0,"total":0}},
    }
}
AB_DATEI = "ab_test.json"

def ab_test_laden():
    import json, os as _os
    global AB_FILTER
    if _os.path.exists(AB_DATEI):
        try:
            with open(AB_DATEI) as f: AB_FILTER = json.load(f)
        except Exception: pass

def ab_test_speichern():
    import json
    try:
        with open(AB_DATEI,"w") as f: json.dump(AB_FILTER,f,indent=2)
    except Exception: pass

def ab_test_variante(test_name):
    if test_name not in AB_FILTER: return "A"
    erg = AB_FILTER[test_name]["ergebnisse"]
    return "A" if erg["A"]["total"] <= erg["B"]["total"] else "B"

def ab_test_auswerten():
    for test,data in AB_FILTER.items():
        erg = data["ergebnisse"]
        if erg["A"]["total"]<20 or erg["B"]["total"]<20: continue
        q_a = erg["A"]["g"]/erg["A"]["total"]
        q_b = erg["B"]["g"]/erg["B"]["total"]
        w   = "A" if q_a>=q_b else "B"
        send_telegram(f"🧪 <b>A/B Test: {test}</b>\nA: {round(q_a*100)}% | B: {round(q_b*100)}%\n🏆 Gewinner: <b>{w}</b>")

# ============================================================
#  VIRTUELLE KONTEN
# ============================================================

VIRTUELLE_KONTEN = {
    "safe":     {"name":"Shield Safe",   "bots":["ecken","hz1tore","torwart"],   "bankroll":100.0,"start":100.0},
    "highrise": {"name":"Flash High",    "bots":["comeback","druck","rotkarte"], "bankroll":50.0, "start":50.0},
    "value":    {"name":"Diamond Value", "bots":["value","arbitrage","sharp"],   "bankroll":75.0, "start":75.0},
}
VK_DATEI = "virtuelle_konten.json"

def vk_laden():
    import json, os as _os
    global VIRTUELLE_KONTEN
    if _os.path.exists(VK_DATEI):
        try:
            with open(VK_DATEI) as f: VIRTUELLE_KONTEN = json.load(f)
        except Exception: pass

def vk_speichern():
    import json
    try:
        with open(VK_DATEI,"w") as f: json.dump(VIRTUELLE_KONTEN,f,indent=2)
    except Exception: pass

def vk_update(typ, gewonnen, quote=1.85):
    konto = next((k for k,v in VIRTUELLE_KONTEN.items() if typ in v.get("bots",[])), "safe")
    br    = VIRTUELLE_KONTEN[konto]["bankroll"]
    eins  = min(br*0.05, 10)
    VIRTUELLE_KONTEN[konto]["bankroll"] = round(br+eins*(quote-1) if gewonnen else max(1,br-eins),2)
    vk_speichern()

def vk_status_text():
    """v59.2: Nutzt Discord-Markdown (**fett**) statt HTML <b>-Tags, da diese Funktion
    ausschließlich im Discord-Tagesbericht-Embed verwendet wird (Discord rendert kein HTML)."""
    zeilen = []
    for k,v in VIRTUELLE_KONTEN.items():
        diff = round(v["bankroll"]-v["start"],2)
        zeilen.append(f"{v['name']}: **{v['bankroll']}€** ({'+' if diff>=0 else ''}{diff}€)")
    return "\n".join(zeilen)


# ============================================================
#  RANG & XP SYSTEM
# ============================================================

XP_QUELLEN = {"tipp_gewonnen":50,"tipp_verloren":10,"daily_checkin":20,
               "einladung":200,"streak_5":50,"streak_10":150,"woche_gewonnen":100}
LEVEL_STUFEN = [
    (0,1,"🆕 Newcomer"),(100,2,"🎯 Einsteiger"),(300,3,"📊 Tipper"),
    (600,4,"🔍 Analyst"),(1000,5,"🥉 Bronze Tipster"),(1500,6,"📈 Fortgeschrittener"),
    (2200,7,"🎲 Tipp-Profi"),(3000,8,"🥈 Silber Tipster"),(4000,9,"🔥 Heißer Draht"),
    (5500,10,"🥇 Gold Tipster"),(7500,11,"💡 Stratege"),(10000,12,"⚡ Streak Hunter"),
    (13000,13,"🎯 Sniper"),(17000,14,"💎 Diamond Tipster"),(22000,15,"🌟 Superstar"),
]
RANG_DATEI   = "rang_system.json"
_rang_daten  = {}
_checkin_heute = {}

def rang_laden():
    import json, os as _os
    global _rang_daten
    if _os.path.exists(RANG_DATEI):
        try:
            with open(RANG_DATEI) as f: _rang_daten = json.load(f)
            print(f"  [Rang] {len(_rang_daten)} User geladen")
        except Exception as e: print(f"  [Rang] Fehler: {e}")

def rang_speichern():
    import json
    try:
        with open(RANG_DATEI,"w") as f: json.dump(_rang_daten,f,indent=2)
    except Exception: pass

def berechne_level(xp):
    akt = LEVEL_STUFEN[0]
    for min_xp,level,name in LEVEL_STUFEN:
        if xp>=min_xp: akt = (min_xp,level,name)
    idx = next((i for i,s in enumerate(LEVEL_STUFEN) if s[1]==akt[1]),0)
    bis = LEVEL_STUFEN[idx+1][0]-xp if idx<len(LEVEL_STUFEN)-1 else 0
    return akt[1],akt[2],bis

def gib_xp(user_id, username, xp, grund):
    if user_id not in _rang_daten:
        _rang_daten[user_id] = {"name":username,"xp":0,"level":1,"level_name":"🆕 Newcomer",
                                 "gewinne":0,"verluste":0,"streak":0}
    d = _rang_daten[user_id]; d["name"] = username; altes = d["level"]
    d["xp"] += xp
    nl,nname,_ = berechne_level(d["xp"]); d["level"]=nl; d["level_name"]=nname
    rang_speichern()
    if nl > altes:
        send_telegram_gruppe(f"🎉 <b>LEVEL UP! {username}</b>\n{altes}→<b>{nl}</b> {nname}")
    return d

def mache_daily_checkin(user_id, username):
    heute_str = de_now().strftime("%Y-%m-%d")
    if _checkin_heute.get(user_id)==heute_str: return f"⏰ Heute bereits eingecheckt!"
    _checkin_heute[user_id] = heute_str
    d = gib_xp(user_id,username,XP_QUELLEN["daily_checkin"],"daily_checkin")
    return f"✅ +{XP_QUELLEN['daily_checkin']} XP | Gesamt: {d['xp']:,} | {d['level_name']}"

def xp_rangliste():
    sort = sorted(_rang_daten.items(),key=lambda x:x[1].get("xp",0),reverse=True)[:10]
    medals = ["🥇","🥈","🥉"]
    return "\n".join([f"{medals[i] if i<3 else str(i+1)+'.'} <b>{d.get('name','?')}</b> Lv.{d.get('level',1)} | {d.get('xp',0):,} XP"
                      for i,(uid,d) in enumerate(sort)]) or "Noch keine Einträge"

def wöchentliche_xp_auswertung(): pass

# ============================================================
#  COMMUNITY SYSTEM
# ============================================================

COMMUNITY_DATEI    = "community_system.json"
_community_system  = {"einladungen":{},"challenges":{},"rollen":{}}

def community_system_laden():
    import json, os as _os
    global _community_system
    if not _os.path.exists(COMMUNITY_DATEI): return
    try:
        with open(COMMUNITY_DATEI) as f: _community_system.update(json.load(f))
        print("  [Community] System geladen")
    except Exception as e: print(f"  [Community] Fehler: {e}")

def community_system_speichern():
    import json
    try:
        with open(COMMUNITY_DATEI,"w") as f: json.dump(_community_system,f,indent=2)
    except Exception: pass

# ============================================================
#  DISCORD VOTE SYSTEM
# ============================================================

_discord_votes  = {}
_discord_punkte = {}
DISCORD_VOTE_DATEI = "discord_votes.json"

def discord_votes_laden():
    import json, os as _os
    global _discord_votes, _discord_punkte
    for datei,ziel in [(DISCORD_VOTE_DATEI,"votes"),("discord_punkte.json","punkte")]:
        if _os.path.exists(datei):
            try:
                with open(datei) as f:
                    d = json.load(f)
                if ziel=="votes": _discord_votes=d
                else: _discord_punkte=d
            except Exception: pass

def discord_votes_speichern():
    import json
    for datei,daten in [(DISCORD_VOTE_DATEI,_discord_votes),("discord_punkte.json",_discord_punkte)]:
        try:
            with open(datei,"w") as f: json.dump(daten,f)
        except Exception: pass

def discord_vote_auswerten(signal_key, gewonnen):
    for uid,vote in _discord_votes.get(signal_key,{}).items():
        richtig = (vote=="ja" and gewonnen) or (vote=="nein" and not gewonnen)
        _discord_punkte.setdefault(uid,{"punkte":0,"gewinne":0,"verluste":0,"name":"?"})
        d = _discord_punkte[uid]
        if richtig: d["punkte"]+=10; d["gewinne"]+=1
        else:       d["punkte"]=max(0,d["punkte"]-3); d["verluste"]+=1
    discord_votes_speichern()

def sende_discord_rangliste():
    if not _discord_punkte: return
    sort = sorted(_discord_punkte.items(),key=lambda x:x[1].get("punkte",0),reverse=True)[:10]
    medals = ["🥇","🥈","🥉"]
    felder = [{"name":f"{medals[i] if i<3 else str(i+1)+'.'} {d.get('name','User')}",
               "value":f"{d['punkte']} Pkt | {d['gewinne']}✅ {d['verluste']}❌","inline":True}
              for i,(uid,d) in enumerate(sort)]
    send_discord_embed(DISCORD_WEBHOOK_BILANZ,{"title":"🏆 Discord Rangliste","color":0xFFD700,
        "fields":felder,"footer":{"text":f"BetlabLIVE • {heute()}"}})

def sende_einladungs_leaderboard():
    einladungen = _community_system.get("einladungen",{})
    if not einladungen: return
    sort    = sorted(einladungen.items(),key=lambda x:x[1].get("count",0),reverse=True)[:10]
    medals  = ["🥇","🥈","🥉"]
    zeilen  = [f"{medals[i] if i<3 else str(i+1)+'.'} <b>{d.get('name','?')}</b>: {d.get('count',0)}"
               for i,(uid,d) in enumerate(sort)]
    msg = "📊 <b>Einladungs-Leaderboard</b>\n━━━━━━━━━━━━━━━━━━━━\n"+"\n".join(zeilen)
    send_telegram(msg); send_telegram_gruppe(msg)

def sende_monatliche_challenge():
    monat = de_now().strftime("%Y-%m")
    if monat in _community_system.get("challenges",{}): return
    _community_system.setdefault("challenges",{})[monat] = {"teilnehmer":{},"aktiv":True}
    community_system_speichern()
    msg = (f"🏆 <b>Monatliche Challenge startet!</b>\n━━━━━━━━━━━━━━━━━━━━\n"
           f"📅 {de_now().strftime('%B %Y')}\n💰 Wer macht aus 10€ am meisten?\n"
           f"🎁 Preise: Premium, Analyse, VIP\n━━━━━━━━━━━━━━━━━━━━\n/challenge join")
    send_telegram(msg); send_telegram_gruppe(msg)

# ============================================================
#  KONFIDENZ-KALIBRIERUNG
# ============================================================

def kalibriere_konfidenz():
    with _tracker_lock:
        alle = [s for s in _signal_tracker.values()
                if s.get("status")=="ausgewertet" and s.get("konfidenz")]
    if len(alle)<40: return {}  # v59.17: von 100 auf 40 gesenkt (weniger Signale aktuell)
    nach_k = {}
    for s in alle:
        k = s.get("konfidenz",6)
        nach_k.setdefault(k,{"gewonnen":0,"total":0})
        nach_k[k]["total"] += 1
        if s.get("gewonnen"): nach_k[k]["gewonnen"] += 1
    meldungen = []
    for k,data in sorted(nach_k.items()):
        if data["total"]<5: continue  # v59.17: von 10 auf 5 gesenkt
        echte = round(data["gewonnen"]/data["total"]*100)
        erw   = k*10
        if abs(echte-erw)>=15:
            meldungen.append(f"{'⚠️' if echte<erw else '✅'} K{k}/10: erwartet {erw}%, echt {echte}%")
    if meldungen:
        send_telegram("🧠 <b>Konfidenz-Kalibrierung</b>\n━━━━━━━━━━━━━━━━━━━━\n"+"\n".join(meldungen))

# ============================================================
#  COMPOUND BANKROLL SIMULATION
# ============================================================

def simuliere_compound_bankroll(startkapital=None, wochen=20):
    if startkapital is None: startkapital = bankroll_laden()
    gw  = sum(statistik[t]["gewonnen"] for t in statistik)
    vl  = sum(statistik[t]["verloren"] for t in statistik)
    ges = gw+vl
    if ges<10: return {"ausreichend":False}
    p     = gw/ges; q_avg = 1.85
    kelly = max(0,min((p*q_avg-1)/(q_avg-1),0.05))
    tpw   = max(1,ges//max(1,de_now().isocalendar()[1]))
    br    = startkapital; verlauf = [round(br,2)]
    for _ in range(wochen):
        for _ in range(tpw):
            eins = br*kelly
            br   = round(br+eins*(q_avg-1) if p>0.5 else max(1,br-eins),2)
        verlauf.append(br)
    return {"ausreichend":True,"start":round(startkapital,2),"end":br,
            "rendite_pct":round((br-startkapital)/startkapital*100,1),
            "kelly_pct":round(kelly*100,1),"wochen":wochen,"verlauf":verlauf,
            "trefferquote":round(p*100,1)}

# ============================================================
#  SIGNAL-ARCHIV SUCHE
# ============================================================

def suche_signale(suchbegriff):
    sb = suchbegriff.lower().strip()
    with _tracker_lock:
        treffer = [s for s in _signal_tracker.values()
                   if any(sb in (s.get(k) or "").lower()
                          for k in ("home","away","competition","liga"))]
    if not treffer: return f"❌ Keine Signale: <b>{suchbegriff}</b>"
    sort = sorted(treffer,key=lambda x:x.get("signal_zeit",0),reverse=True)[:10]
    import datetime
    zeilen = [f"🔍 <b>'{suchbegriff}'</b> ({len(treffer)} gesamt)\n━━━━━━━━━━━━━━━━━━━━"]
    for s in sort:
        try: dt = datetime.datetime.fromtimestamp(s.get("signal_zeit",0)).strftime("%d.%m %H:%M")
        except: dt = "?"
        e = "✅" if s.get("gewonnen") else ("❌" if s.get("status")=="ausgewertet" else "⏳")
        zeilen.append(f"{e} {dt} | {s.get('home','?')} vs {s.get('away','?')} | {s.get('typ','?')}")
    return "\n".join(zeilen)

# ============================================================
#  WETTMARKT-EFFIZIENZ
# ============================================================

def berechne_markt_effizienz():
    with _tracker_lock:
        tipps = [s for s in _signal_tracker.values()
                 if s.get("typ") in ("value","arbitrage") and s.get("status")=="ausgewertet"]
    if len(tipps)<10: return {"ausreichend":False}
    kat = {"sehr_gut":[t for t in tipps if t.get("ev",0)>=0.15],
           "gut":     [t for t in tipps if 0.08<=t.get("ev",0)<0.15],
           "gering":  [t for t in tipps if t.get("ev",0)<0.08]}
    erg = {}
    for name,lst in kat.items():
        if not lst: continue
        gw  = sum(1 for t in lst if t.get("gewonnen"))
        roi = round(sum((t.get("quote",1)-1) if t.get("gewonnen") else -1 for t in lst)/len(lst)*100,1)
        erg[name] = {"pct":round(gw/len(lst)*100),"roi":roi,"tipps":len(lst)}
    return {"ausreichend":True,"kategorien":erg}

# ============================================================
#  PATTERN RECOGNITION
# ============================================================

_pattern_cache = {}
PATTERN_TTL    = 7200

def analysiere_team_muster(team_id, team_name):
    now = time.time()
    if team_id in _pattern_cache and now-_pattern_cache[team_id]["ts"]<PATTERN_TTL:
        return _pattern_cache[team_id]["data"]
    try:
        matches = _af_team_history(team_id,10)
        if len(matches)<5: return {}
        hz1,hz2 = [],[]
        for m in matches:
            svz = (m.get("scores") or {}).get("score","")
            shz = (m.get("scores") or {}).get("ht_score","")
            if not svz or not shz: continue
            hv,av = parse_score(svz); hh,ah = parse_score(shz)
            hz1.append(hh+ah); hz2.append((hv-hh)+(av-ah))
        a1 = round(sum(hz1)/len(hz1),2) if hz1 else 0
        a2 = round(sum(hz2)/len(hz2),2) if hz2 else 0
        result = {"team":team_name,"avg_hz1_tore":a1,"avg_hz2_tore":a2,
                  "hz2_staerker":a2>a1*1.3,"spiele":len(matches)}
        _pattern_cache[team_id] = {"data":result,"ts":now}
        return result
    except Exception as e: print(f"  [Pattern] {e}"); return {}

# ============================================================
#  EXCEL EXPORT
# ============================================================

def erstelle_excel_export():
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font
        wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Tipps"
        header = ["Datum","Zeit","Liga","Heim","Gast","Typ","Quote","Status","Ergebnis"]
        for i,h in enumerate(header,1):
            c = ws.cell(row=1,column=i,value=h)
            c.fill = PatternFill("solid",fgColor="1F3A5F"); c.font = Font(bold=True,color="FFFFFF")
        grenze = time.time()-30*24*3600
        with _tracker_lock:
            sigs = sorted([s for s in _signal_tracker.values() if s.get("signal_zeit",0)>=grenze],
                          key=lambda x:x.get("signal_zeit",0),reverse=True)
        for row,s in enumerate(sigs,2):
            import datetime as _dt
            try: d=_dt.datetime.fromtimestamp(s.get("signal_zeit",0)); dat=d.strftime("%d.%m.%Y"); uhr=d.strftime("%H:%M")
            except: dat=uhr="?"
            erg = "✅ GEWONNEN" if s.get("gewonnen") else ("❌ VERLOREN" if s.get("status")=="ausgewertet" else "⏳")
            for col,val in enumerate([dat,uhr,s.get("competition",s.get("liga","?")),
                    s.get("home","?"),s.get("away","?"),s.get("typ","?"),
                    s.get("quote",""),s.get("status",""),erg],1):
                ws.cell(row=row,column=col,value=val)
        for col in ws.columns: ws.column_dimensions[col[0].column_letter].width=16
        pfad = "/tmp/betlab_export.xlsx"; wb.save(pfad); return pfad
    except ImportError: print("  [Export] openpyxl fehlt"); return None
    except Exception as e: print(f"  [Export] {e}"); return None

# ============================================================
#  CHARTS & GRAFIKEN
# ============================================================

def erstelle_performance_chart():
    try:
        import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
        with _tracker_lock: alle = sorted(_signal_tracker.values(),key=lambda x:x.get("signal_zeit",0))
        qk = []; gw = 0
        for i,s in enumerate([a for a in alle if a.get("status")=="ausgewertet"]):
            gw += 1 if s.get("gewonnen") else 0; qk.append(round(gw/(i+1)*100,1))
        if len(qk)<3: return None
        fig,ax = plt.subplots(figsize=(10,4),facecolor="#0d1117"); ax.set_facecolor("#161b22")
        ax.plot(range(len(qk)),qk,color="#58a6ff",linewidth=2.5)
        ax.axhline(y=50,color="#f85149",linestyle="--",linewidth=1,alpha=0.7)
        ax.axhline(y=55,color="#3fb950",linestyle="--",linewidth=1,alpha=0.5)
        ax.fill_between(range(len(qk)),qk,50,where=[q>=50 for q in qk],alpha=0.15,color="#3fb950")
        ax.fill_between(range(len(qk)),qk,50,where=[q<50  for q in qk],alpha=0.15,color="#f85149")
        ax.set_title("BetlabLIVE Trefferquote",color="#e6edf3",fontsize=14)
        ax.tick_params(colors="#8b949e"); ax.spines[:].set_color("#30363d"); ax.set_ylim(0,100)
        pfad = "/tmp/betlab_chart.png"; plt.tight_layout(); plt.savefig(pfad,dpi=120,bbox_inches="tight",facecolor="#0d1117"); plt.close(); return pfad
    except Exception as e: print(f"  [Chart] {e}"); return None

def sende_chart_telegram(pfad):
    try:
        with open(pfad,"rb") as f:
            requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendPhoto",
                data={"chat_id":TELEGRAM_CHAT_ID,"caption":f"📊 Performance Chart – {heute()}"},
                files={"photo":f},timeout=20)
    except Exception as e: print(f"  [Chart] {e}")

def erstelle_daily_recap_grafik():
    try:
        import matplotlib; matplotlib.use("Agg"); import matplotlib.pyplot as plt
        from matplotlib.patches import FancyBboxPatch
        gw = sum(statistik[t]["gewonnen"] for t in statistik)
        vl = sum(statistik[t]["verloren"] for t in statistik)
        pct= round(gw/max(gw+vl,1)*100)
        fig,ax = plt.subplots(figsize=(9,5),facecolor="#0d1117"); ax.set_facecolor("#0d1117"); ax.axis("off")
        ax.add_patch(FancyBboxPatch((0.02,0.05),0.96,0.9,boxstyle="round,pad=0.02",
            facecolor="#161b22",edgecolor="#30363d",linewidth=2,transform=ax.transAxes))
        ax.text(0.5,0.88,"BetlabLIVE",ha="center",fontsize=22,fontweight="bold",color="#58a6ff",transform=ax.transAxes)
        ax.text(0.5,0.78,f"Daily Recap {de_now().strftime('%d.%m.%Y')}",ha="center",fontsize=12,color="#8b949e",transform=ax.transAxes)
        ax.text(0.2,0.58,str(gw),ha="center",fontsize=48,fontweight="bold",color="#3fb950",transform=ax.transAxes)
        ax.text(0.2,0.43,"Gewonnen",ha="center",fontsize=11,color="#8b949e",transform=ax.transAxes)
        ax.text(0.5,0.58,f"{pct}%",ha="center",fontsize=48,fontweight="bold",color="#58a6ff",transform=ax.transAxes)
        ax.text(0.5,0.43,"Trefferquote",ha="center",fontsize=11,color="#8b949e",transform=ax.transAxes)
        ax.text(0.8,0.58,str(vl),ha="center",fontsize=48,fontweight="bold",color="#f85149",transform=ax.transAxes)
        ax.text(0.8,0.43,"Verloren",ha="center",fontsize=11,color="#8b949e",transform=ax.transAxes)
        pfad = "/tmp/daily_recap.png"; plt.tight_layout(pad=0)
        plt.savefig(pfad,dpi=150,bbox_inches="tight",facecolor="#0d1117"); plt.close(); return pfad
    except Exception as e: print(f"  [Recap] {e}"); return None

def sende_daily_recap():
    pfad = erstelle_daily_recap_grafik()
    if not pfad: return
    gw = sum(statistik[t]["gewonnen"] for t in statistik)
    vl = sum(statistik[t]["verloren"] for t in statistik)
    pct= round(gw/max(gw+vl,1)*100)
    try:
        with open(pfad,"rb") as f:
            requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendPhoto",
                data={"chat_id":TELEGRAM_CHAT_ID,"caption":f"Daily Recap {de_now().strftime('%d.%m.%Y')} | {gw}G {vl}V {pct}%"},
                files={"photo":f},timeout=20)
    except Exception: pass

def erstelle_monatsbericht_pdf():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.colors import HexColor
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.units import cm
        lm   = de_now().replace(day=1) - timedelta(days=1)
        pfad = f"/tmp/betlab_{lm.strftime('%Y_%m')}.pdf"
        doc  = SimpleDocTemplate(pfad,pagesize=A4)
        stl  = getSampleStyleSheet(); stl["Title"].textColor = HexColor("#58a6ff")
        gw   = sum(statistik[t]["gewonnen"] for t in statistik)
        vl   = sum(statistik[t]["verloren"] for t in statistik)
        br   = bankroll_laden()
        pct  = round(gw/max(gw+vl,1)*100)
        doc.build([
            Paragraph(f"BetlabLIVE – {lm.strftime('%B %Y')}",stl["Title"]),
            Spacer(1,0.5*cm),
            Paragraph(f"Gewonnen: {gw} | Verloren: {vl} | Quote: {pct}%",stl["Normal"]),
            Paragraph(f"Bankroll: {br}€",stl["Normal"]),
        ])
        return pfad
    except ImportError: print("  [PDF] reportlab fehlt"); return None
    except Exception as e: print(f"  [PDF] {e}"); return None

def sende_pdf_telegram(pfad, monat):
    try:
        with open(pfad,"rb") as f:
            requests.post(f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument",
                data={"chat_id":TELEGRAM_CHAT_ID,"caption":f"📄 Monatsbericht {monat}"},
                files={"document":f},timeout=30)
    except Exception as e: print(f"  [PDF] {e}")

# ============================================================
#  BOTS: TIPP-DES-TAGES / ODDS-TRACKER / HEDGE-ALARM
#        BONUS-TRACKER / QUOTENVERGLEICH
# ============================================================

def bot_tipp_des_tages():
    print("[TippDesTages-Bot] Gestartet | Täglich 09:00 Uhr | Nur bei Konfidenz 10/10")
    import random; gesendet = set()
    while True:
        try:
            now = de_now(); key = now.strftime("%Y-%m-%d")
            if now.hour==9 and now.minute<5 and key not in gesendet:
                gesendet.add(key)
                fixtures = ls_get_fixtures(now.strftime("%Y-%m-%d"))
                top_alle = filtere_top_spiele(fixtures)
                # v59.30 FIX: Nur echte Prematch-Spiele (noch nicht begonnen), die HEUTE
                # stattfinden – ls_get_fixtures liefert zwar schon nur den heutigen Tag, aber
                # ohne Status-Filter könnte theoretisch ein bereits laufendes frühes Spiel
                # mitgenommen werden. "NS" (Not Started) stellt sicher, dass es wirklich noch
                # vor Anpfiff ist.
                top = [s for s in top_alle if s.get("status") in ("NS","TBD")]
                if not top: time.sleep(60); continue
                bester = None; beste_k = 0
                for spiel in top[:10]:
                    home   = spiel.get("home_name") or (spiel.get("home") or {}).get("name","?")
                    away   = spiel.get("away_name") or (spiel.get("away") or {}).get("name","?")
                    liga   = spiel.get("competition",{}).get("name","?")
                    anstoß = spiel.get("time","?")
                    r      = claude_prematch_analyse(home,away,liga,anstoß,[])
                    if r and r.get("konfidenz",0)>beste_k:
                        beste_k=r["konfidenz"]; bester={**r,"home":home,"away":away,"liga":liga,"anstoß":anstoß,"id":spiel.get("id","")}
                    time.sleep(1)
                # v59.30 FIX: Nur noch senden, wenn die Konfidenz wirklich 10/10 erreicht –
                # ansonsten lieber an dem Tag GAR KEINEN Tipp des Tages senden (Qualität statt
                # Quantität), statt einen schwächeren Tipp als "Tipp des Tages" zu labeln.
                if not bester or beste_k < 10:
                    print(f"  [TippDesTages] Kein Spiel mit Konfidenz 10/10 gefunden (beste: {beste_k}/10) – heute kein Tipp des Tages")
                    time.sleep(60); continue
                ke = konfidenz_emoji(bester["konfidenz"])
                # v59.7: Quote + Einsatz-Empfehlung ergänzt
                bester_fixture_id = bester.get("id","")
                quote_tdt         = get_quote(bester_fixture_id,"prematch",nur_live=False) if bester_fixture_id else None  # v59.24: Pre-Match-Quote hier korrekt/gewollt
                einsatz_tdt       = kelly_einsatz_bankroll(quote_tdt,"ecken") if quote_tdt else EINSATZ
                ql_tdt            = f"\n💶 Quote: <b>{quote_tdt}</b> | 💰 Einsatz: <b>{einsatz_tdt}€</b>" if quote_tdt else f"\n💰 Einsatz: <b>{einsatz_tdt}€</b>"
                msg = (f"⭐ <b>Tipp des Tages – {now.strftime('%d.%m.%Y')}</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🏆 {bester['liga']}\n"
                       f"📌 {bester['home']} vs {bester['away']}\n"
                       f"🕐 Anstoß: <b>{bester['anstoß']} Uhr</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n🎯 Tipp: <b>{bester['tipp']}</b>{ql_tdt}\n"
                       f"{ke} Konfidenz: <b>{bester['konfidenz']}/10</b>\n"
                       f"📊 {bester['analyse']}\n━━━━━━━━━━━━━━━━━━━━\n⚠️ 18+ | Verantwortungsvoll spielen")
                send_telegram(msg); send_telegram_gruppe(msg)
                send_discord_embed(DISCORD_WEBHOOK_TAGESTIPP,{  # v59.29 FIX: eigener Kanal statt Auswertungs-/Bilanz-Kanal
                    "title":f"⭐ Tipp des Tages – {now.strftime('%d.%m.%Y')}","color":0xF1C40F,
                    "fields":[
                        {"name":"🏆 Liga","value":bester["liga"],"inline":True},
                        {"name":"⚽ Spiel","value":f"{bester['home']} vs {bester['away']}","inline":False},
                        {"name":"🎯 Tipp","value":f"**{bester['tipp']}**","inline":True},
                        {"name":ke+" Konfidenz","value":f"**{bester['konfidenz']}/10**","inline":True},
                        {"name":"💶 Quote","value": (f"**{quote_tdt}**" if quote_tdt else "⚠️ Keine zuverlässige Quote"),"inline":True},
                        {"name":"💰 Einsatz-Empfehlung","value":einsatz_empfehlung_text(einsatz_tdt),"inline":True},
                        {"name":"📊 Analyse","value":bester["analyse"],"inline":False},
                    ],"footer":{"text":f"Tipp des Tages • {heute()}"}})
                print(f"  [TippDesTages] ✅ {bester['home']} vs {bester['away']}")
        except Exception as e: print(f"  [TippDesTages] Fehler: {e}")
        time.sleep(60)

_odds_history_tracker = {}
notified_odds_drop    = set()

def bot_odds_tracker():
    """
    v59.1: Beobachtet Quotenbewegungen jetzt über API-Football-Odds pro
    laufendem Spiel, statt wie vorher über einen Bulk-Abruf bei the-odds-api.com.
    """
    print("[Odds-Tracker] Gestartet | Quoten-Bewegungen überwachen")
    while True:
        try:
            matches = get_live_matches()
            now = time.time()
            for m in matches:
                fixture_id = str(m.get("id"))
                h_t = m.get("home",{}).get("name","?"); a_t = m.get("away",{}).get("name","?")
                key = fixture_id
                werte = _af_odds_werte(af_get_odds(fixture_id), ("Goals Over/Under","Over/Under","Match Winner"))
                if not werte: continue
                bm_q = {}
                for w in werte:
                    bm_q.setdefault(w["label"],[]).append({"q":w["odd"],"bm":w["bookmaker"]})
                if key not in _odds_history_tracker:
                    _odds_history_tracker[key]={"ts":now,"q":bm_q}; continue
                alt = _odds_history_tracker[key]
                if now-alt["ts"]<8*60: continue
                for k,ql in bm_q.items():
                    if k not in alt["q"]: continue
                    qn = sum(v["q"] for v in ql)/max(len(ql),1)
                    qa = sum(v["q"] for v in alt["q"][k])/max(len(alt["q"][k]),1)
                    if qa<=1.1 or qn<=1.1: continue
                    bew = round((qn-qa)/qa*100,1)
                    if abs(bew)>=8 and len(ql)>=3:
                        ar_k = f"{key}_{k}"
                        if ar_k in notified_odds_drop: continue
                        notified_odds_drop.add(ar_k)
                        ri = "📉 gefallen" if bew<0 else "📈 gestiegen"
                        msg = (f"📉 <b>Quote stark verändert!</b>\n📌 {h_t} vs {a_t}\n"
                               f"🎯 {k}\n📉 {qa:.2f}→<b>{qn:.2f}</b> ({bew:+.1f}%) {ri}\n🕐 {jetzt()} Uhr")
                        send_telegram(msg)
                        send_discord_embed(DISCORD_WEBHOOK_VALUE,{"title":"📉 Quote verändert","color":0xE74C3C,
                            "fields":[{"name":"⚽","value":f"{h_t} vs {a_t}","inline":True},
                                      {"name":"🎯","value":f"**{k}**","inline":True},
                                      {"name":"📉","value":f"{qa:.2f}→{qn:.2f} ({bew:+.1f}%)","inline":True}],
                            "footer":{"text":f"Odds-Tracker • {heute()} {jetzt()}"}})
                _odds_history_tracker[key]={"ts":now,"q":bm_q}
            bot_fehler_reset("Odds-Tracker")
        except Exception as e: bot_fehler_melden("Odds-Tracker",e)
        time.sleep(5*60)

def bot_hedge_alarm():
    print("[Hedge-Alarm-Bot] Gestartet | Hedging-Empfehlungen")
    while True:
        try:
            offene = tracker_get_offene()
            if not offene: time.sleep(5*60); continue
            for _,sig in offene:
                home   = sig.get("home",""); away  = sig.get("away","")
                orig_q = sig.get("quote")
                if not orig_q or orig_q<=1.0: continue
                details   = get_quote_details(sig.get("match_id"))  # v59.1: jetzt fixture-basiert
                aktuelle_q = details.get("avg_quote",0)
                if not aktuelle_q or aktuelle_q<=1.0: continue
                anstieg = round((aktuelle_q-orig_q)/orig_q*100,1)
                if anstieg>=25:
                    msg = (f"🛡️ <b>Hedge-Empfehlung!</b>\n📌 {home} vs {away}\n"
                           f"💶 Orig: <b>{orig_q}</b> → Aktuell: <b>{aktuelle_q}</b> (+{anstieg}%)\n"
                           f"🛡️ Quote stark gestiegen → Hedge erwägen!\n🕐 {jetzt()} Uhr")
                    send_telegram(msg)
            bot_fehler_reset("Hedge-Alarm-Bot")
        except Exception as e: bot_fehler_melden("Hedge-Alarm-Bot",e)
        time.sleep(10*60)

BOOKIE_AKTIONEN = {
    "Leonbet":  "Keine Wettsteuer, bis zu 5% Cashback",
    "Bet365":   "Regelmäßige Freiwetten für aktive Kunden",
    "Bwin":     "Quoten-Boosts auf ausgewählte Spiele",
    "Unibet":   "Cashback auf erste verlorene Wette",
    "Pinnacle": "Kein Bonus – aber beste Quoten weltweit",
    "Winamax":  "97% Quotenschlüssel, steuerfrei in DE",
}

def bot_bonus_tracker():
    print("[Bonus-Tracker] Gestartet | Erinnerungen Mo/Mi/Fr")
    gesendet = set()
    while True:
        try:
            now = de_now(); key = now.strftime("%Y-%m-%d")
            if now.weekday() in (0,2,4) and now.hour==9 and now.minute<5 and key not in gesendet:
                gesendet.add(key)
                zeilen = [f"📌 <b>{bm}</b>\n   {info}" for bm,info in BOOKIE_AKTIONEN.items()]
                msg = (f"🎁 <b>Bookmaker Aktionen – {now.strftime('%d.%m.%Y')}</b>\n"
                       f"━━━━━━━━━━━━━━━━━━━━\n"+"\n".join(zeilen)+
                       f"\n━━━━━━━━━━━━━━━━━━━━\n⚠️ 18+ | Verantwortungsvoll spielen")
                send_telegram(msg)
            bot_fehler_reset("Bonus-Tracker")
        except Exception as e: bot_fehler_melden("Bonus-Tracker",e)
        time.sleep(60)

def bot_quotenvergleich():
    """
    v59.1: Nutzt jetzt die Fixtures des Tages (bereits über API-Football
    geladen wie beim PreMatch-Bot) und fragt für die Top-Ligen-Spiele die
    Odds einzeln über /odds ab, statt wie vorher einen Bulk-Abruf bei
    the-odds-api.com zu machen.
    """
    print("[Quotenvergleich-Bot] Gestartet | Täglich 08:30 Uhr")
    gesendet = set()
    while True:
        try:
            now = de_now(); key = now.strftime("%Y-%m-%d")
            if now.hour==8 and 30<=now.minute<35 and key not in gesendet:
                gesendet.add(key)
                fixtures   = ls_get_fixtures(now.strftime("%Y-%m-%d"))
                top_spiele = filtere_top_spiele(fixtures)[:8]
                top = []
                for spiel in top_spiele:
                    fixture_id = spiel.get("id","")
                    h_t = spiel.get("home_name") or (spiel.get("home") or {}).get("name","?")
                    a_t = spiel.get("away_name") or (spiel.get("away") or {}).get("name","?")
                    werte = _af_odds_werte(af_get_odds(fixture_id), ("Match Winner","Goals Over/Under","Over/Under"))
                    if not werte: continue
                    bester = max(werte,key=lambda w:w["odd"])
                    if bester["odd"]>=1.8:
                        top.append({"spiel":f"{h_t} vs {a_t}","q":bester["odd"],"bm":bester["bookmaker"],"markt":bester["label"]})
                if top:
                    top.sort(key=lambda x:x["q"],reverse=True)
                    zeilen = "\n".join([f"⭐ {v['spiel']}\n   💶 {v['q']} @ {v['bm']} ({v['markt']})" for v in top[:5]])
                    send_telegram(f"💹 <b>Beste Quoten heute – {now.strftime('%d.%m.%Y')}</b>\n"
                                  f"━━━━━━━━━━━━━━━━━━━━\n{zeilen}\n━━━━━━━━━━━━━━━━━━━━\n🕐 {jetzt()} Uhr")
            bot_fehler_reset("Quotenvergleich-Bot")
        except Exception as e: bot_fehler_melden("Quotenvergleich-Bot",e)
        time.sleep(60)


# ============================================================
#  WATCHDOG
# ============================================================

_bot_targets = {}

def bot_watchdog():
    print("[Watchdog] Gestartet")
    time.sleep(30)
    while True:
        try:
            aktive = {t.name:t for t in threading.enumerate()}
            for name,target in _bot_targets.items():
                if name not in aktive or not aktive[name].is_alive():
                    print(f"  [Watchdog] ⚠️ {name} ist tot – starte neu!")
                    send_telegram(f"⚠️ <b>Watchdog Alert!</b>\nBot <b>{name}</b> neu gestartet.\n🕐 {jetzt()} Uhr")
                    t = threading.Thread(target=target,daemon=True,name=name)
                    t.start()
        except Exception as e:
            print(f"  [Watchdog] Fehler: {e}")
        time.sleep(60)

def cleanup_memory():
    """
    Bereinigt unbegrenzt wachsende Dicts um Memory Leaks zu vermeiden.
    Läuft automatisch täglich um 03:00 Uhr.
    """
    global bereits_getippt, aktive_tipps
    vorher_bt = len(bereits_getippt)
    vorher_at = len(aktive_tipps)
    bereits_getippt = {}
    aktive_tipps    = {}
    print(f"  [Cleanup] bereits_getippt: {vorher_bt}→0 | aktive_tipps: {vorher_at}→0")

def bot_startup_alarm():
    start_str = de_now().strftime("%d.%m.%Y %H:%M")
    print(f"  [Startup] BetlabLIVE {BOT_VERSION} gestartet um {start_str}")
    key_status = "✅ API-Keys OK" if startup_ok else "⚠️ API-Keys prüfen!"
    reaction_status = "✅ Discord-Reaktionen aktiv" if DISCORD_BOT_TOKEN else "⚠️ DISCORD_BOT_TOKEN fehlt – keine ✅/❌-Reaktionen"
    send_telegram(
        f"🚀 <b>BetlabLIVE {BOT_VERSION} gestartet!</b>\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"{key_status}\n{reaction_status}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"📦 Build: <b>{BOT_VERSION}</b>\n"
        f"📝 {BOT_VERSION_INFO}\n"
        f"━━━━━━━━━━━━━━━━━━━━\n"
        f"🕐 {jetzt()} Uhr"
    )

# ============================================================
#  STARTUP
# ============================================================

if __name__ == "__main__":
    print("=" * 55)
    print(f"  ⚽ BETLABLIVE {BOT_VERSION} – {BOT_VERSION_INFO}")
    print(f"  ⚽ BUILD-CHECK: Wenn du diese Zeile im Railway-Log NICHT siehst, läuft ein alter Stand!")
    print("  Auf Basis von v60 (Rückumstellung auf livescore-api.com, Quoten deaktiviert)")
    print("=" * 55 + "\n")

    # ── API-Key Validierung vor Start ─────────────────────
    startup_ok = True
    required_vars = {
        "LIVESCORE_API_KEY":   API_KEY,
        "LIVESCORE_API_SECRET":API_SECRET,
        "TELEGRAM_TOKEN":  TELEGRAM_BOT_TOKEN,
        "TELEGRAM_CHAT_ID":TELEGRAM_CHAT_ID,
    }
    optional_vars = {
        "ODDS_API_KEY":     ODDS_API_KEY,
        "FOOTBALLDATA_KEY": FOOTBALLDATA_KEY,
        "ANTHROPIC_KEY":    ANTHROPIC_API_KEY,
        "GITHUB_TOKEN":     GITHUB_TOKEN,
        "DISCORD_BOT_TOKEN":DISCORD_BOT_TOKEN,   # v59: fuer die ✅/❌-Reaktionen
        "DISCORD_WEBHOOK_TAGESTIPP":DISCORD_WEBHOOK_TAGESTIPP,  # v59.29 NEU
    }
    print("[Startup] Prüfe Konfiguration...")
    for name, val in required_vars.items():
        if not val or val.strip() == "":
            print(f"  ❌ FEHLT: {name} – Pflicht-Variable nicht gesetzt!")
            startup_ok = False
        else:
            print(f"  ✅ {name}: gesetzt")
    for name, val in optional_vars.items():
        status = "✅" if (val and val.strip()) else "⚠️  (deaktiviert)"
        print(f"  {status} {name}")
    if not DISCORD_BOT_TOKEN:
        print("  [Startup] Hinweis: DISCORD_BOT_TOKEN nicht gesetzt -> Auswertungs-Reaktionen (✅/❌) bleiben aus.")
    if not startup_ok:
        print("[Startup] Pflicht-Variablen fehlen – Bot startet, aber API-Calls werden fehlschlagen!")
        print("[Startup] Bitte in Railway die fehlenden Environment Variables setzen.")
    else:
        print("[Startup] Alle Pflicht-Variablen gesetzt")
    print(f"[Startup] API-Tageslimit: {API_DAILY_LIMIT:,} (Hard-Stop bei {API_DAILY_HARD_STOP:,})")
    # ─────────────────────────────────────────────────────────────

    print("[Startup] GitHub Restore...")
    github_restore()

    statistik_laden()
    beobachtete_spiele_laden()
    tracker_laden()
    prematch_tracker_laden()
    notified_sets_laden()
    whitelist_laden()
    admins_laden()
    bekannte_user_laden()
    manuell_tipps_laden()
    telegram_filter_laden()
    liga_auswertung_laden()

    try:
        dynamische_filter_laden()
    except Exception as e:
        print(f"  [Startup] dynamische_filter_laden: {e}")

    try: ab_test_laden();            print("  [Startup] AB-Test geladen")
    except Exception as e: print(f"  [Startup] ab_test_laden: {e}")
    try: vk_laden();                 print("  [Startup] Virtuelle Konten geladen")
    except Exception as e: print(f"  [Startup] vk_laden: {e}")
    try: rang_laden();               print("  [Startup] Rang-System geladen")
    except Exception as e: print(f"  [Startup] rang_laden: {e}")
    try: community_system_laden();   print("  [Startup] Community geladen")
    except Exception as e: print(f"  [Startup] community_system_laden: {e}")
    try: discord_votes_laden();      print("  [Startup] Discord Votes geladen")
    except Exception as e: print(f"  [Startup] discord_votes_laden: {e}")

    # ── Bot-Definitionen
    bot_definitionen = [
        ("Ecken-Bot",        bot_ecken),
        ("Torwart-Bot",      bot_torwart),
        ("Druck-Bot",        bot_druck),
        ("Comeback-Bot",     bot_comeback),
        ("Torflut-Bot",      bot_torflut),
        ("Tore-Bot",         bot_tore_analyse),
        ("PreMatch-Bot",     bot_prematch),
        ("PreMatch-Auswertung-Bot", bot_prematch_auswertung),  # v60.2 NEU
        ("Telegram-Bot",     bot_telegram_befehle),
        ("Erinnerungs-Bot",  bot_prematch_erinnerung),
        ("Backup-Bot",       bot_github_backup),
        ("Auswertung-Bot",   bot_auswertung_und_berichte),
        ("Nachschau-Bot",    bot_nachschau),
        # v60.2: auf Wunsch wieder reaktiviert. WICHTIG: laufen technisch wieder, finden
        # aber praktisch nichts, solange livescore-api.com keine Über/Unter- oder
        # Mehr-Buchmacher-Quoten liefert (siehe Kommentar bei af_get_odds() weiter oben) –
        # beide Bots brauchen echte Linien-Quoten für ihre Berechnung.
        ("Value-Bot",        bot_value_bet),
        ("xG-Bot",           bot_xg),
        ("Arbitrage-Bot",    bot_arbitrage),
        ("EarlyGoal-Bot",    bot_early_goal),
        ("RotkarteEcken-Bot",bot_rotkarte_ecken),
        ("Anomalie-Bot",     bot_anomalie_erkennung),
        # v59.16: Deaktiviert auf Wunsch – bewirbt/beobachtet Buchmacher-Quoten ohne
        # direkten Mehrwert für den aktuellen Fokus. Funktion bleibt im Code erhalten,
        # falls sie später wieder aktiviert werden soll.
        # ("Sharp-Money-Bot",  bot_sharp_money),
        ("HZ2-Tore-Bot",     bot_hz2_tore),
        ("CornerRush-Bot",   bot_corner_rush),
        ("Morgen-Bot",       bot_morgen_uebersicht),
        ("Selbstlern-Bot",   bot_selbstlernend),
        ("Wetter-Bot",       bot_wetter_tipp),
        ("TippDesTages-Bot",  bot_tipp_des_tages),
        # v59.16: Deaktiviert auf Wunsch – bewirbt/beobachtet Buchmacher-Quoten ohne
        # direkten Mehrwert für den aktuellen Fokus. Funktion bleibt im Code erhalten,
        # falls sie später wieder aktiviert werden soll.
        # ("Odds-Tracker",      bot_odds_tracker),
        # v59.16: Deaktiviert auf Wunsch – bewirbt/beobachtet Buchmacher-Quoten ohne
        # direkten Mehrwert für den aktuellen Fokus. Funktion bleibt im Code erhalten,
        # falls sie später wieder aktiviert werden soll.
        # ("Hedge-Alarm-Bot",   bot_hedge_alarm),
        # v59.14: Bonus-Tracker deaktiviert – bewarb Buchmacher ohne eigene Affiliate-Links,
        # kein Nutzen für BettingLab. Funktion bleibt im Code, falls sie später mit echten
        # Affiliate-Links wieder aktiviert werden soll.
        # ("Bonus-Tracker",     bot_bonus_tracker),
        # v59.16: Deaktiviert auf Wunsch – bewirbt/beobachtet Buchmacher-Quoten ohne
        # direkten Mehrwert für den aktuellen Fokus. Funktion bleibt im Code erhalten,
        # falls sie später wieder aktiviert werden soll.
        # ("Quotenvergleich-Bot",bot_quotenvergleich),
        ("Gruppen-Bot",      bot_telegram_gruppe),
    ]

    dashboard = threading.Thread(target=bot_web_dashboard,daemon=True,name="Dashboard")
    dashboard.start()
    time.sleep(2)
    print("  [Startup] Dashboard auf Port 8080 bereit ✅ (/health, /api/stats, /radar)")

    for name,target in bot_definitionen:
        _bot_targets[name] = target

    try:
        bot_startup_alarm()
    except Exception as e:
        print(f"  [Startup] Alarm: {e}")

    threads = []
    for name,target in bot_definitionen:
        t = threading.Thread(target=target,daemon=True,name=name)
        threads.append(t)
        t.start()
        time.sleep(0.3)

    watchdog = threading.Thread(target=bot_watchdog,daemon=True,name="Watchdog")
    watchdog.start()

    print(f"\n✅ Alle {len(bot_definitionen)} Bots gestartet!\n")
    while True:
        time.sleep(60)
